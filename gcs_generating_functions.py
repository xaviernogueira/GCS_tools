import openpyxl as xl
import arcpy
from arcpy import env
from arcpy.sa import *
from arcpy.da import *
import os
from os import listdir
import DEM_detrending_functions
from DEM_detrending_functions import *
import create_centerline_GUI
from create_centerline_GUI import *
import create_station_lines
from create_station_lines import create_station_lines_function
from os.path import isfile, join
from GCS_analysis import *
import classify_landforms_GUI
from classify_landforms_GUI import *
from openpyxl import Workbook
from openpyxl import load_workbook
from matplotlib import pyplot as plt
import numpy as np
import csv
import pandas
import string
import scipy
from scipy import ndimage
from scipy.ndimage import gaussian_filter1d
from scipy.interpolate import UnivariateSpline
import descriptive_statistics_functions as stats


def detrend_to_wetted_poly(detrended_dem, out_folder, raster_units, max_stage=[], step=1):
    '''This function takes the detrended DEM and outputs stage polygons at param_list[0]=max stage, and param_list=[1]=stage interval (INTEGER)
        and makes smoothed, donut free, wetted polygons for each stage appending stage to the shapefile name. Each polygon is used to create a centerline
        from which the width and detrended Z analysis is completed. Raster units takes strings m or ft'''

    if raster_units == "m":
        spur_length = 10
    else:
        spur_length = 30

    if not os.path.exists(out_folder):
        os.makedirs(out_folder)

    if not os.path.exists(out_folder + "\\wetted_polygons"):
        os.makedirs(out_folder + "\\wetted_polygons")

    lines_location = out_folder + '\\analysis_centerline_and_XS'
    if not os.path.exists(lines_location):
        os.makedirs(lines_location)

    wetted_folder = out_folder + "\\wetted_polygons"

    try:
        # We convert raster units into feet for this stage of the analysis.
        # The Int function rounds down, but since a 1.3ft elevation would only be wetted by a 2ft flow, we add 1 to the detrended_int_ras
        detrend_ras = Raster(detrended_dem)
        if raster_units == "m":
            detrended_ras = float(detrend_ras * 3.28082)
        detrend_ras_int = Int(detrend_ras) + 1
        detrend_ras_int.save(out_folder + "\\dtrnd_int_ft.tif")
        print("Stage range is: " + str(range(max_stage[0])))

        flood_stage_ras = Raster(Con(detrend_ras_int <= max_stage[0], detrend_ras_int))
        flood_stage_poly = arcpy.RasterToPolygon_conversion(flood_stage_ras,
                                                            wetted_folder + ("\\flood_stage_poly_%sft" % max_stage[0]),
                                                            simplify=False, raster_field="Value",
                                                            create_multipart_features=True)
        flood_stage_poly = arcpy.Dissolve_management(flood_stage_poly, wetted_folder + ("\\flood_stage_poly_dissolved_%sft" % max_stage[0]), dissolve_field="gridcode", multi_part=True)
        print("Flood stage polygon made with max stage of %sft and is stored: %s" % (max_stage[0], flood_stage_poly))

        for i in range(0, max_stage[0] + 1, int(step)):
            print(range(max_stage[0] + 1))
            query = ('"gridcode" <= %d' % i)
            print(query)
            stage_poly = arcpy.SelectLayerByAttribute_management(flood_stage_poly, selection_type="NEW_SELECTION",
                                                                 where_clause=query)
            stage_poly = arcpy.CopyFeatures_management(stage_poly,
                                                       out_feature_class=wetted_folder + ("\\flood_stage_poly_%sft" % i))
            arcpy.AddField_management(stage_poly, "null_field", "Short")
            stage_poly_no_donuts = arcpy.Union_analysis(stage_poly, wetted_folder + (
                        "\\flood_stage_poly_%sft_no_donuts_predissolve" % i), gaps="NO_GAPS")
            stage_poly_dissolve = arcpy.Dissolve_management(stage_poly,
                                                            wetted_folder + ("\\flood_stage_poly_dissolved_%sft" % i),
                                                            dissolve_field="null_field", multi_part=True)
            stage_poly_dissolve_no_donuts = arcpy.Dissolve_management(stage_poly_no_donuts, wetted_folder + (
                        "\\flood_stage_poly_%sft_no_donuts" % i), dissolve_field="null_field", multi_part=True)

            if raster_units == "m":
                stage_poly_dissolve_no_donuts = arcpy.SmoothPolygon_cartography(stage_poly_dissolve_no_donuts, wetted_folder + (
                        "\\smooth_stage_poly_%sft_donuts" % i), "PAEK", 50)
            else:
                stage_poly_dissolve_no_donuts = arcpy.SmoothPolygon_cartography(stage_poly_dissolve_no_donuts, wetted_folder + ("\\smooth_stage_poly_%sft_donuts" % i), "PAEK", 164)

            stage_poly_dissolve_no_donuts = arcpy.Union_analysis(stage_poly_dissolve_no_donuts, wetted_folder + ("\\smooth_stage_poly_%sft_no_dissolve" % i), gaps="NO_GAPS")
            stage_poly_dissolve_no_donuts = arcpy.Dissolve_management(stage_poly_dissolve_no_donuts, wetted_folder + ("\\smooth_stage_poly_%sft" % i), dissolve_field="null_field", multi_part=True)
            flood_stage_poly = arcpy.SelectLayerByAttribute_management(flood_stage_poly,
                                                                       selection_type="CLEAR_SELECTION")

            # Create folder for centerline and cross sections, Make centerline based on smoothed, donut-less wetted polygons and delete spurs
            print("Creating centerline for %sft stage" % i)
            centerline = arcpy.PolygonToCenterline_topographic(stage_poly_dissolve_no_donuts, lines_location + ("\\stage_centerline_%sft" % i))
            print('Removing spurs smaller than % s units...' % str(spur_length))
            # measure lengths
            arcpy.AddGeometryAttributes_management(centerline, 'LENGTH')
            spurs = arcpy.SelectLayerByAttribute_management(centerline, where_clause=('LENGTH < %s' % str(spur_length)), selection_type="NEW_SELECTION")
            if int(arcpy.GetCount_management(spurs).getOutput(0)) > 0:
                arcpy.DeleteFeatures_management(spurs)
            arcpy.SelectLayerByAttribute_management(centerline, selection_type="CLEAR_SELECTION")

            print("Stage %s dissolved and non-dissolved polygons in %s" % (i, out_folder))

            print("Deleting unnecessary files")
            try:
                if os.path.exists(out_folder + ("\\smooth_stage_poly_%sft_donuts.shp" % i)):
                    os.remove(out_folder + ("\\smooth_stage_poly_%sft_donuts.shp" % i))
                if os.path.exists(out_folder + ("\\smooth_stage_poly_%sft_no_dissolve" % i)):
                    os.remove(out_folder + ("\\smooth_stage_poly_%sft_no_dissolve" % i))
                if os.path.exists(out_folder + ("\\flood_stage_poly_%sft_no_donuts" % i)):
                    os.remove(out_folder + ("\\flood_stage_poly_%sft_no_donuts" % i))
                if os.path.exists(out_folder + ("\\flood_stage_poly_%sft_no_donuts_predissolve" % i)):
                    os.remove(out_folder + ("\\flood_stage_poly_%sft_no_donuts_predissolve" % i))
            except:
                print("Some files not deleted")

        print("wetted polygons created")

    except arcpy.ExecuteError:
        print(arcpy.GetMessages())


def export_to_gcs_ready(out_folder, list_of_error_locations=[]):
    '''Export location, width, and mean depth as tables excluding damaged cross sections specified as a list.
    Table outputs ready for GCS_analysis.py analysis in a new folder'''
    table_location = out_folder + '\\gcs_ready_tables'
    if not os.path.exists(table_location):
        os.makedirs(table_location)

    analysis_folder = out_folder + "\\analysis_shapefiles"
    files_in_folder = [f for f in listdir(analysis_folder) if isfile(join(analysis_folder, f))]
    width_rectangles = [i for i in files_in_folder if i[0] == "w"]
    width_rectangles = [i for i in width_rectangles if i[-4:] == ".shp" and i[-8:-6] != '_0']
    print("Width rectanles found: %s" % width_rectangles)

    list_of_csv_tables = []

    # Delete bad locations and convert data into a dbf table
    for file in width_rectangles:
        file_location = (analysis_folder + "\\%s" % file)
        if len(list_of_error_locations) > 0:
            print("Erroneous cross sections identified...")
            for value in list_of_error_locations:
                query = ('"LOCATION" = %d' % value)
                arcpy.SelectLayerByAttribute_management(file_location, selection_type="ADD_TO_SELECTION", where_clause=query)
            arcpy.DeleteFeatures_management(file_location)
            arcpy.SelectLayerByAttribute_management(file_location, selection_type="CLEAR_SELECTION")
        analysis_table = arcpy.TableToTable_conversion(file_location, out_path=table_location, out_name=("WD_analysis_table_%s.dbf" % file[-8:-4]))
        analysis_xlsx = arcpy.TableToExcel_conversion(Input_Table=file_location, Output_Excel_File=table_location + ("\\WD_analysis_table_%s.xlsx" % file[-8:-4]))

        #Convert xls to csv and change field headers to match GCS_analysis functions
        wb = xl.load_workbook(str(analysis_xlsx))
        ws = wb.active
        if file[-8] == "_":
            csv_file = table_location + ("\\%s_WD_analysis_table.csv" % file[-7:-4])
        else:
            csv_file = table_location + ("\\%s_WD_analysis_table.csv" % file[-8:-4])

        if ws["H1"].value == "Z" and ws["F1"].value == "W" and ws["C1"].value == "dist_down":
            print('CSV already formatted...')
            wb.save(str(analysis_xlsx))
        elif ws["H1"].value == "MEAN" and ws["C1"].value == "LOCATION" and ws["F1"].value == "Width":
            ws["H1"].value = "Z"
            ws["F1"].value = "W"
            ws["C1"].value = "dist_down"
            wb.save(str(analysis_xlsx))
        else:
            print("ERROR! CELL HEADERS IN UNEXPECTED POSITIONS @ %s" % (table_location + ("\\WD_analysis_table_%s.xlsx" % file[-8:-4])))
            return

        with open(csv_file, 'w', newline="") as f:
            col = csv.writer(f)
            for row in ws.rows:
                col.writerow([cell.value for cell in row])
        list_of_csv_tables.append(csv_file)

    print("DBF and CSV tables of width/Z analysis at %s" % table_location)
    print("List of csv tables: %s" % list_of_csv_tables)

    return [list_of_csv_tables, width_rectangles]

def GCS_plotter(table_directory):
    '''Take the, Z, W, and landforms/Ws*Zs and plot vs dist_down using pyplot and open xl'''
    list_of_csv_tables = [f for f in listdir(table_location) if isfile(join(table_location, f))]
    list_of_csv_tables = [f for f in list_of_csv_tables if f[-4:] == ".csv"]
    print(list_of_csv_tables)

    # Set up figure output folder
    figure_output = table_directory + "\\GCS_figures"
    if not os.path.exists(figure_output):
        os.makedirs(figure_output)

    numpy_columns = ['dist_down', 'W_s', 'Z_s', 'W_s_Z_s']

    for table in list_of_csv_tables:
        if table[1] == 'f':
            stage = table[0]
        elif table[2] == 'f':
            stage = table[:2]
        elif table[1] == 'p':
            stage = round(float('%s.%s' % (table[0], table[2]), 1))
        elif table[2] == 'p':
            stage = round(float('%s.%s' % (table[0:2], table[3]), 1))

        table_df = pandas.read_csv(table_location + "\\%s" % table,na_values=[-9999])
        print(table_df)
        table_df["code"].fillna(0, inplace=True)

        zs_plot_name = ("%s_Zs_plot" % stage)
        ws_plot_name = ("%s_Ws_plot" % stage)
        zs_ws_plot_name = ("%s_ZsWs_plot" % stage)

        wide_bar_df = table_df.loc[table_df['code'] == 1, ['dist_down', 'W_s', 'Z_s', 'W_s_Z_s']]  # Landform type data frames
        nozzle_df = table_df.loc[table_df['code'] == 2, ['dist_down', 'W_s', 'Z_s', 'W_s_Z_s']]
        normal_df = table_df.loc[table_df['code'] == 0, ['dist_down', 'W_s', 'Z_s', 'W_s_Z_s']]
        pool_df = table_df.loc[table_df['code'] == -1, ['dist_down', 'W_s', 'Z_s', 'W_s_Z_s']]
        oversized_df = table_df.loc[table_df['code'] == -2, ['dist_down', 'W_s', 'Z_s', 'W_s_Z_s']]

        for i in numpy_columns[1:]:
            plot_1 = wide_bar_df[['dist_down', ('%s' % i)]].to_numpy()
            plot_2 = nozzle_df[['dist_down', ('%s' % i)]].to_numpy()
            plot_3 = normal_df[['dist_down', ('%s' % i)]].to_numpy()
            plot_4 = pool_df[['dist_down', ('%s' % i)]].to_numpy()
            plot_5 = oversized_df[['dist_down', ('%s' % i)]].to_numpy()
            plot_6 = table_df[['dist_down', ('%s' % i)]].to_numpy()

            plt.scatter(plot_1[:, 0], plot_1[:, 1], c='orange', s=0.75, label="Wide bar") # Landform color coded Ws, zs, and C(Ws,Zs) scatter plots
            plt.scatter(plot_2[:, 0], plot_2[:, 1], c='gold', s=0.75, label="Nozzle")
            plt.scatter(plot_3[:, 0], plot_3[:, 1], c='c', s=0.75, label="Normal")
            plt.scatter(plot_4[:, 0], plot_4[:, 1], c='grey', s=0.75, label="Const. Pool")
            plt.scatter(plot_5[:, 0], plot_5[:, 1], c='navy', s=0.75, label="Oversized")

            t1 = plot_5[:, 0]
            t2 = plot_5[:, 1]

            plt.xlabel("Thalweg distance downstream (ft)")
            plt.ylabel("%s" % i)
            axes = plt.gca()
            ymin = min(table_df[i])-1
            if i == 'W_s_Z_s':
                ymax = 5
            else:
                ymax = max(table_df[i])+1
            axes.set_ylim([ymin, ymax])
            axes.set_xlim([0, max(table_df['dist_down'])])
            plt.title("Stage %s %s plot" % (stage, i))
            plt.grid(b=True, which='major', color='#666666', linestyle='-')
            plt.minorticks_on()
            plt.grid(b=True, which='minor', color='#999999', linestyle='-', alpha=0.2)
            plt.legend(('Wide bar', 'Nozzle', 'Normal', 'Pool', 'Oversized'), loc=1)

            fig = plt.gcf()
            fig.set_size_inches(12, 6)
            plt.savefig((figure_output + '\\Stage_%s_%s_plot' % (stage, i)), dpi=300, bbox_inches='tight')
            plt.cla()

            plot_6_sorted = plot_6[plot_6[:, 0].argsort()]  # Gaussian filter plotting
            dist_points = plot_6_sorted[:, 0]
            attribute_points = plot_6_sorted[:, 1]
            gaussian_sigma3 = gaussian_filter1d(attribute_points, 3, axis=0)
            gaussian_sigma6 = gaussian_filter1d(attribute_points, 6, axis=0)

            plt.scatter(dist_points, attribute_points, c='black', s=0.75, ls='--', label=("%s points" % i))
            plt.plot(dist_points, gaussian_sigma3, c='red', ls='-', label="Gaussian filter, sigma=3")
            plt.plot(dist_points, gaussian_sigma6, c='green', ls='-', label="Gaussian filter, sigma=10")
            axes.set_xlim([0, max(table_df['dist_down'])])
            plt.title("Data filtering plot for %s at stage %sft" % (i, stage))
            plt.grid(b=True, which='major', color='#666666', linestyle='-')
            plt.minorticks_on()
            plt.grid(b=True, which='minor', color='#999999', linestyle='-', alpha=0.2)
            plt.legend(('Gaussian filter, sigma=3', 'Gaussian filter, sigma=10'), loc=1)
            fig = plt.gcf()
            fig.set_size_inches(12, 6)
            plt.savefig((figure_output + '\\Stage_%s_%s_filtering_plot' % (stage, i)), dpi=300, bbox_inches='tight')
            plt.cla()




