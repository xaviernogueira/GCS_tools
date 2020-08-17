import arcpy
import csv
import os
from os import listdir
from os.path import isfile, join
from matplotlib import pyplot as plt
import numpy as np
import file_functions
import GCS_statistical_analysis_XRN
import create_station_lines
from create_station_lines import *
import pandas
import openpyxl as xl
import Post_detrend_to_GCS
import classify_landforms_GUI

def find_centerline_nums(detrend_folder):
    '''This function takes a detrend folder location for a given reach, and using string splicing to find the used centerline
    stage numbers. A list of stage numbers from smallest to largest is returned'''
    centerline_nums = []
    centerline_folder = detrend_folder + '\\analysis_centerline_and_XS'

    list = [f for f in listdir(centerline_folder) if f[-6:] == 'DS.shp']

    for line in list:
        if line[18] == 'f':
            stage = int(line[17])
        else:
            stage = int(line[17:19])
        centerline_nums.append(stage)
    centerline_nums.sort()

    return centerline_nums

def loc_stage_finder(stage, centerlines_nums):
    '''Useful function to find the centerline associated with a given stage and list of used stage centerline numbers'''
    if float(stage) > float(centerlines_nums[-1]):
        loc_stage = centerlines_nums[-1]
    elif float(stage) <= float(centerlines_nums[0]):
        loc_stage = centerlines_nums[0]
    else:
        index = 0
        while float(stage) > float(centerlines_nums[index]):
            index += 1
        loc_stage = centerlines_nums[index]

    count = centerlines_nums.index(loc_stage)
    return [loc_stage, count]


def prep_locations(detrend_location,max_stage=20, skip=False):
    '''This function takes a reach and creates a new gcs csv with a location associated with the lowest stage centerline'''
    arcpy.env.overwriteOutput = True

    detrended_raster = detrend_location + "\\ras_detren.tif"
    landform_folder = detrend_location + '\\landform_analysis'  # Make directory for landform analysis xl files and centerline adjusted GCS tables
    centerline_folder = detrend_location + "\\analysis_centerline_and_XS"
    del_files = []
    del_suffix = ['.shp', '.cpg', '.dbf', '.prj', '.sbn', '.sbx', '.shp.xlm', 'shx']

    if not os.path.exists(landform_folder):
        os.makedirs(landform_folder)

    print('Begining centerline reconciliation process...')
    centerlines_nums = []
    centerlines = [f for f in listdir(centerline_folder) if isfile(join(centerline_folder, f)) and f[-5:] == 'S.shp']
    XS_files = [i for i in listdir(centerline_folder) if isfile(join(centerline_folder, i)) and i[-5:] == 't.shp' and len(i) > 32]

    temp_location = []
    cursor = arcpy.SearchCursor(centerline_folder + '\\%s' % XS_files[0])
    for row in cursor:
        temp_location.append(int(row.getValue('LOCATION')))
    temp_location.sort()
    spacing = int(temp_location[1] - temp_location[0])
    print('XS spacing is %sft...' % spacing)

    min_num = 20
    for line in centerlines:
        line_loc = ('%s\\%s' % (centerline_folder, line))
        if line[-11] == '_':
            num = int(line[-10])
        else:
            num = int(line[-11:-9])
        centerlines_nums.append(num)

        if num <= min_num:
            min_num = num
        centerlines_nums.sort()

        station_lines = create_station_lines.create_station_lines_function(line_loc, spacing=spacing, xs_length=5, stage=[])
        station_lines = centerline_folder + ('\\stage_centerline_%sft_DS_XS_%sx5ft.shp' % (num, spacing))

        for suffix in del_suffix:
            del_files.append(station_lines[:-4] + suffix)

        station_points = arcpy.Intersect_analysis([station_lines, line_loc], out_feature_class=(centerline_folder + "\\station_points_%sft.shp" % num), join_attributes="ALL", output_type="POINT")

    print('Using centerlines: %s' % centerlines_nums)
    for num in centerlines_nums:
        station_points = centerline_folder + "\\station_points_%sft.shp" % num

        if num == min_num:
            print("Extracting thalweg elevation for Caamano analysis...")
            loc_field = 'SP_SG_%sFT' % num

            single_station_points = centerline_folder + ("\\%s.shp" % loc_field)

            arcpy.MultipartToSinglepart_management(station_points, out_feature_class=single_station_points)
            z_table = arcpy.sa.Sample(detrended_raster, single_station_points, out_table=(centerline_folder + "\\thalweg_Z.dbf"), unique_id_field='LOCATION')

            for suffix in ['.dbf', '.cpg', '.dbf.xml']:
                del_files.append(centerline_folder + "\\thalweg_Z%s" % suffix)
            for suffix in del_suffix:
                del_files.append(station_points[:-4] + suffix)
                del_files.append(single_station_points[:-4] + suffix)

            centerline_XY_loc = centerline_folder + '\\centerline_XY_%sft.csv' % num  # csv with XY coordinates of the centerlines made
            arcpy.AddXY_management(single_station_points)
            file_functions.tableToCSV(single_station_points, csv_filepath=centerline_XY_loc, fld_to_remove_override=[])

            station_points = arcpy.JoinField_management(station_points, in_field='LOCATION', join_table=z_table, join_field=loc_field, fields=['ras_detren'])
            arcpy.AddField_management(station_points, ('loc_%sft' % num), 'SHORT')
            arcpy.AddField_management(station_points, 'thwg_z', 'FLOAT')
            arcpy.CalculateField_management(station_points, 'loc_%sft' % num, expression='!LOCATION!', expression_type='PYTHON3')
            arcpy.CalculateField_management(station_points, 'thwg_z', expression='!ras_detren!', expression_type='PYTHON3')

            for stage in range(0, max_stage+1):
                stage_f = float(stage)
                arcpy.AddField_management(station_points, ('Dz_%sft' % stage), 'FLOAT')
                arcpy.CalculateField_management(station_points, ('Dz_%sft' % stage), expression=('%s - !thwg_z!' % stage_f), expression_type='PYTHON3')

            del_fields = [f.name for f in arcpy.ListFields(station_points) if f.name[:2] != 'Dz']
            for field in [('loc_%sft' % num), 'FID', 'thwg_z', 'Shape']:
                try:
                    del_fields.remove(field)
                except:
                    "Can't delete field: %s" % field

            arcpy.DeleteField_management(station_points, del_fields)
            print('Fields deleted: %s' % del_fields)

        if num != min_num:
            theis_loc = (centerline_folder + "\\thiessen_%sft.shp" % num)
            arcpy.CreateThiessenPolygons_analysis(station_points, theis_loc, 'ALL')
            arcpy.AddField_management(theis_loc, ('loc_%sft' % num), 'SHORT')
            arcpy.CalculateField_management(theis_loc, ('loc_%sft' % num), expression='!LOCATION!', expression_type='PYTHON3')
            del_fields = [f.name for f in arcpy.ListFields(theis_loc)]
            for field in [('loc_%sft' % num), 'FID', 'Shape']:
                try:
                    del_fields.remove(field)
                except:
                    "Can't delete field: %s" % field
            arcpy.DeleteField_management(theis_loc,del_fields)

    max_count = 0
    for counter, num in enumerate(centerlines_nums):
        theis_loc = (centerline_folder + "\\thiessen_%sft.shp" % num)
        out_points = centerline_folder + ("\\align_points%s.shp" % counter)

        for suffix in del_suffix:
            del_files.append(theis_loc[:-4] + suffix)
            del_files.append(out_points[:-4] + suffix)

        if counter >= max_count:
            max_count = counter
        if counter == 1:
            arcpy.Identity_analysis(centerline_folder + "\\station_points_%sft.shp" % min_num, theis_loc, out_feature_class=out_points, join_attributes='ALL', )
        elif counter > 1:
            arcpy.Identity_analysis(centerline_folder + ("\\align_points%s.shp" % (int(counter-1))), theis_loc, out_feature_class=out_points, join_attributes='ALL', )

    code_csv_loc = landform_folder + '\\all_stages_table.csv'
    file_functions.tableToCSV(out_points, csv_filepath=code_csv_loc, fld_to_remove_override=['FID_statio', 'FID_thiess'])
    print('Empty stages csv created @ %s' % code_csv_loc)

    print('Deleting files: %s' % del_files)
    for file in del_files:
        if os.path.exists(file):
            try:
                os.remove(file)
            except:
                print("Couldn't delete %s" % file)

    return[code_csv_loc,centerlines_nums]

def prep_small_inc(detrend_folder,interval=0.1,max_stage=20):
    '''IN: Folder containing detrended DEM ras_detren.tif, an stage interval length, a maximum flood stafe height.
    RETURNS: None. This function creates a folder containing wetted polygons for a 0.1ft increments as well as a clippped detrended DEM and contours.'''
    del_files = []
    del_suffix = ['.shp', '.cpg', '.dbf', '.prj', '.sbn', '.sbx', '.shp.xml', '.shx', '.tif', '.tif.aux.xml', '.tfw', '.tif.ovr', '.tif.vat.cpg', '.tif.vat.dbf']
    channel_clip_poly = detrend_folder + '\\raster_clip_poly.shp'
    small_wetted_poly_loc = detrend_folder + '\\wetted_polygons\\small_increments'

    if not os.path.exists(small_wetted_poly_loc):  # Make a new folder for the 0.1ft increment wetted polygons
        os.makedirs(small_wetted_poly_loc)

    in_ras = arcpy.sa.Raster(detrend_folder + '\\ras_detren.tif')
    print('Making wetted polygons...')
    for inc in np.arange(0, max_stage+interval, float(interval)):  # Create a polygon representing the portion of the detrended DEM below a stage height interval
        if inc >= 10.0:
            inc_str = (str(inc)[0:2] + 'p' + str(inc)[3])
        else:
            inc_str = (str(inc)[0] + 'p' + str(inc)[2])
        names = [('\\wt_rs_%sft.tif' % inc_str), ('\\wetted_poly_%sft_noclip.shp' % inc_str)]
        wetted_ras = arcpy.sa.Con(in_ras <= inc, 1)
        wetted_ras.save(small_wetted_poly_loc + names[0])
        arcpy.RasterToPolygon_conversion(in_raster=wetted_ras, out_polygon_features=(small_wetted_poly_loc + names[1]), simplify=False)
        arcpy.Clip_analysis(small_wetted_poly_loc + names[1], channel_clip_poly, out_feature_class=(small_wetted_poly_loc + '\\wetted_poly_%sft.shp' % inc_str))

        for name in names:
            del_files.append(small_wetted_poly_loc + name[:-4])
    print('Wetted polygons located @ %s' % small_wetted_poly_loc)

    contour_loc = detrend_folder + '\\detrended_contours.shp'
    clipped_ras_loc = detrend_folder + '\\rs_dt_clip.tif'  # Clipped to channel clip poly

    if not os.path.isfile(contour_loc):  # Create a clipped detrended DEM to the max stage height value, and the channel_clip_poly file
        print('Making contours...')
        max_stage_ras = arcpy.sa.Con(in_ras <= float(max_stage), in_ras)
        max_stage_ras.save(detrend_folder + '\\rs_dt_clip1.tif')
        clipped_ras = arcpy.Clip_management(detrend_folder + '\\rs_dt_clip1.tif', "", clipped_ras_loc, in_template_dataset=channel_clip_poly, clipping_geometry='ClippingGeometry', maintain_clipping_extent='MAINTAIN_EXTENT')
        del_files.append(detrend_folder + '\\rs_dt_clip1')
        contour_ras = arcpy.sa.Contour(clipped_ras_loc, contour_loc, contour_interval=interval)
        print('Contour file and clipped detrended raster made @ %s' % detrend_folder)
    else:
        'Contour file already made @ %s' % contour_loc

    print('Deleting files: %s' % del_files)
    for prefix in del_files:
        for suffix in del_suffix:
            path = prefix + suffix
            if os.path.exists(path):
                try:
                    os.remove(path)
                except:
                    print("Couldn't delete %s" % prefix + suffix)


def align_csv(code_csv_loc, centerlines_nums, max_stage=20):
    '''IN: Aligned csv location, list of used centerline nums, key Zs (optional)
    RETURNS: An aligned csv containing all stages at 1ft increments is returned as a dataframe. '''
    print('Calculating cross-correlation matrix...')
    landform_folder = out_folder + '\\landform_analysis'

    for stage in range(1, max_stage + 1):
        out_points_df = pd.read_csv(code_csv_loc, na_values=-9999)
        out_points_df.sort_values(by=['loc_%sft' % centerlines_nums[0]], inplace=True)

        gcs_csv = out_folder + ('\\gcs_ready_tables\\%sft_WD_analysis_table.csv' % int(stage))

        loc_stage = loc_stage_finder(stage, centerlines_nums)[0]
        j_loc_field = 'loc_%sft' % loc_stage

        temp_df = pd.read_csv(gcs_csv)
        temp_df.sort_values(by=['dist_down'], inplace=True)
        temp_df_mini = temp_df.loc[:, ['dist_down', 'code', 'W', 'W_s', 'Z_s']]
        temp_df_mini.rename({'dist_down': j_loc_field, 'code': ('code_%sft' % stage), 'W': ('W_%sft' % stage),
                             'W_s': ('Ws_%sft' % stage), 'Z_s': ('Zs_%sft' % stage), }, axis=1, inplace=True)
        temp_df_mini.sort_values(by=[j_loc_field], inplace=True)
        result = out_points_df.merge(temp_df_mini, left_on=j_loc_field, right_on=j_loc_field, how='left')
        result = result.replace(np.nan, 0)
        result = result.loc[:, ~result.columns.str.contains('^Unnamed')]
        result.to_csv(code_csv_loc)
        print('Stage %sft added to the all stages csv' % stage)

    print('Stage alignment completed...')

    return result

def key_zs_gcs(detrend_folder, key_zs=[]):
    '''This function does a full GCS analysis using three specific key Zs that can include any float. Results saved
    to the gcs_ready_tables, as well as plotted. Results are aligned to the existing csv to facilitate landform analysis'''
    centerline_nums = find_centerline_nums(detrend_folder)
    wetted_folder = detrend_folder + '\\wetted_polygons\\small_increments'
    centerline_folder = detrend_folder + '\\analysis_centerline_and_XS'
    width_poly_folder = detrend_folder + '\\analysis_shapefiles'
    out_folder = detrend_folder + '\\gcs_ready_tables'
    detrended_DEM = detrend_folder + '\\ras_detren.tif'

    table_list = []

    for z in key_zs:
        if isinstance(z, float) == False:
            z = float(z)
        loc_stage = loc_stage_finder(z, centerline_nums)[0]
        if z >= 10.0:
            z_str = (str(z)[0:2] + 'p' + str(z)[3])
        else:
            z_str = (str(z)[0] + 'p' + str(z)[2])

        wetted_loc = wetted_folder + '\\wetted_poly_%sft.shp' % z_str
        cross_sections = centerline_folder + '\\stage_centerline_2ft_DS_XS_%sft.shp' % loc_stage

        temp_location = []
        cursor = arcpy.SearchCursor(cross_sections)
        for row in cursor:
            temp_location.append(int(row.getValue('LOCATION')))
        temp_location.sort()
        spacing = int(temp_location[1] - temp_location[0])
        print('XS spacing is %sft...' % spacing)

        clipped_XS = arcpy.Clip_analysis(cross_sections,wetted_loc,out_feature_class=width_poly_folder + 'clipped_station_lines_%sft.shp' % z_str)
        width_poly = arcpy.Buffer_analysis(clipped_XS, width_poly_folder + 'width_rectangles_%sft.shp' % z_str, spacing / 2, line_side='FULL', line_end_type='FLAT')
        arcpy.AddField_management(width_poly, "Width", field_type="FLOAT")
        expression = ("(float(!Shape.area!)) / %d" % spacing)
        arcpy.CalculateField_management(width_poly, "Width", expression, "PYTHON3")
        print('Width polygons for %sft stage created...' % z)

        arcpy.AddField_management(width_poly, field_name="loc_id", field_type="SHORT")
        field_calc = "(int(!LOCATION!))"
        arcpy.CalculateField_management(width_poly, field="loc_id", expression=field_calc, expression_type="PYTHON3")
        zonal_table = arcpy.sa.ZonalStatisticsAsTable(width_poly, "loc_id", detrended_DEM, out_table=(width_poly_folder + '\\stats_table_%s.dbf' % z_str), statistics_type="MEAN")
        width_poly = arcpy.JoinField_management(width_poly, "loc_id",  join_table=zonal_table, join_field="loc_id", fields=["MEAN"])

        csv_loc = "WD_analysis_table_%s.csv" % z_str
        #analysis_table = arcpy.TableToTable_conversion(width_poly, out_path=table_location, out_name=("WD_analysis_table_%s.dbf" % z_str))
        #analysis_xlsx = arcpy.TableToExcel_conversion(Input_Table=width_poly, Output_Excel_File=table_location + ("\\WD_analysis_table_%s.xlsx" % z_str))
        tableToCSV(width_poly, csv_filepath=csv_loc, fld_to_remove_override=[])
        df = pd.read_csv(csv_loc)
        df.sort_values(by=['dist_down'], inplace=True)
        df.rename({'LOCATION': 'dist_down', 'Width': 'W', 'MEAN': 'Z'}, axis=1, inplace=True)
        df.to_csv(csv_loc)
        table_list.append(csv_loc)

    classify_landforms_GUI.main_classify_landforms(table_list, w_field='W', z_field='Z', dist_field='dist_down', out_folder=detrend_folder, make_plots=False)



def key_z_finder(out_folder, channel_clip_poly, code_csv_loc, centerlines_nums, key_zs=[], max_stage=20, small_increments=0):
    '''INPUT: Linear detrending output folder, clip polygon capturing all relevent wetted area, pearson correlation threshold (optional), maximum stage for plotting
    RETURNS: Pearson correlation matrix comaparing the width series of each stage with every other stage. CDF and PDF plots of accumulating wetted areas
    Used to guide key Z selection for the following nested landform analysis'''
    print('Calculating cross-correlation matrix...')
    landform_folder = out_folder + '\\landform_analysis'
    result = pd.read_csv(code_csv_loc)

    col_row_heads = [('%sft' % f) for f in range(1, max_stage + 1)]
    col_list = [('Ws_%sft') % f for f in range(1, max_stage + 1)]

    cross_corrs = []
    in_data = result.loc[:, col_list]
    cross_corrs_df = in_data.corr()

    for num in range(1, max_stage + 1):  # Putting cross correlation dataframe in a maptplot format
        row_data = cross_corrs_df.loc[:, ['Ws_%sft' % num]].astype(float)
        row_data = row_data.squeeze()
        row_list = row_data.values.tolist()
        cross_corrs.append(row_list)

    for list in cross_corrs:  # Making sure all lists in cross_corrs are the same length to avoid plotting error
        if len(list) != len(col_row_heads):
            print('Error found. Repairing a coefficient list...')
            gap = int(len(col_row_heads) - len(list))
            for g in range(gap):
                list.append(0.0)

    fig, ax = plt.subplots()  # Plotting cross-correlation matrix
    im = ax.imshow(np.array(cross_corrs, dtype=float))
    ax.set_xticks(np.arange(len(col_row_heads)))
    ax.set_yticks(np.arange(len(col_row_heads)))
    ax.set_xticklabels(col_row_heads)
    ax.set_yticklabels(col_row_heads)

    for i in range(len(col_row_heads)):
        for j in range(len(col_row_heads)):
            text = ax.text(j, i, round(cross_corrs[i][j],2),
                           ha="center", va="center", fontsize=6, color="r")

    ax.set_title("Cross-correlation of stage width series")
    fig.tight_layout()
    fig.set_size_inches(16, 8)
    plt.savefig((landform_folder + '\\cross_corrs_table.png'), dpi=300, bbox_inches='tight')
    plt.cla()
    print('Stage width profile correlation matrix: %s' % (landform_folder + '\\cross_corrs_table.png'))

    print('CDF and PDF wetted area analysis initiated...')
    if small_increments == 0:
        print('Making 1ft flood stage height interval plots...')
        clipped_wetted_folder = out_folder + "\\clipped_wetted_polygons"
        if not os.path.exists(clipped_wetted_folder):
            os.makedirs(clipped_wetted_folder)

        wetted_polys = [f for f in listdir(out_folder+'\\wetted_polygons') if f[:26]=='flood_stage_poly_dissolved' and f[-3:]=='shp']

        print('Calculating wetted areas...')
        wetted_areas = [None]*len(wetted_polys)
        for poly in wetted_polys:
            poly_loc = ('%s\\wetted_polygons\\%s' % (out_folder, poly))
            if poly[28] == 'f':
                stage = int(poly[27])
            else:
                stage = int(poly[27:29])
            if stage <= max_stage:
                clip_poly = arcpy.Clip_analysis(poly_loc, channel_clip_poly, out_feature_class=('%s\\clipped_wetted_poly_%sft' % (clipped_wetted_folder, stage)))
                geometries = arcpy.CopyFeatures_management(clip_poly, arcpy.Geometry())
                poly_area = 0
                for geometry in geometries:
                    poly_area += float(geometry.area)
                wetted_areas[stage] = poly_area

    else:
        print('Making small increment plots...')
        wetted_areas = []
        wetted_polys = [f for f in listdir(out_folder + '\\wetted_polygons\\small_increments') if f[:11] == 'wetted_poly' and f[-6:] == 'ft.shp']
        flood_stage_incs = np.arange(0, max_stage+small_increments, float(small_increments)).tolist() # A list storing all small flood stage increments

        print('Calculating wetted areas...')
        for poly in wetted_polys:
            poly_loc = out_folder + '\\wetted_polygons\\small_increments\\%s' % poly
            geometries = arcpy.CopyFeatures_management(poly_loc, arcpy.Geometry())
            poly_area = 0
            for geometry in geometries:
                poly_area += float(geometry.area)
            wetted_areas.append(poly_area)

    wetted_areas = [i for i in wetted_areas if i != None]
    wetted_areas.sort()

    print('Calculating centerline lengths, d(wetted area), and d(XS length)...')
    centerline_lengths = [None]*len(centerlines_nums)

    for count, line in enumerate(centerlines_nums):
        line_loc = ('%s\\analysis_centerline_and_XS\\stage_centerline_%sft_DS.shp' % (out_folder,line))
        geometries = arcpy.CopyFeatures_management(line_loc, arcpy.Geometry())
        poly_length = 0
        for geometry in geometries:
            poly_length += float(geometry.length)
        centerline_lengths[count] = poly_length

    mean_XS_length = []  # Calculates mean width per stage as wetted area / centerline length
    for count, area in enumerate(wetted_areas):
        if small_increments != 0:
            stage = flood_stage_incs[count]
        else:
            stage = count
        index = loc_stage_finder(stage,centerlines_nums)[1]
        length = centerline_lengths[index]
        if length != None:
            mean_XS_length.append(float(area/length))
    mean_XS_length = [i for i in mean_XS_length if i != None]

    d_area = []  # Calculates the change in wetted area between stages
    for count, area in enumerate(wetted_areas):
        if count == 0:
            d_area.append(area)
        else:
            d_area.append(float(area-wetted_areas[count-1]))

    d_XS_length = []  # Calculates the change in mean width between stages
    for count, length in enumerate(mean_XS_length):
        if count == 0:
            d_XS_length.append(length)
        else:
            d_XS_length.append(float(length - mean_XS_length[count - 1]))

    max_area = wetted_areas[-1]
    print('Plotting CDF and PDF plots')

    if small_increments == 0:
        x1 = np.array(range(0, len(wetted_areas)))
        title = (out_folder + '\\CDF_plot.png')
    else:
        x1 = np.arange(0, max_stage+small_increments, small_increments)
        title = (out_folder + '\\CDF_plot_small_inc.png')
    y1 = np.array([(float(f/max_area))*100 for f in wetted_areas])
    plt.figure()
    plt.plot(x1,y1)
    plt.xlabel('Flood stage height (ft)')
    plt.ylabel('Percent of %sft stage area' % max_stage)
    plt.title('CDF chart')
    plt.grid(b=True, which='major', color='#666666', linestyle='-')
    plt.xlim(0,max_stage)
    plt.ylim(0,max(y1))
    plt.xticks(np.arange(0, (max_stage+1), step=1))
    if len(key_zs) != 0:
        for stage in key_zs:
            plt.axvline(x=stage, color='r', linestyle='--')
            title = (out_folder + '\\CDF_plot_with_Zs.png')
    fig = plt.gcf()
    fig.set_size_inches(12, 6)
    plt.savefig(title, dpi=300, bbox_inches='tight')
    plt.clf()
    plt.close('all')

    if small_increments == 0:
        x2 = np.array(range(0, len(d_area))) #Add saving optionality
        y2 = np.array(d_area)
        title = (out_folder + '\\PDF_plot.png')
    else:
        x2 = np.arange(small_increments, max_stage + small_increments, small_increments)
        y2 = np.array(d_area[1:])
        title = (out_folder + '\\PDF_plot_small_inc.png')
    plt.figure()
    plt.plot(x2, y2)
    plt.xlabel('Flood stage height (ft)')
    plt.ylabel('Change in area (sq ft)')
    plt.title('PDF chart')
    plt.grid(b=True, which='major', color='#666666', linestyle='-')
    plt.xlim(0, max_stage)
    plt.ylim(0, None)
    plt.xticks(np.arange(0, (max_stage+1), step=1))
    if len(key_zs) != 0:
        for stage in key_zs:
            plt.axvline(x=stage, color='r', linestyle='--')
            title = (out_folder + '\\PDF_plot_with_Zs.png')
    fig = plt.gcf()
    fig.set_size_inches(12, 6)
    plt.savefig(title, dpi=300, bbox_inches='tight')
    plt.clf()
    plt.close('all')

    if small_increments == 0:
        x3 = np.array(range(0, len(wetted_areas)))
        title = (out_folder + '\\wetted_areas_plot.png')
    else:
        x3 = np.arange(0, max_stage + small_increments, small_increments)
        title = (out_folder + '\\wetted_areas_plot_small_inc.png')
    y3 = np.array(wetted_areas)
    plt.figure()
    plt.plot(x3, y3)
    plt.xlabel('Flood stage height (ft)')
    plt.ylabel('Wetted area (sq ft)')
    plt.title('Cumulative wetted area chart')
    plt.grid(b=True, which='major', color='#666666', linestyle='-')
    plt.xlim(0, max_stage)
    plt.ylim(0, None)
    plt.xticks(np.arange(0, (max_stage + 1), step=1))
    if len(key_zs) != 0:
        for stage in key_zs:
            plt.axvline(x=stage, color='r', linestyle='--')
            title = (out_folder + '\\wetted_area_plot_with_Zs.png')
    fig = plt.gcf()
    fig.set_size_inches(12, 6)
    plt.savefig(title, dpi=300, bbox_inches='tight')
    plt.clf()
    plt.close('all')

    x4 = np.array(mean_XS_length)
    if small_increments == 0:
        y4 = np.array(range(0, len(mean_XS_length)))
        title = (out_folder + '\\XS_length_plot.png')
    else:
        y4 = np.arange(0, max_stage+small_increments, small_increments)
        title = (out_folder + '\\XS_length_plot_small_inc.png')
    plt.figure()
    plt.plot(x4, y4)
    plt.xlabel('Mean XS length')
    plt.ylabel('Flood stage height (ft)')
    plt.title('Mean XS length chart')
    plt.grid(b=True, which='major', color='#666666', linestyle='-')
    plt.xlim(0, max(x4))
    plt.ylim(0, max_stage)
    plt.xticks(np.arange(0, int(max(x4)), step=20))
    if len(key_zs) != 0:
        for stage in key_zs:
            plt.axvline(x=stage, color='r', linestyle='--')
            title = (out_folder + '\\XS_length_plot_with_Zs.png')
    fig = plt.gcf()
    fig.set_size_inches(12, 6)
    plt.savefig(title, dpi=300, bbox_inches='tight')
    plt.clf()
    plt.close('all')

    if small_increments == 0:
        x5 = np.array(range(0, len(d_XS_length)))
        y5 = np.array(d_XS_length)
        title = (out_folder + '\\PDF_XS_plot.png')
    else:
        x5 = np.arange(small_increments, max_stage + small_increments, small_increments)
        y5 = np.array(d_XS_length[1:])
        title = (out_folder + '\\PDF_XS_plot_small_inc.png')
    plt.figure()
    plt.plot(x5, y5)
    plt.xlabel('Flood stage height (ft)')
    plt.ylabel('Change in mean XS length (ft)')
    plt.title('PDF XS chart')
    plt.grid(b=True, which='major', color='#666666', linestyle='-')
    plt.xlim(0, max_stage)
    plt.ylim(0, None)
    plt.xticks(np.arange(0, (max_stage + 1), step=1))
    if len(key_zs) != 0:
        for stage in key_zs:
            plt.axvline(x=stage, color='r', linestyle='--')
            title = (out_folder + '\\PDF_XS_plot_with_Zs.png')
    fig = plt.gcf()
    fig.set_size_inches(12, 6)
    plt.savefig(title, dpi=300, bbox_inches='tight')
    plt.clf()
    plt.close('all')

def nested_landform_analysis(aligned_csv, key_zs):
    '''IN: Aligned csv with landform codes for each XS. A list (key_zs) containing three stages
    RETURNS: A xl table containing the abundance of each unique nested landform set'''
    landform_folder = str(os.path.dirname(aligned_csv))
    code_dict = {-2: 'O', -1: 'CP', 0: 'NC', 1: 'WB', 2: 'NZ'}  # code number and corresponding MU
    print('Starting nested landform analysis...')

    if len(key_zs) == 0:
        return print('No key Zs selected, please add parameters')

    key_zs.sort()
    aligned_df = pd.read_csv(aligned_csv)

    code_df_list = []  # code_df_list[0] is baseflow, [1] is bankful, and [2] is flood stage
    for key_z in key_zs:
        if isinstance(key_z, float) == True:
            if key_z >= 10.0:
                key_z = (str(key_z)[0:2] + 'p' + str(key_z)[3])
            else:
                key_z = (str(key_z)[0] + 'p' + str(key_z)[2])
        code_df_temp = aligned_df.loc[:, [('code_%sft' % key_z)]].squeeze()
        code_df_list.append(code_df_temp.values.tolist())

    nested_landforms = list(zip(code_df_list[0], code_df_list[1], code_df_list[2]))
    unique_nests = list(set(nested_landforms))

    unique_nest_counts = list(np.zeros(len(unique_nests), dtype=int))  # initialize list of lists to count abundance

    for nest in nested_landforms:
        i = unique_nests.index(nest)
        unique_nest_counts[i] += 1

    nest_abundances = list(zip(unique_nests, unique_nest_counts))
    nest_abundances.sort(key=lambda x: x[1], reverse=True)

    nested_analysis_xl = (landform_folder + '\\nested_landforms.xlsx')
    wb = xl.Workbook()
    wb.save(nested_analysis_xl)
    ws = wb.active
    ws.title = 'Nested landforms abundances'
    ws.cell(row=1, column=1).value = 'Nested landform set [baseflow, BF, flood]'
    ws.cell(row=1, column=2).value = 'Abundances'
    ws.cell(row=1, column=3).value = '% of unique sets'
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 16

    for count, unique_set in enumerate(nest_abundances):
        string = '%s, %s, %s' % (code_dict[unique_set[0][0]],code_dict[unique_set[0][1]],code_dict[unique_set[0][2]])
        ws.cell(row=2 + count, column=1).value = str(string)
        ws.cell(row=2 + count, column=2).value = unique_set[1]
        ws.cell(row=2 + count, column=3).value = round((unique_set[1] / len(nested_landforms)) * 100, 2)

    wb.save(nested_analysis_xl)
    wb.close()
    print('Nested landform analysis complete. Results @ %s' % nested_analysis_xl)

def heat_plotter(comids, geo_class, key_zs=[], max_stage=20):
    '''IN: a list with either one or multiple comids. Key_zs list if filled makes subplots, if not a plot for each stage is made.
    If multiple comids are present, key_zs list structure is important. EXAMPLE: comids= [123,456,789], key_zs = [[baseflow, BF, flood],[baseflow, BF, flood],[baseflow, BF, flood]
    Heatmaps plotted and saved in landform folder, or new folder for class averaged figures'''
    titles = []

    if len(comids) > 1:
        print('Making hexagaon heatplot for geomorphic class %s' % geo_class)
        x_list_of_arrays = [[],[],[]] #Initialize list containing [baseflow, bankful, flood] values
        y_list_of_arrays = [[],[],[]]

        for count, comid in enumerate(comids):
            landform_folder = (r"Z:\users\xavierrn\SoCoast_Final_ResearchFiles\SCO%s\COMID%s\LINEAR_DETREND\landform_analysis" % (geo_class, comid))

            for index, z in enumerate(key_zs[count]):
                if isinstance(z, float) ==  True:
                    if z >= 10.0:
                        z = (str(z)[0:2] + 'p' + str(z)[3])
                    else:
                        z = (str(z)[0] + 'p' + str(z)[2])

                data = pd.read_csv(landform_folder[:-18] + '\\gcs_ready_tables\\%sft_WD_analysis_table.csv' % z)
                x_temp = data.loc[:, [('W_s')]].squeeze().to_list()
                y_temp = data.loc[:, [('Z_s')]].squeeze().to_list()
                for value in range(len(x_temp)):
                    x_list_of_arrays[index].append(x_temp[value])
                    y_list_of_arrays[index].append(y_temp[value])

        fig, axs = plt.subplots(ncols=int(len(key_zs[0])), sharey=True, figsize=(7, 7))
        fig.subplots_adjust(hspace=0.5, left=0.07, right=0.93)

        xmax = 0
        ymax = 0
        key_z_meanings = {0:'Baseflow',1:'Bankful',2:'Flood'}

        for count, z in enumerate(key_zs[0]):
            titles.append('Class %s, %s stage' % (geo_class, key_z_meanings[count]))

            x = np.asarray(x_list_of_arrays[count])
            y = np.asarray(y_list_of_arrays[count])
            if count == 0:
                xmax = float(np.percentile(x, 99))
                ymax = float(np.percentile(y, 99))
                xmin = float(np.percentile(x, 1))
                ymin = float(np.percentile(y, 1))

            if float(np.percentile(x, 99)) >= xmax:
                xmax = float(np.percentile(x, 99))
            if float(np.percentile(y, 99)) >= ymax:
                ymax = float(np.percentile(y, 99))
            if float(np.percentile(x, 1)) <= xmin:
                xmin = float(np.percentile(x, 1))
            if float(np.percentile(y, 1)) <= ymin:
                ymin = float(np.percentile(y, 1))

        for count, ax in enumerate(axs):
            x = np.asarray(x_list_of_arrays[count])
            y = np.asarray(y_list_of_arrays[count])

            hb = ax.hexbin(x, y, gridsize=40, cmap='YlOrRd')
            ax.set(xlim=(xmin, xmax), ylim=(xmin, ymax))
            ax.set_title(titles[count])
            ax.grid(b=True, which='major', color='#9e9e9e', linestyle='--')
            ax.set_xlabel('Standardized width (Ws)')
            ax.set_ylabel('Standardized detrended elevation (Zs)')

        save_title = (r"Z:\users\xavierrn\SoCoast_Final_ResearchFiles\SCO%s" % geo_class) + ('\\class%s_keyZs_heatplot.png' % geo_class)
        fig = plt.gcf()
        fig.set_size_inches(16, 10)
        plt.savefig(save_title, dpi=300, bbox_inches='tight')
        plt.show()
        plt.clf()
        plt.close('all')
        print('Plot comparing key Zs for all class %s reaches completed. Located @ %s' % (geo_class, save_title))

    elif len(key_zs) != 0:
        fig, axs = plt.subplots(ncols=int(len(key_zs)), sharey=True, figsize=(7, 7))
        fig.subplots_adjust(hspace=0.5, left=0.07, right=0.93)

        xmax = 0
        ymax = 0

        for count, z in enumerate(key_zs):
            if isinstance(z, float) == True:
                if z >= 10.0:
                    z = (str(z)[0:2] + 'p' + str(z)[3])
                else:
                    z = (str(z)[0] + 'p' + str(z)[2])

            titles.append('COMID%s, class %s, stage %sft' % (comids[0], geo_class, z))
            landform_folder = (r"Z:\users\xavierrn\SoCoast_Final_ResearchFiles\SCO%s\COMID%s\LINEAR_DETREND\landform_analysis" % (geo_class, comids[0]))
            data = pd.read_csv(landform_folder[:-18] + '\\gcs_ready_tables\\%sft_WD_analysis_table.csv' % z)
            data = data.loc[:, ~data.columns.str.contains('^Unnamed')]

            x = data.loc[:, [('W_s')]].to_numpy()
            y = data.loc[:, [('Z_s')]].to_numpy()

            if count == 0:
                xmax = float(np.percentile(x, 99))
                ymax = float(np.percentile(y, 99))
                xmin = float(np.percentile(x, 1))
                ymin = float(np.percentile(y, 1))

            if float(np.percentile(x, 99)) >= xmax:
                xmax = float(np.percentile(x, 99))
            if float(np.percentile(y, 99)) >= ymax:
                ymax = float(np.percentile(y, 99))
            if float(np.percentile(x, 1)) <= xmin:
                xmin = float(np.percentile(x, 1))
            if float(np.percentile(y, 1)) <= ymin:
                ymin = float(np.percentile(y, 1))

        for count, ax in enumerate(axs):
            key_z = key_zs[count]
            if isinstance(key_z, float) == True:
                if key_z >= 10.0:
                    key_z = (str(key_z)[0:2] + 'p' + str(key_z)[3])
                else:
                    key_z = (str(key_z)[0] + 'p' + str(key_z)[2])
            landform_folder = (r"Z:\users\xavierrn\SoCoast_Final_ResearchFiles\SCO%s\COMID%s\LINEAR_DETREND\landform_analysis" % (geo_class, comids[0]))
            data = pd.read_csv(landform_folder[:-18] + '\\gcs_ready_tables\\%sft_WD_analysis_table.csv' % key_z)
            data = data.loc[:, ~data.columns.str.contains('^Unnamed')]

            x = data.loc[:, [('W_s')]].to_numpy()
            y = data.loc[:, [('Z_s')]].to_numpy()

            hb = ax.hexbin(x, y, gridsize=30, cmap='YlOrRd')
            ax.set(xlim=(xmin, xmax), ylim=(xmin, ymax))
            ax.set_title(titles[count])
            ax.grid(b=True, which='major', color='#9e9e9e', linestyle='--')
            ax.set_xlabel('Standardized width (Ws)')
            ax.set_ylabel('Standardized detrended elevation (Zs)')

        save_title = landform_folder + '\\comid%s_KEY_Zs_heatplot.png' % comids[0]
        fig = plt.gcf()
        fig.set_size_inches(16, 10)
        plt.savefig(save_title, dpi=300, bbox_inches='tight')
        plt.show()
        plt.clf()
        plt.close('all')
        print('Plot comparing key Zs %s for comid %s. Located @ %s' % (key_zs,comids[0],save_title))

    else:
        for stage in range(1, max_stage + 1):
            print('Making hexagon heatplot for comid %s, geomorphic class %s, stage %sft...' % (comids, geo_class,stage))
            landform_folder = (r"Z:\users\xavierrn\SoCoast_Final_ResearchFiles\SCO%s\COMID%s\LINEAR_DETREND\landform_analysis" % (geo_class, comids[0]))
            data = pd.read_csv(landform_folder[:-18] + '\\gcs_ready_tables\\%sft_WD_analysis_table.csv' % stage)
            data = data.loc[:, ~data.columns.str.contains('^Unnamed')]

            x = data.loc[:, ['W_s']].to_numpy()
            y = data.loc[:, ['Z_s']].to_numpy()

            xmax = float(np.percentile(x,99))
            ymax = float(np.percentile(y, 99))
            xmin = float(np.percentile(x, 1))
            ymin = float(np.percentile(y, 1))

            save_title = landform_folder + '\\comid%s_stage%sft_heatplot.png' % (comids[0], stage)
            titles.append('COMID%s, class %s, %sft stage' % (comids[0], geo_class, stage))

            plt.hexbin(x, y, gridsize=50, cmap='YlOrRd')
            plt.axis([xmin, xmax, ymin, ymax])
            plt.title(titles[stage-1])
            plt.grid(b=True, which='major', color='#9e9e9e', linestyle='--')
            plt.xlabel('Standardized width (Ws)')
            plt.ylabel('Standardized detrended elevation (Zs)')

            fig = plt.gcf()
            fig.set_size_inches(10, 10)
            plt.savefig(save_title, dpi=300, bbox_inches='tight')
            plt.clf()
            plt.close('all')

    print('Plots completed')


def caamano_analysis(aligned_csv):
    '''IN: Aligned csv with landform codes for each XS.
    OUT:'''

###### INPUTS ######
comid_list = [17610661,17586504,17610257,17573013,17573045,17586810,17609015,17585738,17586610,17610235,17595173,17607455,17586760,17563722,17594703,17609699,17570395,17585756,17611423,17609755,17569841,17563602,17610541,17610721,17610671]
              #17586504,17610257,17573013,17573045,17586810,17609015,17585738,17586610,17610235,17595173,17607455,17586760,17563722,17594703,17609699,17570395,17585756,17611423,17609755,17569841,17563602,17610541,17610721,17610671]
#[17569535,22514218,17607553,17609707,17609017,17610661]
#[17585738,17586610,17610235,17595173,17607455,17586760,17563722,17594703,17609699,17570395,17585756,17611423,17609755,17569841,17563602,17610541,17610721,17610671]
SCO_list = [1,2,2,2,2,2,2,3,3,3,3,3,3,4,4,4,4,4,4,5,5,5,5,5,5]
            #2,2,2,2,2,2,3,3,3,3,3,3,4,4,4,4,4,4,5,5,5,5,5,5]


for count, comid in enumerate(comid_list):
    SCO_number = SCO_list[count]
    direct = (r"Z:\users\xavierrn\SoCoast_Final_ResearchFiles\SCO%s\COMID%s" % (SCO_number, comid))
    out_folder = direct + '\\LINEAR_DETREND'
    process_footprint = direct + '\\las_footprint.shp'
    table_location = out_folder + "\\gcs_ready_tables"
    channel_clip_poly = out_folder + '\\raster_clip_poly.shp'
    code_csv_loc = out_folder + '\\landform_analysis\\all_stages_table.csv'
    aligned_csv_loc = out_folder + '\\landform_analysis\\all_stages_table.csv'

    arcpy.env.overwriteOutput = True

    #prep_small_inc(detrend_folder=out_folder, interval=0.1, max_stage=20) #Ran with all reaches!
    #out_list = prep_locations(detrend_location=out_folder, max_stage=20) #out_list[0]=code_csv_loc, centerline_nums = out_list[1]
    #align_csv(code_csv_loc, centerlines_nums=out_list[1], max_stage=20)

    key_z_finder(out_folder, channel_clip_poly, code_csv_loc=aligned_csv_loc, centerlines_nums=find_centerline_nums(detrend_folder=out_folder), key_zs=[], max_stage=20, small_increments=0.1)
    #nested_landform_analysis(aligned_csv=aligned_csv_loc, key_zs=[]) #Update so a float as a key z can refer to the float to string system
    #heat_plotter(comids=comid_list, geo_class=3, key_zs=[[1,3,6],[2,3,7]], max_stage=20) #Make sure updates for float key zs work

