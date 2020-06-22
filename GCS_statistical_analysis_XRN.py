import numpy as np
import matplotlib
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
import scipy as sp
import pandas as pd
import GCS_analysis
import openpyxl as xl
from openpyxl import Workbook
import os
from os import listdir
from os.path import isfile, join

#This py file updates functions in the GCS_analysis for Python 3#
def analysis_setup(table_directory):
    '''This function takes a directory containing csv files with landform analysis for each flow, deletes all extra in the directory files and creates output xl sheets for statistical results.
    Args: Table directory, Returns: [stages_dict, stages_stats_xl_dict, GCS stats directory]. Stages_dict stores pandas data frames for each stage, stages_stats_xl_dict stores stage level xlsx tables.
    Keys for both dictionaries are StageXft where X is stage height'''

    list_of_files_in_out_folder = [f for f in listdir(table_directory) if isfile(join(table_directory, f))]
    csv_tables = []
    for file in list_of_files_in_out_folder:
        if file[-4:] == ".csv":
            csv_tables.append(table_directory + "\\" + file)
        else:
            os.remove(table_directory + "\\" + file)
    print("List of tables ready to be formatted: %s" % csv_tables)
    stat_table_location = table_directory + "\\GCS_stat_tables_and_plots"
    if not os.path.exists(stat_table_location):
        os.makedirs(stat_table_location)

    stages_dict = {}
    stages_stats_xl_dict = {}
    max_stage = 0
    for table in csv_tables:
        base_name = os.path.basename(table)
        if base_name[1] == "f":
            stage = int(base_name[0])
        else:
            stage = int(base_name[:2])
        if int(stage) > max_stage:
            max_stage = stage
        stage_stats_xl_name = (stat_table_location + '\\%sft_stats_table.xlsx' % stage)
        wb = xl.Workbook()
        wb.save(stage_stats_xl_name)
        wb.close()

        stage_df = pd.read_csv(table)
        stages_dict['Stage_%sft' % stage] = stage_df
        stages_stats_xl_dict['Stage_%sft' % stage] = stage_stats_xl_name
    print("Max stage is %sft" % max_stage)
    return [stages_dict,stages_stats_xl_dict,max_stage,stat_table_location]

def stage_level_descriptive_stats(stages_dict,stages_stats_xl_dict,max_stage,box_and_whisker=False):
    '''This function takes the stages_dict of pandas dataframes, and the stage_stats_xl dictionary, as well as calculated max stage as arguments.
    It fills out the xlsx file for each respective stage height with mean, std, max, min, and median values for the stage as a hole and each landform
    If the box_and_whisker parameter is set to True (not default), then box and whisker plots comparing the W, Z, and C(W,Z)'''
    for stage in range(1,max_stage+1):
        print("Writing descriptive stats for stage %sft" % stage)
        stage_df = stages_dict['Stage_%sft' % stage]
        stage_stat_xl = stages_stats_xl_dict['Stage_%sft' % stage]

        list_of_fields = ['W', 'Z', 'W_s_Z_s','W_s','Z_s']
        means = []
        stds = []
        high = []
        low = []
        medians = []
        for field in list_of_fields[:3]: # add the mean and std_deviation of each field in list_of_fields to a list
            means.append(np.mean(stage_df.loc[:, field].to_numpy()))
            stds.append(np.std(stage_df.loc[:, field].to_numpy()))
            high.append(np.max(stage_df.loc[:, field].to_numpy()))
            low.append(np.min(stage_df.loc[:, field].to_numpy()))
            medians.append(np.median(stage_df.loc[:, field].to_numpy()))

        wb = xl.load_workbook(stage_stat_xl)
        ws = wb.active
        ws.title = 'Descriptive stats'

        ws.cell(row=2, column=1).value = 'MEAN' #Setting up titles on xl
        ws.cell(row=3, column=1).value = 'STD'
        ws.cell(row=4, column=1).value = 'MAX'
        ws.cell(row=5, column=1).value = 'MIN'
        ws.cell(row=6, column=1).value = 'MEDIAN'

        for field in list_of_fields[:3]: #add values in each field column
            field_index = int(list_of_fields.index(field))

            ws.cell(row=1, column=(2 + field_index)).value = field
            ws.cell(row=2, column=(2 + field_index)).value = means[field_index]
            ws.cell(row=3, column=(2 + field_index)).value = stds[field_index]
            ws.cell(row=4, column=(2 + field_index)).value = high[field_index]
            ws.cell(row=5, column=(2 + field_index)).value = low[field_index]
            ws.cell(row=6, column=(2 + field_index)).value = medians[field_index]

        wb.save(stage_stat_xl)
        print("Descriptive stats for W, Z, and W_s_W_Z_s saved for stage %sft..." % stage)

        list_of_codes = [-2,-1,0,1,2]
        landform_dict = {-2:'Oversized',-1:'Constricted pool',0:'Normal',1:'Wide riffle',2:'Nozzle'}
        ws['F1'].value = '*Code: -2 for oversized, -1 for constricted pool, 0 for normal channel, 1 for wide riffle, and 2 for nozzle'
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['A'].width = 11

        w_codes_list = [] #initiate lists of arrays to store data for multiple box plot
        z_codes_list = []
        cwz_codes_list = []
        list_of_field_lists = [w_codes_list,z_codes_list,cwz_codes_list]
        box_plot_dict = dict(zip(list_of_fields,list_of_field_lists)) #This has W,Z, and W_s_Z_s as keys referncing lists that will store the data for each subplot

        total_rows = len(stage_df.index)
        above_1_list = [0,0,0] #'W', 'Z', 'W_s_Z_s'
        above_half_list = [0,0,0]
        for index, row in stage_df.iterrows():
            if abs(row['W_s_Z_s']) >= stds[2]:
                above_1_list[2] += 1
                above_half_list[2] += 1
            elif abs(row['W_s_Z_s']) >= (0.5*stds[2]):
                above_half_list[2] += 1

            for field in list_of_fields[3:]: #List splice: ['W_s', 'Z_s']
                field_index = int(list_of_fields[3:].index(field))
                if abs(row[field]) >= 1:
                    above_1_list[field_index] += 1
                    above_half_list[field_index] += 1
                elif abs(row[field]) >= 0.5:
                    above_half_list[field_index] += 1

        ws.cell(row=7, column=1).value = "% >= 0.5 STD"
        ws.cell(row=8, column=1).value = "% >= 1 STD"

        for index in range(len(above_1_list)): #Calculates % of W, Z, and W_s_Z_s that are greater than 0.5 and 1 of their standard deviations
            percent_above_half = float((above_half_list[index]/total_rows)*100)
            percent_above_1 = float((above_1_list[index]/total_rows)*100)
            ws.cell(row=7, column=(2 + index)).value = percent_above_half
            ws.cell(row=8, column=(2 + index)).value = percent_above_1


        row_num = 2
        for code in list_of_codes: #Calculating same descriptive stats for each landform, each table is spaced 7 cells apart
            code_df = stage_df.loc[stage_df['code'] == code, ['dist_down', 'W', 'W_s', 'Z', 'Z_s', 'W_s_Z_s']]
            ws.cell(row=row_num, column=7).value = (str(landform_dict[code])) #Preparing the table
            ws.cell(row=row_num + 1, column=7).value = 'MEAN'
            ws.cell(row=row_num + 2, column=7).value = 'STD'
            ws.cell(row=row_num + 3, column=7).value = 'MAX'
            ws.cell(row=row_num + 4, column=7).value = 'MIN'
            ws.cell(row=row_num + 5, column=7).value = 'MEDIAN'
            ws.cell(row_num + 6, column=7).value = '% Abundance:'

            for field in list_of_fields:
                field_index = int(list_of_fields.index(field))
                ws.cell(row=row_num,column=(8+field_index)).value = str(field)
                ws.cell(row=row_num+1,column=(8+field_index)).value = (np.mean(code_df.loc[:, field].to_numpy()))
                ws.cell(row=row_num+2,column=(8+field_index)).value = (np.std(code_df.loc[:, field].to_numpy()))
                ws.cell(row=row_num+3,column=(8+field_index)).value = (np.max(code_df.loc[:, field].to_numpy()))
                ws.cell(row=row_num+4,column=(8+field_index)).value = (np.min(code_df.loc[:, field].to_numpy()))
                ws.cell(row=row_num+5,column=(8+field_index)).value = (np.median(code_df.loc[:, field].to_numpy()))

                if field_index <=2 and box_and_whisker == True:
                    box_plot_dict[field].append(code_df.loc[:, field].to_numpy())

            ws.cell(row_num + 6, column=8).value = float(code_df.shape[0]/total_rows)*100 #Calculates % of XS with the given landform designation

            row_num += 8

        wb.save(stage_stat_xl)
        print("Landform stats for stage %sft completed..." % stage)

        if box_and_whisker==True:
            directory = os.path.dirname(stage_stat_xl)
            plot_dirs = directory + "\\Box_plots"
            if not os.path.exists(plot_dirs):
                os.makedirs(plot_dirs)

            for field in list_of_fields[:3]:
                fig, ax = plt.subplots()
                ax.set_title('Stage %s ft, %s boxplots by landform' % (str(int(stage)), field))
                ax.set_xlabel('Landform')
                if field != 'W_s_Z_s':
                    ax.set_ylabel('US feet')
                else:
                    ax.set_ylabel('C(Ws,Zs)')
                plt.boxplot(box_plot_dict[field], patch_artist=True, labels=['Oversized', 'Const. Pool', 'Normal', 'Wide riffle', 'Nozzle'])
                plt.savefig(plot_dirs + ('\\stage_%sft_%s_boxplot.png' % (stage, field)), dpi=400,bbox_inches='tight')
                plt.close(fig)

    print("All descriptive stats completed!")

def compare_flows(stages_stats_xl_dict, max_stage,save_plots=False):
    list_of_lists = [[],[],[]] #W, Z, C(Ws,Zs), used to make line plots of mean values vs stage
    list_of_landforms = [[],[],[],[],[]] #-2,-1,0,1,2

    for stage in range(1, max_stage + 1):
        stage_stat_xl = stages_stats_xl_dict['Stage_%sft' % stage]

        wb = xl.load_workbook(stage_stat_xl)
        ws = wb.active
        list_of_lists[0].append(float(ws.cell(row=2, column=2).value))
        list_of_lists[1].append(float(ws.cell(row=2, column=3).value))
        list_of_lists[2].append(float(ws.cell(row=2, column=4).value))

        for num in range(len(list_of_landforms)):
            row_num = 8 + (8*num)
            list_of_landforms[num].append(float(ws.cell(row=row_num, column=8).value))
        wb.close()

        directory = os.path.dirname(stage_stat_xl)
        plot_dirs = directory + "\\Line_plots"
        if not os.path.exists(plot_dirs):
            os.makedirs(plot_dirs)

    x_values = np.arange(start=1,stop=max_stage+1,step=1) #Setting up subplots showing W, Z, and C(W,Z) vs stage

    ax1 = plt.subplot(311)
    plt.plot(x_values, np.array(list_of_lists[0]),color='g')
    plt.ylabel('Mean width (US ft)')
    plt.ylim(0, np.max(np.array(list_of_lists[0])))
    plt.setp(ax1.get_xticklabels(), visible=False)
    plt.grid(True)

    ax2 = plt.subplot(312, sharex=ax1)
    plt.plot(x_values,np.array(list_of_lists[1]),color='m')
    plt.ylabel('Mean detrended Z (US ft)')
    plt.ylim(0, np.max(np.array(list_of_lists[1])))
    plt.setp(ax2.get_xticklabels(), visible=False)
    plt.grid(True)

    ax3 = plt.subplot(313, sharex=ax1)
    plt.plot(x_values,np.array(list_of_lists[2]),color='darkorange')
    plt.ylabel('Mean C(Ws,Zs)')
    plt.xlabel("Flood stage height (US ft)")
    plt.ylim(np.min(np.array(list_of_lists[2])), np.max(np.array(list_of_lists[2])))
    plt.setp(ax3.get_xticklabels(), fontsize=12)
    ax3.xaxis.set_major_locator(MaxNLocator(nbins=40,integer=True))
    plt.grid(True)

    if save_plots == False:
        plt.show()
        plt.cla()
    else:
        fig = plt.gcf()
        fig.set_size_inches(12, 6)
        plt.savefig(plot_dirs + "\\w_z_csz_stage_plot", dpi=300, bbox_inches='tight')
        plt.cla()
        print("W, Z, C(W,Z) plot vs flood stage saved @ %s" % plot_dirs)

    x_values = np.arange(start=1, stop=max_stage + 1, step=1) #Setting up lineplots of landform abundance vs stage

    list_of_land_labels = ['Oversized', 'Const. Pool', 'Normal', 'Wide Riffle', 'Nozzle']
    list_of_land_colors = ['navy', 'orange', 'c', 'grey', 'gold']

    fig, ax = plt.subplots()

    for landform in list_of_landforms:
        landform_index = list_of_landforms.index(landform)
        ax.plot(x_values,np.array(landform),color=list_of_land_colors[landform_index],label=list_of_land_labels[landform_index])

    ax.set_xlabel('Flood stage height (US ft)')
    ax.set_ylabel('% Abundance')
    ax.set_title("Landform abundance vs. flood stage height")
    ax.legend()
    ax.set_ylim(0,100)
    ax.set_xlim(1,max_stage)
    plt.setp(ax.get_xticklabels(), fontsize=12)
    ax.xaxis.set_major_locator(MaxNLocator(nbins=40, integer=True))
    plt.grid(True)

    if save_plots == False:
        plt.show()
        plt.cla()
    else:
        fig = plt.gcf()
        fig.set_size_inches(12, 6)
        plt.savefig(plot_dirs + "\\Landform_abundance", dpi=300, bbox_inches='tight')
        plt.cla()
        print("Landform abundace plot saved @ %s" % plot_dirs)




#INPUTS#
table_directory = r"C:\Users\xavierrn\Documents\RESEARCH\test_csv_folder" #A folder with stage csv files in it. Other files can occupy the directory as well.
##
out_list = analysis_setup(table_directory)
stages_dict = out_list[0]
stages_stats_xl_dict = out_list[1]
max_stage = out_list[2]
stats_table_location = out_list[3]

stage_level_descriptive_stats(stages_dict,stages_stats_xl_dict,max_stage,box_and_whisker=False)
compare_flows(stages_stats_xl_dict, max_stage,save_plots=True)



