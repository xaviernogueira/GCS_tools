import scipy as sp
import scipy.signal as sig
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
from matplotlib.lines import Line2D
import matplotlib.tri as tri
import pandas as pd
import GCS_analysis
import openpyxl as xl
from openpyxl import Workbook
import os
from os import listdir
from os.path import isfile, join
import itertools
from itertools import combinations
import file_functions
from file_functions import *


def analysis_setup(table_director, key_zs=[]):
    '''This function takes a directory containing csv files with landform analysis for each flow, deletes all extra in the directory files and creates output xl sheets for statistical results.
    Args: Table directory, Returns: [stages_dict, stages_stats_xl_dict, GCS stats directory]. Stages_dict stores pandas data frames for each stage, stages_stats_xl_dict stores stage level xlsx tables.
    Keys for both dictionaries are StageXft where X is stage height'''
    print('Prepping csv file inputs...')
    list_of_files_in_out_folder = [f for f in listdir(table_directory) if isfile(join(table_directory, f))]
    csv_tables = []

    if len(key_zs) >= 1:
        ind = 0
        while list_of_files_in_out_folder[ind][-4:] != '.csv':
            ind += 1
        suffix = find_suffix(list_of_files_in_out_folder[ind])

        for z in key_zs:
            z_str = float_keyz_format(z)
            csv_tables.append(table_directory + "\\%s%s" % (z_str, suffix))

    else:
        for file in list_of_files_in_out_folder:
            if file[-4:] == ".csv":
                csv_tables.append(table_directory + "\\" + file)
            else:
                os.remove(table_directory + "\\" + file)

    print("List of tables ready to be formatted: %s" % csv_tables)
    stat_table_location = table_directory + "\\GCS_stat_tables_and_plots"
    if not os.path.exists(stat_table_location):
        os.makedirs(stat_table_location)

    stages_dict = {}
    stages_stats_xl_dict = {}
    stages = []
    max_stage = 0
    for table in csv_tables:
        base_name = os.path.basename(table)
        if base_name[1] == 'f':
            stage = int(base_name[0])
        elif base_name[2] == 'f':
            stage = int(base_name[:2])
        elif base_name[1] == 'p':
            stage = base_name[:3]
        elif base_name[2] == 'p':
            stage = base_name[:4]

        stages.append(stage)
        if  len(key_zs) == 0:
            if float(stage) > max_stage:
                max_stage = stage

        stage_stats_xl_name = (stat_table_location + '\\%sft_stats_table.xlsx' % stage)
        wb = xl.Workbook()
        wb.save(stage_stats_xl_name)
        wb.close()

        stage_df = pd.read_csv(table, na_values=[-9999])
        stage_df['code'].fillna(0, inplace=True)
        stages_dict['Stage_%sft' % stage] = stage_df
        stages_stats_xl_dict['Stage_%sft' % stage] = stage_stats_xl_name

    return [stages_dict, stages_stats_xl_dict, max_stage, stat_table_location, stages]

def stage_level_descriptive_stats(stages_dict, stages_stats_xl_dict, max_stage, stages, box_and_whisker=False):
    '''This function takes the stages_dict of pandas dataframes, and the stage_stats_xl dictionary, as well as calculated max stage as arguments.
    It fills out the xlsx file for each respective stage height with mean, std, max, min, and median values for the stage as a hole and each landform
    If the box_and_whisker parameter is set to True (not default), then box and whisker plots comparing the W, Z, and C(W,Z)'''
    for stage in stages:
        print("Writing descriptive stats for stage %sft" % stage)
        stage_df = stages_dict['Stage_%sft' % stage]
        stage_stat_xl = stages_stats_xl_dict['Stage_%sft' % stage]

        list_of_fields = ['W', 'Z', 'W_s_Z_s', 'W_s', 'Z_s']
        means = []
        stds = []
        high = []
        low = []
        medians = []
        for field in list_of_fields[:3]:  # add the mean and std_deviation of each field in list_of_fields to a list
            means.append(np.mean(stage_df.loc[:, field].to_numpy()))
            stds.append(np.std(stage_df.loc[:, field].to_numpy()))
            high.append(np.max(stage_df.loc[:, field].to_numpy()))
            low.append(np.min(stage_df.loc[:, field].to_numpy()))
            medians.append(np.median(stage_df.loc[:, field].to_numpy()))

        wb = xl.load_workbook(stage_stat_xl)
        ws = wb.active
        ws.title = 'Descriptive stats'

        ws.cell(row=2, column=1).value = 'MEAN'  # Setting up titles on xl
        ws.cell(row=3, column=1).value = 'STD'
        ws.cell(row=4, column=1).value = 'MAX'
        ws.cell(row=5, column=1).value = 'MIN'
        ws.cell(row=6, column=1).value = 'MEDIAN'

        for field in list_of_fields[:3]:  # add values in each field column
            field_index = int(list_of_fields.index(field))

            ws.cell(row=1, column=(2 + field_index)).value = field
            ws.cell(row=2, column=(2 + field_index)).value = means[field_index]
            ws.cell(row=3, column=(2 + field_index)).value = stds[field_index]
            ws.cell(row=4, column=(2 + field_index)).value = high[field_index]
            ws.cell(row=5, column=(2 + field_index)).value = low[field_index]
            ws.cell(row=6, column=(2 + field_index)).value = medians[field_index]

        wb.save(stage_stat_xl)
        print("Descriptive stats for W, Z, and W_s_W_Z_s saved for stage %sft..." % stage)

        list_of_codes = [-2, -1, 0, 1, 2]
        landform_dict = {-2: 'Oversized', -1: 'Constricted pool', 0: 'Normal', 1: 'Wide riffle', 2: 'Nozzle'}
        ws['F1'].value = '*Code: -2 for oversized, -1 for constricted pool, 0 for normal channel, 1 for wide riffle, and 2 for nozzle'
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['A'].width = 16

        w_codes_list = []  # initiate lists of arrays to store data for multiple box plot
        z_codes_list = []
        cwz_codes_list = []
        list_of_field_lists = [w_codes_list, z_codes_list, cwz_codes_list]
        box_plot_dict = dict(zip(list_of_fields, list_of_field_lists)) #  This has W,Z, and W_s_Z_s as keys referncing lists that will store the data for each subplot

        total_rows = len(stage_df.index)
        above_1_list = [0, 0, 0]  # 'W', 'Z', 'W_s_Z_s'
        above_half_list = [0, 0, 0]
        for index, row in stage_df.iterrows():
            if abs(row['W_s_Z_s']) >= stds[2]:
                above_1_list[2] += 1
                above_half_list[2] += 1
            elif abs(row['W_s_Z_s']) >= (0.5*stds[2]):
                above_half_list[2] += 1

            for field in list_of_fields[3:]:  # List splice: ['W_s', 'Z_s']
                field_index = int(list_of_fields[3:].index(field))
                if abs(row[field]) >= 1:
                    above_1_list[field_index] += 1
                    above_half_list[field_index] += 1
                elif abs(row[field]) >= 0.5:
                    above_half_list[field_index] += 1

        ws.cell(row=7, column=1).value = "% >= 0.5 STD"
        ws.cell(row=8, column=1).value = "% >= 1 STD"

        for index in range(len(above_1_list)):  # Calculates % of W, Z, and W_s_Z_s that are greater than 0.5 and 1 of their standard deviations
            percent_above_half = float((above_half_list[index]/total_rows)*100)
            percent_above_1 = float((above_1_list[index]/total_rows)*100)
            ws.cell(row=7, column=(2 + index)).value = percent_above_half
            ws.cell(row=8, column=(2 + index)).value = percent_above_1

        row_num = 2
        for code in list_of_codes:  # Calculating same descriptive stats for each landform, each table is spaced 7 cells apart
            code_df = stage_df.loc[stage_df['code'] == code, ['dist_down', 'W', 'W_s', 'Z', 'Z_s', 'W_s_Z_s']]
            ws.cell(row=row_num, column=7).value = (str(landform_dict[code]))  # Preparing the table
            ws.cell(row=row_num + 1, column=7).value = 'MEAN'
            ws.cell(row=row_num + 2, column=7).value = 'STD'
            ws.cell(row=row_num + 3, column=7).value = 'MAX'
            ws.cell(row=row_num + 4, column=7).value = 'MIN'
            ws.cell(row=row_num + 5, column=7).value = 'MEDIAN'
            ws.cell(row_num + 6, column=7).value = '% Abundance:'

            if len(code_df.index) == 0:
                for field in list_of_fields:
                    field_index = int(list_of_fields.index(field))
                    ws.cell(row=row_num, column=(8 + field_index)).value = str(field)
                    ws.cell(row=row_num + 1, column=(8 + field_index)).value = 0
                    ws.cell(row=row_num + 2, column=(8 + field_index)).value = 0
                    ws.cell(row=row_num + 3, column=(8 + field_index)).value = 0
                    ws.cell(row=row_num + 4, column=(8 + field_index)).value = 0
                    ws.cell(row=row_num + 5, column=(8 + field_index)).value = 0

                    if field_index <= 2 and box_and_whisker == True:
                        box_plot_dict[field].append(code_df.loc[:, field].to_numpy())

            else:
                for field in list_of_fields:
                    field_index = int(list_of_fields.index(field))
                    ws.cell(row=row_num, column=(8 + field_index)).value = str(field)
                    ws.cell(row=row_num + 1, column=(8 + field_index)).value = (
                        np.mean(code_df.loc[:, field].to_numpy()))
                    ws.cell(row=row_num + 2, column=(8 + field_index)).value = (
                        np.std(code_df.loc[:, field].to_numpy()))
                    ws.cell(row=row_num + 3, column=(8 + field_index)).value = (
                        np.max(code_df.loc[:, field].to_numpy()))
                    ws.cell(row=row_num + 4, column=(8 + field_index)).value = (
                        np.min(code_df.loc[:, field].to_numpy()))
                    ws.cell(row=row_num + 5, column=(8 + field_index)).value = (
                        np.median(code_df.loc[:, field].to_numpy()))

                    if field_index <= 2 and box_and_whisker == True:
                        box_plot_dict[field].append(code_df.loc[:, field].to_numpy())

            ws.cell(row_num + 6, column=8).value = float(code_df.shape[0]/total_rows)*100  # Calculates % of XS with the given landform designation

            row_num += 8

        wb.save(stage_stat_xl)
        print("Landform stats for stage %sft completed..." % stage)

        if box_and_whisker==True:
            directory = os.path.dirname(stage_stat_xl)
            plot_dirs = directory + "\\Box_plots"
            if not os.path.exists(plot_dirs):
                os.makedirs(plot_dirs)

            for field in list_of_fields[:3]:
                fig, ax = plt.subplots()
                ax.set_title('Stage %s ft, %s boxplots by landform' % (str(int(stage)), field))
                ax.set_xlabel('Landform')
                if field != 'W_s_Z_s':
                    ax.set_ylabel('US feet')
                else:
                    ax.set_ylabel('C(Ws,Zs)')
                plt.boxplot(box_plot_dict[field], patch_artist=True, labels=['Oversized', 'Const. Pool', 'Normal', 'Wide riffle', 'Nozzle'])
                plt.savefig(plot_dirs + ('\\stage_%sft_%s_boxplot.png' % (stage, field)), dpi=400,bbox_inches='tight')
                plt.close(fig)

    print("All descriptive stats completed!")

def compare_flows(stages_stats_xl_dict, max_stage,save_plots=False):
    list_of_lists = [[],[],[]] # W, Z, C(Ws,Zs), used to make line plots of mean values vs stage
    list_of_landforms = [[],[],[],[],[]] # -2,-1,0,1,2
    wz_corr_lists = [[], [], [], [], []]  # Pearson correlation coefficients for Ws and Zs for each increasing flood stage for landforms [-2,-1,0,1,2]

    for stage in range(1, max_stage + 1):
        stage_stat_xl = stages_stats_xl_dict['Stage_%sft' % stage]

        wb = xl.load_workbook(stage_stat_xl)
        ws = wb.active
        list_of_lists[0].append(float(ws.cell(row=2, column=2).value))
        list_of_lists[1].append(float(ws.cell(row=2, column=3).value))
        list_of_lists[2].append(float(ws.cell(row=2, column=4).value))

        for num in range(len(list_of_landforms)):
            row_num1 = 8 + (8*num)
            row_num2 = 3 + (8*num)
            list_of_landforms[num].append(float(ws.cell(row=row_num1, column=8).value))
            wz_corr_lists[num].append(float(ws.cell(row_num2, column=10).value))
        wb.close()

        directory = os.path.dirname(stage_stat_xl)
        plot_dirs = directory + "\\Line_plots"
        if not os.path.exists(plot_dirs):
            os.makedirs(plot_dirs)

    x_values = np.arange(start=1, stop=max_stage+1, step=1) #  Setting up subplots showing W, Z, and C(W,Z) vs stage

    ax1 = plt.subplot(311)
    plt.plot(x_values, np.array(list_of_lists[0]), color='g')
    plt.ylabel('Mean width (US ft)')
    plt.ylim(0, np.max(np.array(list_of_lists[0])))
    plt.setp(ax1.get_xticklabels(), visible=False)
    plt.grid(True)

    ax2 = plt.subplot(312, sharex=ax1)
    plt.plot(x_values,np.array(list_of_lists[1]), color='m')
    plt.ylabel('Mean detrended Z (US ft)')
    plt.ylim(0, np.max(np.array(list_of_lists[1])))
    plt.setp(ax2.get_xticklabels(), visible=False)
    plt.grid(True)

    ax3 = plt.subplot(313, sharex=ax1)
    plt.plot(x_values, np.array(list_of_lists[2]), color='darkorange')
    plt.ylabel('Mean C(Ws,Zs)')
    plt.xlabel("Flood stage height (US ft)")
    plt.ylim(np.min(np.array(list_of_lists[2])), np.max(np.array(list_of_lists[2])))
    plt.setp(ax3.get_xticklabels(), fontsize=12)
    ax3.xaxis.set_major_locator(MaxNLocator(nbins=40,integer=True))
    plt.grid(True)

    if save_plots == False:
        plt.show()
        plt.cla()
    else:
        fig = plt.gcf()
        fig.set_size_inches(12, 6)
        plt.savefig(plot_dirs + "\\w_z_csz_stage_plot", dpi=300, bbox_inches='tight')
        plt.cla()
        print("W, Z, C(W,Z) plot vs flood stage saved @ %s" % plot_dirs)

    x_values = np.arange(start=1, stop=max_stage + 1, step=1) #Setting up lineplots of landform abundance vs stage

    list_of_land_labels = ['Oversized', 'Const. Pool', 'Normal', 'Wide Riffle', 'Nozzle']
    list_of_land_colors = ['navy', 'orange', 'c', 'grey', 'gold']

    fig, ax = plt.subplots()

    for landform in list_of_landforms:
        landform_index = list_of_landforms.index(landform)
        ax.plot(x_values,np.array(landform),color=list_of_land_colors[landform_index],label=list_of_land_labels[landform_index])

    ax.set_xlabel('Flood stage height (US ft)')
    ax.set_ylabel('% Abundance')
    ax.set_title("Landform abundance vs. flood stage height")
    ax.legend()
    ax.set_ylim(0,100)
    ax.set_xlim(1,max_stage)
    plt.setp(ax.get_xticklabels(), fontsize=12)
    ax.xaxis.set_major_locator(MaxNLocator(nbins=40, integer=True))
    plt.grid(True)

    if save_plots == False:
        plt.show()
        plt.cla()
    else:
        fig = plt.gcf()
        fig.set_size_inches(12, 6)
        plt.savefig(plot_dirs + "\\Landform_abundance", dpi=300, bbox_inches='tight')
        plt.cla()
        print("Landform abundace plot saved @ %s" % plot_dirs)

    x_values = np.arange(start=1, stop=max_stage + 1, step=1)  # Setting up mean C(Ws,Zs) for each landform plotted vs flood stage height

    fig, ax = plt.subplots()

    for landform in wz_corr_lists:
        landform_index = wz_corr_lists.index(landform)
        ax.plot(x_values, np.array(landform), color=list_of_land_colors[landform_index],
                label=list_of_land_labels[landform_index])

    ax.set_xlabel('Flood stage height (US ft)')
    ax.set_ylabel('C(Ws,Zs)')
    ax.set_title("Landform mean C(Ws,Zs) vs. flood stage height")
    ax.legend()
    ax.set_xlim(1, max_stage)
    plt.setp(ax.get_xticklabels(), fontsize=12)
    ax.xaxis.set_major_locator(MaxNLocator(nbins=40, integer=True))
    plt.grid(True)

    if save_plots == False:
        plt.show()
        plt.cla()
    else:
        fig = plt.gcf()
        fig.set_size_inches(12, 6)
        plt.savefig(plot_dirs + "\\Landform_cwz_plot", dpi=300, bbox_inches='tight')
        plt.cla()
        print("Landform C(Ws,Zs) vs. flood stage plot saved @ %s" % plot_dirs)

def autocorr_and_powerspec(stages_dict, stages_stats_xl_dict, max_stage, save_plots=False):
    # Set up autocorrelation subplots
    plt.rcParams.update({'figure.max_open_warning': 0})

    for stage in range(1, max_stage+1):
        fig, ax = plt.subplots(3, 1, sharex=True, sharey=True)
        stage_stat_xl = stages_stats_xl_dict['Stage_%sft' % stage]
        directory = os.path.dirname(stage_stat_xl)  # Define folder for autocorrlation spot
        plot_dirs = directory + "\\Autocorrelation_plots"
        if not os.path.exists(plot_dirs):
            os.makedirs(plot_dirs)

        stage_df = stages_dict['Stage_%sft' % stage]
        stage_df = stage_df.sort_values(by=['dist_down'])
        stage_df = stage_df.reset_index(drop=True)
        ws = stage_df['W_s']
        zs = stage_df['Z_s']
        cwz = stage_df['W_s_Z_s']
        dist_down = stage_df['dist_down']
        spacing = abs(dist_down[1] - dist_down[0])
        maxlags = int(len(dist_down) / 2)
        lags, lower_white, upper_white = white_noise_acf_ci(ws, maxlags=maxlags)

        lags, ar1_acorrs, lower_red, upper_red = ar1_acorr(ws, maxlags=maxlags)
        ax[0].plot(*cox_acorr(ws, maxlags=maxlags))
        ax[0].plot(lags, ar1_acorrs, color='red')
        ax[0].plot(lags, lower_red, '--', color='salmon')
        ax[0].plot(lags, upper_red, '--', color='salmon')
        ax[0].plot(lags, lower_white, '--', color='grey')
        ax[0].plot(lags, upper_white, '--', color='grey')
        ax[0].set_ylabel(r'Ws' + ' Autocorrelation')

        lags, ar1_acorrs, lower_red, upper_red = ar1_acorr(zs, maxlags=maxlags)
        ax[1].plot(*cox_acorr(zs, maxlags=maxlags))
        ax[1].plot(lags, ar1_acorrs, color='red')
        ax[1].plot(lags, lower_red, '--', color='salmon')
        ax[1].plot(lags, upper_red, '--', color='salmon')
        ax[1].plot(lags, lower_white, '--', color='grey')
        ax[1].plot(lags, upper_white, '--', color='grey')
        ax[1].set_ylabel(r'Zs Autocorrelation')

        lags, ar1_acorrs, lower_red, upper_red = ar1_acorr(cwz, maxlags=maxlags)
        ax[2].plot(*cox_acorr(cwz, maxlags=maxlags))
        ax[2].plot(lags, ar1_acorrs, color='red')
        ax[2].plot(lags, lower_red, '--', color='salmon')
        ax[2].plot(lags, upper_red, '--', color='salmon')
        ax[2].plot(lags, lower_white, '--', color='grey')
        ax[2].plot(lags, upper_white, '--', color='grey')
        ax[2].set_ylabel('C(Ws,Zs) Autocorrelation')

        ax[0].set_title('Stage %sft autocorrelation' % stage)
        ax[2].set_xlabel('Lag (US ft)')
        for j in range(3):
            ax[j].grid()
            ax[j].set_xlim(0, maxlags)
            ticks = map(int, ax[j].get_xticks() * spacing)
            ax[j].set_xticklabels(ticks)
            fig.set_size_inches(12, 6)
            plt.grid(True,which='both')
        if save_plots == False:
            plt.show()
            plt.cla()
        else:
            fig = plt.gcf()
            fig.set_size_inches(12, 6)
            plt.savefig(plot_dirs + ("\\Autocorr_plot_stage%sft" % stage), dpi=300, bbox_inches='tight')
            plt.cla()
    print("Autocorrelation plots created for each stage @ %s" % plot_dirs)

# Autocorrelation Heat Map plotting
    x = []
    y = []
    z = []
    lower_whites = []
    upper_whites = []
    for stage in range(1, max_stage + 1):
        stage_stat_xl = stages_stats_xl_dict['Stage_%sft' % stage]
        directory = os.path.dirname(stage_stat_xl)  # Define folder for autocorrlation spot
        plot_dirs = directory + "\\Autocorrelation_plots"
        if not os.path.exists(plot_dirs):
            os.makedirs(plot_dirs)

        stage_df = stages_dict['Stage_%sft' % stage]
        stage_df = stage_df.sort_values(by=['dist_down'])
        stage_df = stage_df.reset_index(drop=True)
        ws = stage_df['W_s']
        zs = stage_df['Z_s']
        cwz = stage_df['W_s_Z_s']
        dist_down = stage_df['dist_down']
        spacing = abs(dist_down[1] - dist_down[0])
        maxlags = int(len(dist_down) / 2)
        lags, lower_white, upper_white = white_noise_acf_ci(cwz, maxlags=maxlags)
        lags, acorrs = cox_acorr(cwz, maxlags=maxlags)
        # only include positive autocorrelations
        for j, acorr_val in enumerate(acorrs):
            if acorr_val > 0:
                x.append('%s' % stage)
                y.append(lags[j])
                z.append(acorrs[j])
                lower_whites.append(lower_white[j])
                upper_whites.append(upper_white[j])

    triang = tri.Triangulation(x, y)
    # mask where acorr (z) is below the white noise threshold at that lag (y)
    mask = []
    for triangle in triang.triangles:
        z_vals = [z[vertex] for vertex in triangle]
        lws = [lower_whites[vertex] for vertex in triangle]
        uws = [upper_whites[vertex] for vertex in triangle]
        cond_1 = all(z_val > lw for z_val, lw in zip(z_vals, lws))
        cond_2 = all(z_val < uw for z_val, uw in zip(z_vals, uws))
        if cond_1 and cond_2:
            mask_val = 1
        else:
            mask_val = 0
        mask.append(mask_val)
    triang.set_mask(mask)

    fig, ax = plt.subplots(1, 1)
    fig.set_size_inches(12, 6)
    tpc = ax.tripcolor(triang, z, cmap='jet')
    ax.set(xlabel='Flood stage height (US ft)', ylabel='Lag (US ft)')
    qlabels = range(1, max_stage + 1)
    xticks = [1*q for q in qlabels]
    ax.set_xticks(xticks)
    xticklabels = [q for q in qlabels]
    ax.set_xticklabels(xticklabels)
    plt.xticks(rotation=-45)
    yticklabels = map(int, ax.get_yticks() * spacing)
    ax.set_yticklabels(yticklabels)
    fig.colorbar(tpc)
    plt.grid(True, which='both')
    if save_plots == False:
        plt.show()
        plt.cla()
    else:
        fig = plt.gcf()
        fig.set_size_inches(12, 6)
        plt.savefig(plot_dirs + ("\\Autocorr_heat_plot"), dpi=300, bbox_inches='tight')
        plt.cla()
    print("Autocorrelation heat plot created @ %s" % plot_dirs)

    # Power spectral density plotting
    x2 = [] #Stores flood stage heights
    y2 = [] #Stores frequencies
    z2 = [] #Stores power spectral density for a given frequency at a given stage height
    lower_whites = []
    upper_whites = []
    corr_list = [] #Stores C(Ws,Zs) for each stage
    qs = [float(stage) for stage in range(1,max_stage+1)]
    for stage in range(1, max_stage + 1):
        stage_stat_xl = stages_stats_xl_dict['Stage_%sft' % stage]
        directory = os.path.dirname(stage_stat_xl)
        plot_dirs = directory + "\\Autocorrelation_plots"
        if not os.path.exists(plot_dirs):
            os.makedirs(plot_dirs)

        stage_df = stages_dict['Stage_%sft' % stage]
        stage_df = stage_df.sort_values(by=['dist_down'])
        stage_df = stage_df.reset_index(drop=True)
        ws = stage_df['W_s']
        zs = stage_df['Z_s']
        cwz = stage_df['W_s_Z_s']
        dist_down = stage_df['dist_down']

        corr = np.corrcoef(ws, zs)[0][1]
        corr_list.append(corr)
        spacing = abs(dist_down[1] - dist_down[0])
        frequencies, psd = sig.periodogram(cwz, 1.0 / 3, window=sig.get_window('hamming', len(cwz)))
        index = stage - 1

        maxlags = int(len(dist_down) / 2)
        lags, lower_white, upper_white = white_noise_acf_ci(cwz, maxlags=maxlags)
        for j, psd_val in enumerate(psd):
            x2.append(qs[index])
            y2.append(frequencies[j])
            z2.append(psd[j])
            lower_whites.append(lower_white[j])
            upper_whites.append(upper_white[j])
    std_z = [z_val * 1.0 / np.std(z2) for z_val in z2]
    qs, corr_list = zip(*sorted(zip(qs, corr_list)))
    triang = tri.Triangulation(x2, y2)

    mask = []
    for triangle in triang.triangles:
        z_vals = [z2[vertex] for vertex in triangle]
        lws = [lower_whites[vertex] for vertex in triangle]
        uws = [upper_whites[vertex] for vertex in triangle]
        cond_1 = all(z_val > lw for z_val, lw in zip(z_vals, lws))
        cond_2 = all(z_val < uw for z_val, uw in zip(z_vals, uws))
        if cond_1 and cond_2:
            mask_val = 1
        else:
            mask_val = 0
        mask.append(mask_val)
    triang.set_mask(mask)

    fig, ax = plt.subplots(1, 1)
    fig.set_size_inches(12, 6)
    tpc = ax.tripcolor(triang, std_z, cmap='jet')
    ax.set(xlabel='Flood stage height (US ft)', ylabel='Spatial Frequency (cycles/ft)')
    qlabels = range(1, max_stage+1)  # show less discharge ticks to make plot less cluttered
    xticks = [1*q for q in qlabels]
    ax.set_xticks(xticks)
    xticklabels = [('%s' % q) for q in qlabels]
    ax.set_xticklabels(xticklabels)
    plt.xticks(rotation=-45)
    ax.set_ylim(min(y2), 0.01)
    fig.colorbar(tpc)
    plt.grid(True, which='both')
    if save_plots == False:
        plt.show()
        plt.cla()
    else:
        fig = plt.gcf()
        fig.set_size_inches(12, 6)
        plt.savefig(plot_dirs + ("\\PSD_plot.png"), dpi=300, bbox_inches='tight')
        plt.cla()
    print("Power spectral density plot created @ %s" % plot_dirs)


#INPUTS#
GCS_process_on=False

if GCS_process_on=True:
    sc_class = 2
    comid = 17586810
    direct = (r"Z:\users\xavierrn\SoCoast_Final_ResearchFiles\SCO%s\COMID%s" % (sc_class, comid))
    out_folder = direct + r'\LINEAR_DETREND'


    table_directory = direct + '\\LINEAR_DETREND\\gcs_ready_tables'

    out_list = analysis_setup(table_directory, key_zs=[0.6, 3.6, 8.1])
    stages_dict = out_list[0]
    stages_stats_xl_dict = out_list[1]
    max_stage = out_list[2]
    stats_table_location = out_list[3]
    stages = out_list[4]

    stage_level_descriptive_stats(stages_dict, stages_stats_xl_dict, max_stage, stages=stages, box_and_whisker=False)
    #compare_flows(stages_stats_xl_dict, max_stage, save_plots=True)
    #autocorr_and_powerspec(stages_dict, stages_stats_xl_dict, max_stage, save_plots=True)




