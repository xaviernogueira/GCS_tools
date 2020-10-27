import openpyxl as xl
import arcpy
import os as os
import pandas as pd
from arcpy import env
from openpyxl import Workbook
from openpyxl import load_workbook
import matplotlib as mpl
from matplotlib import pyplot as plt
import numpy as np
import csv


def prep_xl_file(xyz_table_location, listofcolumn=['LOCATION', 'POINT_X', 'POINT_Y', 'Value']):
    listoflist = [None, None, None, None]

    if xyz_table_location[-4:] == ".csv":
        elevation_df = pd.read_csv(xyz_table_location)
        for j, header in enumerate(listofcolumn):
            listoflist[j] = elevation_df.loc[:, [(header)]].squeeze().to_numpy()

    else:
        listofcolumn = ['E', 'K', 'L', 'N']
        wb = load_workbook(xyz_table_location)
        ws = wb.active
        print("Workbook %s loaded" % xyz_table_location)

        for i, col in enumerate(listofcolumn):
            temp_list = []
            for cell in ws[col]:
                temp_list.append(cell.value)
            del temp_list[0]
            listoflist[i] = np.array(temp_list)
        wb.save(xyz_table_location)

    location = np.int_(listoflist[0])
    z = np.around(listoflist[-1], 9)

    point_spacing = int(location[1]) - int(location[0])
    print("Point spacing: " + str(point_spacing))
    number_of_points = int(int(location[-1]) / int(point_spacing))

    print("Z array: %s" % z)

    return [location, z, xyz_table_location]


def quadratic_fit(location_np, location, z_np, ws):
    print("Applying quadratic fit...")
    constants = np.polyfit(location_np, z_np, 2)

    fit_params = []
    for i in constants:
        fit_params.append(i)
    print(fit_params)

    #Use the found regresssion down channel
    fitted_z = []
    for i in location:
        z_fit_temp = (fit_params[0]*i**(2) + fit_params[1]*i + fit_params[2])
        fitted_z.append(z_fit_temp)

    print(fitted_z)

    residual = []
    if len(fitted_z) == len(z):
        print("Lists are same length, continue as planned...")
        for i in range(len(fitted_z)):
            residual.append(z[i] - fitted_z[i])

    else:
        print("Lists are wrong lengths/not matching, something is wrong...")

    mean_z = (sum(z) / len(z))
    squared_real = []
    squared_res = []

    #Calculate the total sum of squares, sum of residuals, and R^2
    for i in range(len(residual)):
        squared_real.append((z[i] - mean_z)**2)
        squared_res.append(residual[i]**2)

    print("List of squared variance from mean: " + str(squared_real))
    print("List of squared residuals: " + str(squared_res))

    R_squared = 1 - (sum(squared_res)/sum(squared_real))
    print("The coefficient of determination is: " + str(R_squared))

    # Calculate mean and SD for residual to calculate residual z scores as a list
    mean_res = sum(residual)/len(residual)
    sd_res = np.std(residual)
    res_z_score = [(z_res_value - mean_res) * 1.0 / sd_res for z_res_value in residual]
    print(residual)
    print("The mean residual is %.2f" %mean_res)

    # Add fitted z values to the xyz table
    cell_test = ws["D1"]
    print(cell_test.value)
    cell_test.value = "Quadratic fit z values"
    print(cell_test.value)

    if ws["D1"].value == "Quadratic fit z values":
        for i in range(2, len(fitted_z)):
            cell = ws.cell(row=i, column=4)
            cell.value = float(fitted_z[i])
    else:
        print("Something is wrong with writing to the excel sheet")

    wb.save(filename=xyz_table)

    print("Excel file ready for Arc processing!")


def linear_fit(location, z, xyz_table_location, list_of_breakpoints=[], transform=0, chosen_fit_index=[]):
    # Applies a linear fit to piecewise sections of the longitudinal profile, each piece is stored in split_list

    print("Applying linear fit...")

    if len(list_of_breakpoints) != 0:
        list_of_breakpoints.insert(0, 0)
        list_of_breakpoints.append(int(location[-1]))
        print("Breakpoints imported...")
    else:
        print("No breakpoint imported...")

    split_location_list = []
    split_z_list = []

    point_spacing = int(location[1]) - int(location[0])  # Split by breakpoints into a list of lists

    location = np.int_(location)  # Format input numpy arrays
    z = np.float_(z)
    z = np.around(z, 9)  # Round z to 9 decimal points

    if len(list_of_breakpoints) > 0:
        fit_params = []

        slope_break_indices = [int(int(distance)/int(point_spacing)) for distance in list_of_breakpoints]

        for i in slope_break_indices[1:]:
            temp_location_list = []
            temp_z_list = []
            temp_break_index = i
            index = slope_break_indices.index(temp_break_index)

            if index == 1:
                split_location_list.append(location[:temp_break_index])
                split_z_list.append(z[:temp_break_index])
            elif i != slope_break_indices[-1]:
                temp_location_list = location[slope_break_indices[index-1]:temp_break_index]
                split_location_list.append(temp_location_list)
                split_z_list.append(z[slope_break_indices[index-1]:temp_break_index])
            elif i == slope_break_indices[-1]:
                split_location_list.append(location[slope_break_indices[index-1]:])
                split_z_list.append(z[slope_break_indices[index-1]:])

        print("Breakpoints added...")

        if len(split_z_list) == len(split_location_list):  # Get fit parameters for each section of the data
            for i in range(len(split_location_list)):
                temp_loc_array_unformatted = np.array(split_location_list[i])
                temp_loc_array = np.int_(temp_loc_array_unformatted)
                temp_z_array_unformatted = np.array(split_z_list[i])
                temp_z_array = np.float_(temp_z_array_unformatted)

                m, b = np.polyfit(temp_loc_array, temp_z_array, 1)
                fit_params.append([m, b])
            print("Fit param list: " + str(fit_params))
        else:
            print("Something went wrong, list lengths do not match...")

        # Make list of lists storing fitted z values
        list_of_lengths = []
        z_fit_list = []
        for i in range(len(split_z_list)):
            list_of_lengths.append(len(split_z_list[i]))

        # Add fitted z's into a list
        print(list_of_lengths)
        i = 0

        while i < len(list_of_lengths):
            if len(chosen_fit_index) == 0:
                index = i
            else:
                index = int(chosen_fit_index[0])
            for j in range(list_of_lengths[i]):
                z_fit_list.append((split_location_list[i][j] * fit_params[index][0] + fit_params[index][1]) + float(transform))
            i += 1

        if len(z_fit_list) == len(z):
            print("List lengths are compatible, next we export to excel")
        else:
            print("Something went wrong, length of z =/ z_fit_list")
        print(z_fit_list)

    else:
        m, b = np.polyfit(location, z, 1)
        fit_params = [[m, b]]
        print("Fit params [m, b]: " + str(fit_params))

        z_fit_list = []

        location_list = location.tolist()
        for j in location_list:
            z_fit_list.append((j*fit_params[0][0]+fit_params[0][1]) + float(transform))

    #Calculate residual and R^2
    residual = []
    for i in range(len(z_fit_list)):
        residual.append(z[i]-z_fit_list[i])
    mean_z = (sum(z) / len(z))
    squared_real = []
    squared_res = []

    #Calculate the total sum of squares, sum of residuals, and R^2
    for i in range(len(residual)):
        squared_real.append((z[i] - mean_z)**2)
        squared_res.append(residual[i]**2)

    R_squared = 1 - (sum(squared_res)/sum(squared_real))
    print("The coefficient of determination is: " + str(R_squared))

    # Calculate mean and SD for residual to calculate residual z scores as a list
    mean_res = sum(residual)/len(residual)
    sd_res = np.std(residual)
    res_z_score = [(z_res_value - mean_res) * 1.0 / sd_res for z_res_value in residual]

    if xyz_table_location[-4:] == 'xlsx':  # Add fitted z values to the xyz table
        wb = load_workbook(xyz_table_location)
        ws = wb.active
        cell_test = ws["F1"]
        cell_test.value = ("z_fit")

        if ws["F1"].value == cell_test.value:
            for i in z_fit_list:
                index = int(z_fit_list.index(i))
                row_index = index + 2
                cell = ws.cell(row=row_index, column=6)
                cell.value = float(z_fit_list[index])
        else:
            print("Something is wrong with writing to the excel sheet")

        ws["A1"].value == "OID"
        wb.save(filename=xyz_table_location)

    elif xyz_table_location[-4:] == '.csv':
        elevation_df = pd.read_csv(xyz_table_location)
        elevation_df['z_fit'] = np.array(z_fit_list)
        elevation_df.to_csv(xyz_table_location)

    print("Excel file ready for wetted-polygon processing!")

    return [fit_params, z_fit_list, residual, R_squared]

def moving_window_linear_fit(location, z, xyz_table_location, window_size):
    point_spacing = int(location[1]) - int(location[0]) #Calculate point spacing
    number_of_complete_windows = 0
    while ((number_of_complete_windows*window_size) + window_size) <= location[-1]:
        number_of_complete_windows += 1
    print("Total number of full %sft detrending windows is %s" % (window_size, number_of_complete_windows))

    if xyz_table_location[-3:] == 'csv': #Make sure we are accessing and xlsx file to use openpyxl module
        xyz_table_location = (xyz_table_location[:-3] + "xlsx")

    wb = load_workbook(xyz_table_location)
    ws = wb.active

    # Format input numpy arrays
    location = np.int_(location)
    z = np.float_(z)
    z = np.around(z, 9)  # Round z to 9 decimal points

    #Get window indice value
    window_breakpoints = []
    for num in range(0,number_of_complete_windows):
        window_breakpoints.append(int(num*window_size))
    window_breakpoints.remove(0)
    print("Window breakpoints @ %s" % window_breakpoints)
    window_breakpoint_indices = [int(f/point_spacing) for f in window_breakpoints]
    print("Window breakpoint indices @ %s" % window_breakpoint_indices)

    list_of_fit_params = []
    z_fit_list = []
    window_breakpoint_indices.sort()
    for bp in window_breakpoint_indices:
        if bp == window_breakpoint_indices[0]:
            split_loc_np = np.array(location[:bp])
            split_z_np = np.array(z[:bp])
            m, b = np.polyfit(split_loc_np, split_z_np, 1)
            list_of_fit_params.append([m, b])
            print("Fit params [m, b] for the first window of %sft are: %s" % (window_size, list_of_fit_params[0]))
            for i in location[:bp]:
                z_fit_list.append((i * m) + b) #add fit values to list
        elif bp != window_breakpoint_indices[-1]:
            index = int(window_breakpoint_indices.index(bp))
            last_index = window_breakpoint_indices[int(index - 1)]
            split_loc_np = np.array(location[last_index:bp])
            split_z_np = np.array(z[last_index:bp])
            m, b = np.polyfit(split_loc_np, split_z_np, 1)
            list_of_fit_params.append([m, b])
            print("Fit params [m, b] for window (%s to %s)ft are: %s" % ((point_spacing*window_breakpoint_indices[index-1]),(bp*point_spacing), list_of_fit_params[index]))
            for i in location[last_index:index]:
                z_fit_list.append((i * m) + b)
        elif bp == window_breakpoint_indices[-1]:
            split_loc_np = np.array(location[bp:])
            split_z_np = np.array(z[bp:])
            m, b = np.polyfit(split_loc_np, split_z_np, 1)
            list_of_fit_params.append([m, b])
            print("Fit params [m, b] for the last window from %sft ro the end are: %s" % ((bp*point_spacing), list_of_fit_params[-1]))
            for i in location[bp:]:
                z_fit_list.append((i * m) + b)

    print("Z fit list is ready for excel: %s" % z_fit_list)

    #residual = [location[i]-z_fit_list[i] for i in range(len(location))]

    # Add fitted z values to the xyz table
    cell_test = ws["F1"]
    print(cell_test.value)
    cell_test.value = ("z_fit_window%s" % window_size)
    print(cell_test.value)

    if ws["F1"].value == cell_test.value:
        print("Sheet activated...")
        for i in range(2, len(z_fit_list)): #index begins at two to ignore 0 (doesn't exist in xl) and 1 (xl titles)
            cell = ws.cell(row=i, column=6)
            cell.value = float(z_fit_list[i])
    else:
        print("Something is wrong with writing to the excel sheet")
    ws["A1"].value == "OID"

    wb.save(filename=xyz_table_location)

    print("Excel file ready for wetted-polygon processing!")

    return [z_fit_list]


def detrend_that_raster(detrend_location, fit_z_xl_file, original_dem, stage=0, window_size=0, list_of_breakpoints=[]):
    # Turn fitted xl file to a csv
    wb = xl.load_workbook(fit_z_xl_file)
    ws = wb.active
    arcpy.env.workspace = detrend_location
    arcpy.overwriteoutput = True
    spatial_ref = arcpy.Describe(original_dem).spatialReference
    arcpy.env.extent = arcpy.Describe(original_dem).extent

    if not os.path.exists(detrend_location):
        os.makedirs(detrend_location)

    if fit_z_xl_file[-4:] == 'xlsx':
        csv_name = fit_z_xl_file[:-5] + "_fitted.csv"
        with open(csv_name, 'w', newline="") as f:
            col = csv.writer(f)
            for row in ws.rows:
                col.writerow([cell.value for cell in row])

    elif fit_z_xl_file[-4] == '.csv':
        csv_name = fit_z_xl_file

    else:
        print('Invalid input table. Please re-save file as either an csv or xlsx')

    if window_size != 0:
        column = ('z_fit_window%s' % window_size)
    else:
        column = 'z_fit'

    if stage != 0:
        detrended_raster_file = detrend_location + ("\\rs_dt_s%s.tif" % stage)
    else:
        detrended_raster_file = detrend_location + "\\ras_detren.tif"
        print("0th stage marks non-stage specific centerline")
    points = arcpy.MakeXYEventLayer_management(csv_name, "POINT_X", "POINT_Y",
                                                   out_layer=("fitted_station_points_stage%s" % stage),
                                                   spatial_reference=spatial_ref, in_z_field=column)
    points = arcpy.SaveToLayerFile_management(points, ("fitted_station_points_stage%sft" % stage).replace('.csv', '.lyr'))
    points = arcpy.CopyFeatures_management(points)
    print("Creating Thiessen polygons...")

    fields = [f.name for f in arcpy.ListFields(points)]  # Delete non-relevent fields tp reduce errors
    dont_delete_fields = ['FID', 'Shape', 'POINT_X', 'POINT_Y', column]
    fields2delete = list(set(fields) - set(dont_delete_fields))
    points = arcpy.DeleteField_management(points, fields2delete)

    cell_size1 = arcpy.GetRasterProperties_management(DEM, "CELLSIZEX")
    cell_size = float(cell_size1.getOutput(0))
    thiessen = arcpy.CreateThiessenPolygons_analysis(points, "thiespoly_stage%s.shp" % stage, fields_to_copy='ALL')
    z_fit_raster = arcpy.PolygonToRaster_conversion(thiessen, column, ('theis_raster_stage%sft.tif' % stage),
                                                        cell_assignment="CELL_CENTER", cellsize=cell_size)
    detrended_DEM = arcpy.Raster(DEM) - arcpy.Raster(z_fit_raster)
    detrended_DEM.save(detrended_raster_file)
    print("DEM DETRENDED!")


###### Define plotting function ######
def diagnostic_quick_plot(location_np, z_np, xlim=0):
    x_plot = location_np
    y_plot = z_np
    plt.plot(x_plot, y_plot, 'r', label="Actual elevation profile")
    plt.xlabel("Thalweg distance downstream (ft)")
    plt.ylabel("Bed elevation (ft)")
    if xlim != 0:
        plt.xlim(0,xlim)
    else:
        plt.xlim(0,None)
    plt.grid(b=True, which='major', color='#666666', linestyle='-')
    plt.minorticks_on()
    plt.grid(b=True, which='minor', color='#999999', linestyle='-', alpha=0.2)

    return plt.show()


def make_quadratic_fit_plot(location_np, z_np, fit_params, stage=0, location=''):
    # Make a plot of the quadratic fit
    # Plot longitudinal elevation profile and detrending fit line
    x_plot = location_np
    y1_plot = z_np
    y2_plot = (fit_params[0]*x_plot**2 + fit_params[1]*x_plot + fit_params[2])
    plt.plot(x_plot, y1_plot, 'r', label="Actual elevation profile")
    plt.xlabel("Thalweg distance downstream (ft)")
    plt.ylabel("Bed elevation (ft)")
    plt.title("Quadratic detrended elevation profile")
    plt.plot(x_plot, y2_plot, 'b', label= "Elevation profile detrending quadratic: %s * x^2 + %.4f * x + %.4f" % (fit_params[0], fit_params[1], fit_params[2]))
    plt.grid(b=True, which='major', color='#666666', linestyle='-')
    plt.minorticks_on()
    plt.grid(b=True, which='minor', color='#999999', linestyle='-', alpha=0.2)
    plt.legend(loc=1)
    if stage == 0 and location == '':
        plt.show()
    else:
        fig = plt.gcf()
        fig.set_size_inches(12, 6)
        plt.savefig((location + '\\Stage_%s_quadratic_detrend_plot' % stage), dpi=300, bbox_inches='tight')
        plt.cla()

def make_linear_fit_plot(location_np, z_np, fit_params, stage=0, xmin=0, xmax=0, ymin=0, ymax=0, location='', transform=0):
    x_plot = location_np
    y1_plot = z_np
    y2_plots = []

    for list in fit_params:
        y2_plots.append((list[0]*x_plot + list[1])+float(transform))

    plt.plot(x_plot, y1_plot, 'r', label="Actual elevation profile")
    plt.xlabel("Thalweg distance downstream (ft)")
    plt.ylabel("Bed elevation (ft)")
    plt.title("Linear piecewise detrended elevation profile")
    for i in range(len(y2_plots)):
        plt.plot(x_plot, y2_plots[i], 'b', label="Elevation profile linear detrending: %.4f * x + %.4f" % (y2_plots[i][0], y2_plots[i][1]))
    if ymin != 0 and ymax == 0:
        plt.ylim(ymin, None)
    if ymax != 0 and ymin ==0:
        plt.ylim(None, ymax)
    if ymax != 0 and ymin != 0:
        plt.ylim(ymin, ymax)

    if xmin != 0 and xmax == 0:
        plt.xlim(xmin, None)
    if xmax != 0 and xmin ==0:
        plt.xlim(None, xmax)
    if xmax != 0 and xmin != 0:
        plt.xlim(xmin, xmax)

    plt.grid(b=True, which='major', color='#666666', linestyle='-')
    plt.minorticks_on()
    plt.grid(b=True, which='minor', color='#999999', linestyle='-', alpha=0.2)
    plt.legend(loc=1)

    if stage == 0 and location == '':
        plt.show()
    else:
        fig = plt.gcf()
        fig.set_size_inches(12, 6)
        plt.savefig((location + '\\Stage_%sft_linear_detrend_plot' % stage), dpi=300, bbox_inches='tight')
        plt.cla()

def make_residual_plot(location_np, residual, R_squared, stage=0, xmin=0, xmax=0, location=''):
    '''Plot residuals across longitudinal profile, show R^2. Inputs are a numpy array of location, a list of residuals, and a float for R-squared'''
    x_plot = location_np
    y_plot = np.array(residual)
    y_zero = 0*x_plot
    plt.scatter(x_plot, y_plot, s=1, c='r')
    plt.plot(x_plot, y_zero, c='b')
    plt.xlim(0, max(location_np))
    plt.ylim(min(y_plot), max(y_plot))
    plt.xlabel("Distance down stream centerline (ft)")
    plt.ylabel("Residual")
    plt.title("Residuals: R^2 = %.4f" % R_squared)
    plt.grid(b=True, which='major', color='#666666', linestyle='-')

    if xmin != 0 and xmax == 0:
        plt.xlim(xmin, None)
    if xmax != 0 and xmin == 0:
        plt.xlim(None, xmax)
    if xmax != 0 and xmin != 0:
        plt.xlim(xmin, xmax)

    plt.minorticks_on()
    plt.grid(b=True, which='minor', color='#999999', linestyle='-', alpha=0.2)

    if stage == 0 and location == '':
        plt.show()
    else:
        fig = plt.gcf()
        fig.set_size_inches(12, 6)
        plt.savefig((location + '\\Stage_%sft_linear_residual_plot' % stage), dpi=300, bbox_inches='tight')
        plt.cla()


################## CALL FUNCTIONS AS NECESSARY ####################
process_on = True
detrend_or_diagnostic = False  # False plots graphs to help make breakpoint decision, True saves plots and detrends the DEM.

###### INPUTS ######
# excel file containing xyz data for station points
comid = 17586552
SCO_number = '00_new_adds'

direct = (r"Z:\users\xavierrn\SoCoast_Final_ResearchFiles\SC%s\COMID%s" % (SCO_number, comid))
xyz_table = direct + '\\XYZ_elevation_table.csv'
centerline = direct + '\\las_files\\centerline\\smooth_centerline.shp'
DEM = direct + '\\las_files\\ls_nodt.tif'
process_footprint = direct + '\\las_footprint.shp'
detrend_workplace = direct + '\\LINEAR_DETREND'
spatial_ref = arcpy.Describe(process_footprint).spatialReference
######

breakpoints = [800, 1420, 2300]
transform_value = (0.0)  # Leave at 0.0
xlimits = [0, 0]  # [xmin, xmax] default is [0, 0]
ylimits = [0, 0]  # [ymin, ymax] default is [0, 0]
chosen_fit_index = []  # Allows one piecewise segment to be used for the whole DEM. Helpful with poor-centerline quality. Leave empty to have all included

if process_on == True:
    loc = prep_xl_file(xyz_table_location=xyz_table)[0]
    z = prep_xl_file(xyz_table_location=xyz_table)[1]
    ws = prep_xl_file(xyz_table_location=xyz_table)[2]

    if detrend_or_diagnostic==False:
        save_location = ''
        diagnostic_quick_plot(location_np=loc, z_np=z, xlim=0)
        fit_list = linear_fit(location=loc, z=z, xyz_table_location=xyz_table, list_of_breakpoints=breakpoints, transform=transform_value, chosen_fit_index=chosen_fit_index)
        make_linear_fit_plot(location_np=loc, z_np=z, fit_params=fit_list[0], stage=0, xmin=xlimits[0], xmax=xlimits[1], ymin=ylimits[0], ymax=ylimits[1],
                             location=save_location, transform=transform_value)
    else:
        save_location=direct
        fit_list = linear_fit(location=loc, z=z, xyz_table_location=xyz_table, list_of_breakpoints=breakpoints,
                              transform=transform_value, chosen_fit_index=chosen_fit_index)
        make_linear_fit_plot(location_np=loc, z_np=z, fit_params=fit_list[0], stage=0, xmin=xlimits[0], xmax=xlimits[1], ymin=ylimits[0], ymax=ylimits[1],
                             location=save_location, transform=transform_value)
        make_residual_plot(location_np=loc, residual=fit_list[2], R_squared=fit_list[3], stage=0, xmin=xlimits[0], xmax=xlimits[1], location=save_location)
        detrend_that_raster(detrend_location=detrend_workplace, fit_z_xl_file=xyz_table, original_dem=DEM, stage=0, list_of_breakpoints=breakpoints)
