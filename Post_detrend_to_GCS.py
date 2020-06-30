import openpyxl as xl
import arcpy
from arcpy import env
from arcpy.sa import *
from arcpy.da import *
import os
from os import listdir
import Xavier_detrend_postGIS
from Xavier_detrend_postGIS import *
import create_centerline_GUI
from create_centerline_GUI import *
import create_station_lines
from create_station_lines import create_station_lines_function
from os.path import isfile, join
from GCS_analysis import *
#from Lidar_to_detrend_ready_XRN_functions import *
#from classify_landforms_GUI import *
from openpyxl import Workbook
from openpyxl import load_workbook
from matplotlib import pyplot as plt
import numpy as np
import csv
import pandas
import string
import scipy
from scipy import ndimage
from scipy.ndimage import gaussian_filter1d
from scipy.interpolate import UnivariateSpline
import GCS_statistical_analysis_XRN as stats


def detrend_to_wetted_poly(detrended_dem, out_folder, raster_units, max_stage=[], step=1):
    '''This function takes the detrended DEM and outputs stage polygons at param_list[0]=max stage, and param_list=[1]=stage interval (INTEGER)
        and makes smoothed, donut free, wetted polygons for each stage appending stage to the shapefile name. Each polygon is used to create a centerline
        from which the width and detrended Z analysis is completed. Raster units takes strings m or ft'''

    if raster_units == "m":
        spur_length = 10
    else:
        spur_length = 30

    if not os.path.exists(out_folder):
        os.makedirs(out_folder)

    if not os.path.exists(out_folder + "\\wetted_polygons"):
        os.makedirs(out_folder + "\\wetted_polygons")

    lines_location = out_folder + '\\analysis_centerline_and_XS'
    if not os.path.exists(lines_location):
        os.makedirs(lines_location)

    wetted_folder = out_folder + "\\wetted_polygons"

    try:
        # We convert raster units into feet for this stage of the analysis.
        # The Int function rounds down, but since a 1.3ft elevation would only be wetted by a 2ft flow, we add 1 to the detrended_int_ras
        detrend_ras = Raster(detrended_dem)
        if raster_units == "m":
            detrended_ras = float(detrended_ras * 3.28082)
        detrend_ras_int = Int(detrend_ras) + 1
        detrend_ras_int.save(out_folder + "\\dtrnd_int_ft.tif")
        print("Stage range is: " + str(range(max_stage[0])))

        flood_stage_ras = Raster(Con(detrend_ras_int <= max_stage[0], detrend_ras_int))
        flood_stage_poly = arcpy.RasterToPolygon_conversion(flood_stage_ras,
                                                            wetted_folder + ("\\flood_stage_poly_%sft" % max_stage[0]),
                                                            simplify=False, raster_field="Value",
                                                            create_multipart_features=True)
        flood_stage_poly = arcpy.Dissolve_management(flood_stage_poly, wetted_folder + ("\\flood_stage_poly_dissolved_%sft" % max_stage[0]), dissolve_field="gridcode", multi_part=True)
        print("Flood stage polygon made with max stage of %sft and is stored: %s" % (max_stage[0], flood_stage_poly))

        for i in range(0, max_stage[0] + 1, int(step)):
            print(range(max_stage[0] + 1))
            query = ('"gridcode" <= %d' % i)
            print(query)
            stage_poly = arcpy.SelectLayerByAttribute_management(flood_stage_poly, selection_type="NEW_SELECTION",
                                                                 where_clause=query)
            stage_poly = arcpy.CopyFeatures_management(stage_poly,
                                                       out_feature_class=wetted_folder + ("\\flood_stage_poly_%sft" % i))
            arcpy.AddField_management(stage_poly, "null_field", "Short")
            stage_poly_no_donuts = arcpy.Union_analysis(stage_poly, wetted_folder + (
                        "\\flood_stage_poly_%sft_no_donuts_predissolve" % i), gaps="NO_GAPS")
            stage_poly_dissolve = arcpy.Dissolve_management(stage_poly,
                                                            wetted_folder + ("\\flood_stage_poly_dissolved_%sft" % i),
                                                            dissolve_field="null_field", multi_part=True)
            stage_poly_dissolve_no_donuts = arcpy.Dissolve_management(stage_poly_no_donuts, wetted_folder + (
                        "\\flood_stage_poly_%sft_no_donuts" % i), dissolve_field="null_field", multi_part=True)

            if raster_units == "m":
                stage_poly_dissolve_no_donuts = arcpy.SmoothPolygon_cartography(stage_poly_dissolve_no_donuts, wetted_folder + (
                        "\\smooth_stage_poly_%sft_donuts" % i), "PAEK", 50)
            else:
                stage_poly_dissolve_no_donuts = arcpy.SmoothPolygon_cartography(stage_poly_dissolve_no_donuts, wetted_folder + ("\\smooth_stage_poly_%sft_donuts" % i), "PAEK", 164)

            stage_poly_dissolve_no_donuts = arcpy.Union_analysis(stage_poly_dissolve_no_donuts, wetted_folder + ("\\smooth_stage_poly_%sft_no_dissolve" % i), gaps="NO_GAPS")
            stage_poly_dissolve_no_donuts = arcpy.Dissolve_management(stage_poly_dissolve_no_donuts, wetted_folder + ("\\smooth_stage_poly_%sft" % i), dissolve_field="null_field", multi_part=True)
            flood_stage_poly = arcpy.SelectLayerByAttribute_management(flood_stage_poly,
                                                                       selection_type="CLEAR_SELECTION")

            # Create folder for centerline and cross sections, Make centerline based on smoothed, donut-less wetted polygons and delete spurs
            print("Creating centerline for %sft stage" % i)
            centerline = arcpy.PolygonToCenterline_topographic(stage_poly_dissolve_no_donuts, lines_location + ("\\stage_centerline_%sft" % i))
            print('Removing spurs smaller than % s units...' % str(spur_length))
            # measure lengths
            arcpy.AddGeometryAttributes_management(centerline, 'LENGTH')
            spurs = arcpy.SelectLayerByAttribute_management(centerline, where_clause=('LENGTH < %s' % str(spur_length)), selection_type="NEW_SELECTION")
            if int(arcpy.GetCount_management(spurs).getOutput(0)) > 0:
                arcpy.DeleteFeatures_management(spurs)
            arcpy.SelectLayerByAttribute_management(centerline, selection_type="CLEAR_SELECTION")

            print("Stage %s dissolved and non-dissolved polygons in %s" % (i, out_folder))

            print("Deleting unnecessary files")
            try:
                if os.path.exists(out_folder + ("\\smooth_stage_poly_%sft_donuts.shp" % i)):
                    os.remove(out_folder + ("\\smooth_stage_poly_%sft_donuts.shp" % i))
                if os.path.exists(out_folder + ("\\smooth_stage_poly_%sft_no_dissolve" % i)):
                    os.remove(out_folder + ("\\smooth_stage_poly_%sft_no_dissolve" % i))
                if os.path.exists(out_folder + ("\\flood_stage_poly_%sft_no_donuts" % i)):
                    os.remove(out_folder + ("\\flood_stage_poly_%sft_no_donuts" % i))
                if os.path.exists(out_folder + ("\\flood_stage_poly_%sft_no_donuts_predissolve" % i)):
                    os.remove(out_folder + ("\\flood_stage_poly_%sft_no_donuts_predissolve" % i))
            except:
                print("Some files not deleted")

        print("wetted polygons created")

    except arcpy.ExecuteError:
        print(arcpy.GetMessages())


## GOOD TO RUN A 3m smoothing during centerline editing

def width_series_analysis(out_folder, float_detrended_DEM, raster_units, biggest_stage, spacing=[], centerlines=[], XS_lengths=[], ft_smoothing_tolerance=75, clip_poly=''):
    ''' For each wetted polygon produced within the max_stage range set in the detrend_to_wetted_poly function, this function splits the polygon
    into rectangular slices which are used to extract mean depth and width for each filling out an excel sheet and some descriptive stats.

    Centerlines created in previous function should be visually inspected for quality, edited, and adjusted. Centerlines list holds the LAST stage to use a specified stage centerline
    NOTE: This function splicing the file names used for the dissolved poly, so make sure that doesn't change in the other function without updating this.'''
    print("Width analysis comencing. XS spacing is %s%s, analysis will be conducted up to flood stage %s%s" % (spacing[0],raster_units,biggest_stage,raster_units))

    wetted_folder = (out_folder + "\\wetted_polygons")
    list_of_files_in_out_folder = [f for f in listdir(wetted_folder) if isfile(join(wetted_folder, f))]
    list_of_dissolved_polygons = []
    for file in list_of_files_in_out_folder:
        if file[17:26] == "dissolved" and file[-4:] == ".shp":
            list_of_dissolved_polygons.append(file)
    print("Unsorted list of dissolved polygons:" + str(list_of_dissolved_polygons))

    lines_location = out_folder + '\\analysis_centerline_and_XS'
    shapefile_location = out_folder + '\\analysis_shapefiles'
    if not os.path.exists(shapefile_location):
        os.makedirs(shapefile_location)

    try:
        for stage_line in centerlines:
            index = centerlines.index(stage_line)
            length = XS_lengths[index]
            # Make station lines for chosen centerlines
            centerline_location = (lines_location + "\\stage_centerline_%sft.shp" % stage_line)
            centerline_dissolve = arcpy.Dissolve_management(centerline_location, (lines_location + "\\stage_centerline_%sft_D.shp" % stage_line), dissolve_field="ObjectID", multi_part="MULTI_PART")
            centerline_dissolve = (lines_location + "\\stage_centerline_%sft_D.shp" % stage_line)
            arcpy.AddField_management(centerline_dissolve, "Id", "SHORT")
            if raster_units == "m":
                tolerance = float(ft_smoothing_tolerance / 3)
            else:
                tolerance = ft_smoothing_tolerance

            centerline_dissolve = arcpy.SmoothLine_cartography(centerline_dissolve, (lines_location + "\\stage_centerline_%sft_DS.shp" % stage_line), algorithm="PAEK", tolerance=tolerance)
            centerline_dissolve = (lines_location + "\\stage_centerline_%sft_DS.shp" % stage_line)
            create_station_lines.create_station_lines_function(centerline_dissolve, spacing=float(spacing[0]), xs_length=float(length), stage=[int(stage_line)])
            station_lines = lines_location + ("\\stage_centerline_%sft_DS_XS_%sft.shp" % (int(stage_line), spacing[0]))

            if os.path.exists(lines_location + "\\stage_centerline_%sft_D.shp" % stage_line):
                os.remove(
                    lines_location + "\\stage_centerline_%sft_D.shp" % stage_line)  # Remove non-smoothed, dissolved centerline
            print("Station lines file at: " + str(station_lines))

        for file in list_of_dissolved_polygons: #remove dissolved polygons above designated biggest stage from list
            if file[-8] == "_":
                stage = int(file[-7])
            else:
                stage = int(file[-8:-6])
            if stage > biggest_stage:
                list_of_dissolved_polygons.remove(file)

        for file in list_of_dissolved_polygons: #Assign stage value to the wetted polygon
            if file[-8] == "_":
                stage = int(file[-7])
            else:
                stage = int(file[-8:-6])
            print("Doing width analysis for stage %s%s" % (stage, raster_units))

            centerline_number = 0
            if stage < int(centerlines[0]) or stage == int(centerlines[0]):
                centerline = (lines_location + "\\stage_centerline_%sft_DS.shp" % centerlines[0])
                station_lines = lines_location + ("\\stage_centerline_%sft_DS_XS_%sft.shp" % (centerlines[0], spacing[0]))
            elif stage > centerlines[-1]:
                centerline = (lines_location + "\\stage_centerline_%sft_DS.shp" % centerlines[-1])
                station_lines = lines_location + ("\\stage_centerline_%sft_DS_XS_%sft.shp" % (centerlines[-1], spacing[0]))
            else:
                while int(stage) > centerlines[centerline_number]:
                    centerline_number += 1
                centerline = (lines_location + "\\stage_centerline_%sft_DS.shp" % centerlines[centerline_number])
                station_lines = lines_location + ("\\stage_centerline_%sft_DS_XS_%sft.shp" % (centerlines[centerline_number], spacing[0]))


            print('Using station lines found @ %s' % station_lines) #XS station lines are assigned to the stage

            spacing_half = float(spacing[0] / 2)
            file_location = (wetted_folder + "\\%s" % file) # might need to string splice here
            clipped_lines = arcpy.Clip_analysis(station_lines, file_location, out_feature_class=(
                        shapefile_location + "\\clipped_station_lines1_%s" % file[-8:]))

            if clip_poly != '':
                clipped_lines = arcpy.Clip_analysis(clipped_lines, file_location, out_feature_class=(shapefile_location + "\\clipped_station_lines2_%s" % file[-8:]))

            rectangles = arcpy.Buffer_analysis(clipped_lines, out_feature_class=(
                        shapefile_location + "\\width_rectangles_%sft.shp" % stage),
                                               buffer_distance_or_field=spacing_half, line_side="FULL",
                                               line_end_type="FLAT")
            arcpy.AddField_management(rectangles, "Width", field_type="FLOAT")
            expression = ("(float(!Shape.area!)) / %d" % spacing[0])
            arcpy.CalculateField_management(rectangles, "Width", expression, "PYTHON3")

            #print((shapefile_location + "\\stats_table_%s.dbf" % stage))

            rectangles = (shapefile_location + "\\width_rectangles_%s" % file[-8:])

    except arcpy.ExecuteError:
        print(arcpy.GetMessages())

    print("Clipped station lines and width rectangles in %s" % shapefile_location)

def z_value_analysis1(out_folder, detrended_DEM):
    '''This function used zonal statistics with the stage dependent centerlines to extract mean detrended elevation from each width polygon.
    NO RE-DETRENDING is done, as opposed to z_value_analysis2 where the elevation profile is detrended based on different centerlines'''
    width_series_shapefile_folder = out_folder + "\\analysis_shapefiles"
    list_of_files_width_folder = [f for f in listdir(width_series_shapefile_folder) if
                                  isfile(join(width_series_shapefile_folder, f))]  # add width rectangles to a list
    list_of_width_polygons = []
    for file in list_of_files_width_folder:
        if file[:16] == "width_rectangles" and file[-4:] == ".shp" and file[-8:-6] != '_0':
            list_of_width_polygons.append(file)
    print("Unsorted list of width polygons:" + str(list_of_width_polygons))

    for file in list_of_width_polygons:
        width_file = width_series_shapefile_folder + "\\" + file

        if file[-8] == "_": # assign stage int value to label files correctly
            stage = int(file[-7])
        else:
            stage = int(file[-8:-6])


        arcpy.AddField_management(width_file, field_name="loc_id", field_type="SHORT")
        field_calc = "(int(!LOCATION!))"
        arcpy.CalculateField_management(width_file, field="loc_id", expression=field_calc, expression_type="PYTHON3")

        zonal_table = arcpy.sa.ZonalStatisticsAsTable(width_file, zone_field="loc_id", in_value_raster=detrended_DEM,
                                                      out_table=(
                                                                  width_series_shapefile_folder + "\\stats_table_%s.dbf" % stage),
                                                      statistics_type="MEAN")
        rectangles = arcpy.JoinField_management(width_series_shapefile_folder + "\\" + file, in_field="loc_id",
                                                join_table=zonal_table, join_field="loc_id", fields=["MEAN"])

        print("Z statistics added for stage %s" % stage)
    print("Z analysis completed, use following functions to export to excel and plot")


def z_value_analysis2(out_folder, original_DEM, spacing, breakpoint, centerlines=[]):
    '''This function iteratively detrends the lidar DEM by the used centerlines, and then is used to extract zonal statistics for the width rectanges
    and export as tables for the GCS analysis. Breakpoint much be an integer that is a multiple of the spacing'''
    lines_location = out_folder + '\\analysis_centerline_and_XS'
    for stage_num in centerlines:
        if os.path.exists(lines_location + "\\stage_centerline_%sft_DS.shp" % stage_num):
            print("Smoothed and edited stage centerline %s exists and is ready to be processed" % ("stage_centerline_%sft_DS.shp" % stage_num))

    for stage_num in centerlines:
        line = ("\\stage_centerline_%sft_DS.shp" % stage_num)
        line_location = lines_location + ("\\%s" % line)
        # We now make non-self-intersecting temporary station lines for the analysis
        station_lines = create_station_lines.create_station_lines_function(line_location, spacing=spacing, xs_length=5, stage=[stage_num])
        station_lines = lines_location + ("\\stage_centerline_%sft_DS_XS_%sft.shp" % (int(stage_num), spacing))
        print("Station line made for %sft stage: %s" % (stage_num, station_lines))

        detrend_file_location = lines_location + "\\detrend_files"
        if not os.path.exists(detrend_file_location):
            os.makedirs(detrend_file_location)

        station_points = arcpy.Intersect_analysis([station_lines, line_location], out_feature_class=(detrend_file_location + "\\stage_%s_dtpoints_%sft_SP.shp" % (stage_num, spacing)),
                                                  join_attributes="ALL", output_type="POINT")
        station_points = arcpy.MultipartToSinglepart_management(station_points, (detrend_file_location + "\\stage_%s_dtpoints_%sftP.shp" % (stage_num, spacing)))
        station_points = arcpy.AddXY_management(station_points)
        print("Station points made for %sft stage: %s" % (stage_num, station_points))

        elevation_table = arcpy.ExtractValuesToTable_ga(station_points, in_rasters=original_DEM, out_table=(detrend_file_location + "\\stage_%s_dtpoints_%sft.dbf" % (stage_num, spacing)))
        arcpy.JoinField_management(in_data=station_points, in_field="FID", join_table=elevation_table, join_field="OID", fields=["Value"]) #Add Z value from original DEM to station points as "Value"

        if os.path.exists(detrend_file_location + "\\stage_%s_dtpoints_%sft.dbf" % (stage_num, spacing)):
            os.remove(detrend_file_location + "\\stage_%s_dtpoints_%sft.dbf" % (stage_num, spacing))  # Remove dbf table with point elevations
        if os.path.exists(detrend_file_location + "\\stage_%s_dtpoints_%sft_SP.shp" % (stage_num, spacing)):
            os.remove(detrend_file_location + "\\stage_%s_dtpoints_%sft_SP.shp" % (stage_num, spacing)) #Remove multipart station points

        delete_field_names = ["FID_stage_", "Id", "FID_stage1", "ObjectID", "ID1", "ID_1", "InLine_FID", "ORIG_FID", "POINT_M"]
        print("The following list of fields will be deleted before exporting to xlsx: %s" % delete_field_names)
        arcpy.DeleteField_management(station_points, drop_field=delete_field_names)
        elevation_table = arcpy.TableToTable_conversion(station_points,
                    out_path=detrend_file_location, out_name="stage_%sft_XYZ_table_%sft.csv" % (stage_num, spacing))
        if os.path.exists(detrend_file_location + "\\stage_%sft_XYZ_table_%sft.csv" % (stage_num, spacing)):
            print("Stage %sft XYZ table for detrending is created and located at: %s" % (stage_num, elevation_table))
        else:
            print("Something went wrong...")
        elevation_table = detrend_file_location + "\\stage_%sft_XYZ_table_%sft.csv" % (stage_num, spacing)

        ##### See if we can change the detrending functions to be able to run with just xl csv and points so we can call the function iteratively
        #without repeating ourselves. We should then use this function to do the spatial analysis and add the values to analysis tables in proper form

        location_and_z = Xavier_detrend_postGIS.prep_xl_file(xyz_table_location=elevation_table,listofcolumn=['B', 'A', 'E', 'C', 'D'])
        location_np = location_and_z[0]
        z_np = location_and_z[1]
        #Xavier_detrend_postGIS.diagnostic_quick_plot(location_np,z_np) #Un-hash out if first time interacting with data
        linear_fit_return_list = linear_fit(location=location_np, z=z_np, xyz_table_location=elevation_table, list_of_breakpoints=[0,int(breakpoint)])
        fit_params = linear_fit_return_list[0]
        z_fit_list = linear_fit_return_list[1]
        residual = linear_fit_return_list[2]
        R_squared = linear_fit_return_list[3]
        make_linear_fit_plot(location_np, z_np, fit_params, stage=stage_num, location=detrend_file_location)
        elevation_table_xlsx = elevation_table[:-3] + "xlsx"
        detrend_that_raster(detrend_location=detrend_file_location, fit_z_xl_file=elevation_table_xlsx, original_dem=original_DEM, stage=stage_num,
                            list_of_breakpoints=[int(breakpoint)])
        detrended_raster = detrend_file_location + "\\rs_dt_s%s.tif" % stage_num
        print("Detrended raster for stage %s is made @: %s" % (stage_num, detrended_raster))

        print("Deleting intermediary files...")
        if os.path.exists(detrend_file_location + "\\thiespoly_stage%s.shp" % stage_num):
            os.remove(detrend_file_location + "\\thiespoly_stage%s.shp" % stage_num)
        if os.path.exists('theis_raster%s_stage%sft.tif' % (breakpoint, stage_num)):
            os.remove('theis_raster%s_stage%sft.tif' % (breakpoint, stage_num))

    width_series_shapefile_folder = out_folder + "\\analysis_shapefiles"
    list_of_files_width_folder = [f for f in listdir(width_series_shapefile_folder) if
                                      isfile(join(width_series_shapefile_folder, f))]  # add width rectangles to a list
    list_of_width_polygons = []
    for file in list_of_files_width_folder:
        if file[:16] == "width_rectangles" and file[-4:] == ".shp" and file[-8:-6] != '_0':
            list_of_width_polygons.append(file)
    print("Unsorted list of width polygons:" + str(list_of_width_polygons))

    for file in list_of_width_polygons:
        if file[-8] == "_":
            stage = int(file[-7])
        else:
            stage = int(file[-8:-6])
        centerline_number = 0
        if stage <= centerlines[centerline_number]:
            detrended_ras = (detrend_file_location + "\\rs_dt_s%s.tif" % centerlines[0])
        elif stage > centerlines[-1]:
            detrended_ras = (detrend_file_location + "\\rs_dt_s%s.tif" % centerlines[-1])
        else:
            while int(stage) > centerlines[centerline_number]:
                centerline_number += 1
            detrended_ras = (detrend_file_location + "\\rs_dt_s%s.tif" % centerlines[centerline_number])
        print("Stage %s width analysis polygon will be joined with Z attributes from %s" % (stage, detrended_ras))

        width_file = width_series_shapefile_folder + "\\" + file
        arcpy.AddField_management(width_file,field_name="loc_id",field_type="SHORT")
        field_calc = "(int(!LOCATION!))"
        arcpy.CalculateField_management(width_file,field="loc_id",expression=field_calc, expression_type="PYTHON3")

        zonal_table = arcpy.sa.ZonalStatisticsAsTable(width_file, zone_field="loc_id", in_value_raster=detrended_ras, out_table=(width_series_shapefile_folder + "\\stats_table_%s.dbf" % stage), statistics_type="MEAN")
        rectangles = arcpy.JoinField_management(width_series_shapefile_folder + "\\" + file, in_field="loc_id", join_table=zonal_table, join_field="loc_id", fields=["MEAN"])

        print("Z statistics added for stage %s" % stage)
    print("Z analysis completed, use following functions to export to excel and plot")


        #### Continue by using the width polygons to get zonal stats as a table and then join based on FID using the respective detrended DEM

def export_to_gcs_ready(out_folder, list_of_error_locations=[]):
    '''Export location, width, and mean depth as tables excluding damaged cross sections specified as a list.
    Table outputs ready for GCS_analysis.py analysis in a new folder'''
    table_location = out_folder + '\\gcs_ready_tables'
    if not os.path.exists(table_location):
        os.makedirs(table_location)

    analysis_folder = out_folder + "\\analysis_shapefiles"
    files_in_folder = [f for f in listdir(analysis_folder) if isfile(join(analysis_folder, f))]
    width_rectangles = [i for i in files_in_folder if i[0] == "w"]
    width_rectangles = [i for i in width_rectangles if i[-4:] == ".shp" and i[-8:-6] != '_0']
    print("Width rectanles found: %s" % width_rectangles)

    list_of_csv_tables = []

    # Delete bad locations and convert data into a dbf table
    for file in width_rectangles:
        file_location = (analysis_folder + "\\%s" % file)
        if len(list_of_error_locations) > 0:
            print("Erroneous cross sections identified...")
            for value in list_of_error_locations:
                query = ('"LOCATION" = %d' % value)
                arcpy.SelectLayerByAttribute_management(file_location, selection_type="ADD_TO_SELECTION", where_clause=query)
            arcpy.DeleteFeatures_management(file_location)
            arcpy.SelectLayerByAttribute_management(file_location, selection_type="CLEAR_SELECTION")
        analysis_table = arcpy.TableToTable_conversion(file_location, out_path=table_location, out_name=("WD_analysis_table_%s.dbf" % file[-8:-4]))
        analysis_xlsx = arcpy.TableToExcel_conversion(Input_Table=file_location, Output_Excel_File=table_location + ("\\WD_analysis_table_%s.xlsx" % file[-8:-4]))

        #Convert xls to csv and change field headers to match GCS_analysis functions
        wb = xl.load_workbook(str(analysis_xlsx))
        ws = wb.active
        if file[-8] == "_":
            csv_file = table_location + ("\\%s_WD_analysis_table.csv" % file[-7:-4])
        else:
            csv_file = table_location + ("\\%s_WD_analysis_table.csv" % file[-8:-4])
        if ws["H1"].value == "MEAN" and ws["C1"].value == "LOCATION" and ws["F1"].value == "Width":
            ws["H1"].value = "Z"
            ws["F1"].value = "W"
            ws["C1"].value = "dist_down"
            wb.save(str(analysis_xlsx))
        else:
            print("ERROR! CELL HEADERS IN UNEXPECTED POSITIONS @ %s" % (table_location + ("\\WD_analysis_table_%s.xlsx" % file[-8:-4])))
            return

        with open(csv_file, 'w', newline="") as f:
            col = csv.writer(f)
            for row in ws.rows:
                col.writerow([cell.value for cell in row])
        list_of_csv_tables.append(csv_file)

    print("DBF and CSV tables of width/Z analysis at %s" % table_location)
    print("List of csv tables: %s" % list_of_csv_tables)

    print("Landform classification is underway..")
    main_classify_landforms(list_of_csv_tables, w_field='W', z_field='Z', dist_field='dist_down', make_plots=False)

    return [list_of_csv_tables, width_rectangles]

def GCS_plotter(table_directory):
    '''Take the, Z, W, and landforms/Ws*Zs and plot vs dist_down using pyplot and open xl'''
    list_of_csv_tables = [f for f in listdir(table_location) if isfile(join(table_location, f))]
    list_of_csv_tables = [f for f in list_of_csv_tables if f[-4:] == ".csv"]
    print(list_of_csv_tables)

    # Set up figure output folder
    figure_output = table_directory + "\\GCS_figures"
    if not os.path.exists(figure_output):
        os.makedirs(figure_output)

    numpy_columns = ['dist_down', 'W_s', 'Z_s', 'W_s_Z_s']

    for table in list_of_csv_tables:
        if table[1] == "f":
            stage = table[0]
        else:
            stage = table[:2]

        table_df = pandas.read_csv(table_location + "\\%s" % table)
        print(table_df)
        zs_plot_name = ("%s_Zs_plot" % stage)
        ws_plot_name = ("%s_Ws_plot" % stage)
        zs_ws_plot_name = ("%s_ZsWs_plot" % stage)

        #Make data frames for each landform type
        wide_bar_df = table_df.loc[table_df['code'] == 1, ['dist_down', 'W_s', 'Z_s', 'W_s_Z_s']]
        nozzle_df = table_df.loc[table_df['code'] == 2, ['dist_down', 'W_s', 'Z_s', 'W_s_Z_s']]
        normal_df = table_df.loc[table_df['code'] == 0, ['dist_down', 'W_s', 'Z_s', 'W_s_Z_s']]
        pool_df = table_df.loc[table_df['code'] == -1, ['dist_down', 'W_s', 'Z_s', 'W_s_Z_s']]
        oversized_df = table_df.loc[table_df['code'] == -2, ['dist_down', 'W_s', 'Z_s', 'W_s_Z_s']]


        for i in numpy_columns[1:]:
            plot_1 = wide_bar_df[['dist_down', ('%s' % i)]].to_numpy()
            plot_2 = nozzle_df[['dist_down', ('%s' % i)]].to_numpy()
            plot_3 = normal_df[['dist_down', ('%s' % i)]].to_numpy()
            plot_4 = pool_df[['dist_down', ('%s' % i)]].to_numpy()
            plot_5 = oversized_df[['dist_down', ('%s' % i)]].to_numpy()
            plot_6 = table_df[['dist_down', ('%s' % i)]].to_numpy()

            plt.scatter(plot_1[:,0], plot_1[:,1], c='orange', s=0.75, label="Wide bar")
            plt.scatter(plot_2[:,0], plot_2[:,1], c='gold', s=0.75, label="Nozzle")
            plt.scatter(plot_3[:,0], plot_3[:,1], c='c', s=0.75, label="Normal")
            plt.scatter(plot_4[:,0], plot_4[:,1], c='grey', s=0.75, label="Const. Pool")
            plt.scatter(plot_5[:,0], plot_5[:,1], c='navy', s=0.75, label="Oversized")

            t1 = plot_5[:,0]
            t2 = plot_5[:,1]

            plt.xlabel("Thalweg distance downstream (ft)")
            plt.ylabel("%s" % i)
            axes = plt.gca()
            ymin = min(table_df[i])-1
            if i == 'W_s_Z_s':
                ymax = 5
            else:
                ymax = max(table_df[i])+1
            axes.set_ylim([ymin, ymax])
            axes.set_xlim([0, max(table_df['dist_down'])])
            plt.title("Stage %s %s plot" % (stage, i))
            plt.grid(b=True, which='major', color='#666666', linestyle='-')
            plt.minorticks_on()
            plt.grid(b=True, which='minor', color='#999999', linestyle='-', alpha=0.2)
            plt.legend(('Wide bar', 'Nozzle', 'Normal', 'Pool', 'Oversized'), loc=1)

            fig = plt.gcf()
            fig.set_size_inches(12, 6)
            plt.savefig((figure_output + '\\Stage_%s_%s_plot' % (stage, i)), dpi=300, bbox_inches='tight')
            plt.cla()

            #Now plot Gaussian filter plot and spine plot with parameters
            plot_6_sorted = plot_6[plot_6[:, 0].argsort()]
            dist_points = plot_6_sorted[:,0]
            attribute_points = plot_6_sorted[:,1]
            gaussian_sigma3 = gaussian_filter1d(attribute_points, 3,axis=0)
            gaussian_sigma6 = gaussian_filter1d(attribute_points, 10,axis=0)

            plt.scatter(dist_points, attribute_points,c='black',s=0.75,ls='--',label=("%s points" % i))
            plt.plot(dist_points,gaussian_sigma3,c='red',ls='-',label="Gaussian filter, sigma=3")
            plt.plot(dist_points,gaussian_sigma6,c='green',ls='-',label="Gaussian filter, sigma=10")
            #plt.plot(x_linespace,spline(x_linespace),c='blue',ls='-',label=("Spline ft with coefficients %s" % spline_coeff))
            axes.set_xlim([0, max(table_df['dist_down'])])
            plt.title("Data filtering plot for %s at stage %sft" % (i,stage))
            plt.grid(b=True, which='major', color='#666666', linestyle='-')
            plt.minorticks_on()
            plt.grid(b=True, which='minor', color='#999999', linestyle='-', alpha=0.2)
            plt.legend(('Gaussian filter, sigma=3', 'Gaussian filter, sigma=10'), loc=1)
            #plt.show()
            fig = plt.gcf()
            fig.set_size_inches(12, 6)
            plt.savefig((figure_output + '\\Stage_%s_%s_filtering_plot' % (stage, i)), dpi=300, bbox_inches='tight')
            plt.cla()



##### INPUTS #####
comid_list = [22514218]
SCO_number = 1

for comid in comid_list:
    direct = (r"Z:\users\xavierrn\SoCoast_Final_ResearchFiles\SCO%s\COMID%s" % (SCO_number, comid))
    out_folder = direct + '\\LINEAR_DETREND'
    original_dem_location = direct + '\\las_files\\ls_nodt.tif'
    detrended_dem_location = out_folder + "\\ras_detren.tif" #change back
    process_footprint = direct + '\\las_footprint.shp'
    spatial_ref = arcpy.Describe(detrended_dem_location).spatialReference
    station_lines = direct + "\\las_files\\centerline\\smooth_centerline_XS_3x5ft"
    table_location = out_folder + "\\gcs_ready_tables"
    channel_clip_poly = out_folder + '\\raster_clip_poly.shp' #optional paramter for width_series_analysis

    arcpy.env.workplace = direct #Set arcpy environment
    arcpy.env.extent = detrended_dem_location
    arcpy.env.overwriteOutput = True

    #Call functions:
    #detrend_to_wetted_poly(detrended_dem=detrended_dem_location, out_folder=out_folder, raster_units="ft", max_stage=[20], step=1)
    width_series_analysis(out_folder, float_detrended_DEM=detrended_dem_location, raster_units="ft",biggest_stage=20, spacing=[6], centerlines=[1,2,9], XS_lengths=[100,225,2000],ft_smoothing_tolerance=75, clip_poly=channel_clip_poly)
    z_value_analysis1(out_folder=out_folder, detrended_DEM=detrended_dem_location)

    export_list = export_to_gcs_ready(out_folder=out_folder, list_of_error_locations=[])
    tables = export_list[0]
    main_classify_landforms(tables, w_field='W', z_field='Z', dist_field='dist_down', out_folder=out_folder, make_plots=False)
    GCS_plotter(table_directory=table_location)

    out_list = stats.analysis_setup(table_directory=table_location)
    stages_dict = out_list[0]
    stages_stats_xl_dict = out_list[1]
    max_stage = out_list[2]
    stats_table_location = out_list[3]

    stats.stage_level_descriptive_stats(stages_dict, stages_stats_xl_dict, max_stage, box_and_whisker=True)
    stats.compare_flows(stages_stats_xl_dict, max_stage, save_plots=True)
    stats.autocorr_and_powerspec(stages_dict, stages_stats_xl_dict, max_stage, save_plots=True)








