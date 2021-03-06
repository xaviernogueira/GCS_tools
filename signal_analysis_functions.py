import math

import scipy as sp
import scipy.signal as sig
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
from matplotlib.ticker import MaxNLocator
from matplotlib.lines import Line2D
import matplotlib.tri as tri
import pandas as pd
import GCS_analysis
import openpyxl as xl
from openpyxl import Workbook
import os
from os import listdir
from os.path import isfile, join
import itertools
from itertools import combinations
import file_functions
from file_functions import *
import key_z_analysis_functions


def powerspec_plotting(in_folder, out_folder, key_zs=[], fields=['W_s', 'Z_s', 'W_s_Z_s'], smoothing=5):
    '''This function saves a plot of the PSD for each input key z to the out_folder using GCS csv files found in the in_folder.
    INPUTS: in_folder containing (KEY Z)ft_WD_analysis_table.csv files for each selected key z. out_folder to save fig.
    key_zs can be either float or int.
    fields can be changed from the defaults is csv headers are different.
    Smoothing (int, default=5) is the moving average window to smoothing the power spectral data. If 0 no moving average is used. '''
    print('Plotting Key Z power spectral densities...')
    value_dict = {}  # Stores Ws and C(Ws,Zs) power spectral density values respectively
    z_str_list = []
    key_zs.sort()
    for field in fields:
        value_dict[field] = []
    labels = ['Base flow', 'Bank full', 'Valley Fill']

    for value in value_dict.keys():
        freq_lists = []
        dens_lists = []
        for z in key_zs:
            z_str = key_z_analysis_functions.float_keyz_format(z)
            z_str_list.append(z_str)

            df = pd.read_csv(in_folder + '\\%sft_WD_analysis_table.csv' % z_str)

            df.sort_values(['dist_down'], inplace=True)
            spacing = df['dist_down'][1] - df['dist_down'][0]
            values = df.loc[:, [value]].squeeze()

            frequencies, psd = sig.periodogram(values, 1.0/spacing, window=sig.get_window('hamming', len(values)), detrend=False)

            if smoothing != 0:
                psd = sp.ndimage.uniform_filter1d(psd, size=smoothing)
            freq_lists.append(frequencies)
            dens_lists.append(psd)

        value_dict[value].append(freq_lists)
        value_dict[value].append(dens_lists)

    for value in value_dict.keys():
        fig, ax = plt.subplots(len(key_zs), 1, sharex=True, sharey=True)
        fig_name = out_folder + '\\Key_z_PSD_%s.png' % value
        ax[0].set_title('%s Power Spectral Density plots' % value)
        ax[0].set_xticks(np.arange(0.0, np.max(value_dict[value][0][0]), round(np.max(value_dict[value][0][0] / 30), 3)))
        ax[0].grid(b=True, which='major', color='grey', linewidth=1.0)
        ax[0].grid(b=True, which='minor', color='grey', linewidth=0.5)

        ax[len(key_zs) - 1].set_xlabel('Frequency (cycles/ft)')

        max_freq = 0
        for count, z in enumerate(key_zs):
            freqs = value_dict[value][0][count]
            if np.max(freqs) >= max_freq:
                max_freq = np.max(freqs)
            power = value_dict[value][1][count]
            ax[count].plot(freqs, power, '*-', color='blue')
            ax[count].set_ylabel('%sft stage PSD' % z)
            ax[count].set_xscale('log')
            ax[count].set_yscale('log')
            ax[count].grid(b=True, which='major', color='grey', linewidth=1.0)
            ax[count].grid(b=True, which='minor', color='grey', linewidth=0.5, linestyle='--')
            ax[count].text(0.75, 0.75, labels[count], transform=ax[count].transAxes, fontsize=16)

        ax[0].set_xlim(None, max_freq)
        fig.set_size_inches(12, 6)
        plt.savefig(fig_name, dpi=300, bbox_inches='tight')
        plt.cla()
    plt.close('all')
    print('PSD plots created!')


def cross_corr_analysis(in_folder, out_folder, detrend_folder, key_zs, fields=['W_s_Z_s', 'W_s', 'Z_s'], in_csv=''):
    '''This function plots key z signals for the given fields, and the cross correlation between each combination of two signals.
    INPUTS: Folder containing aligned all_stages_table.csv. out_folder is where the plot is saved.
    key_zs can be float or int. If in_csv is overridden to equal a csv location, a csv not part of the documented file structure is used.'''

    print('Plotting key z signal correlation')
    value_dict = {}  # Stores Ws and C(Ws,Zs) power spectral density values respectively
    z_str_list = []
    key_zs.sort()
    for field in fields:
        value_dict[field] = []
    labels = ['Base flow', 'Bank full', 'Valley Fill']
    join_field = 'loc_%sft' % min(key_z_analysis_functions.find_centerline_nums(detrend_folder))

    aligned_df = pd.read_csv(in_folder + '\\aligned_locations.csv')
    aligned_df.dropna(inplace=True)
    aligned_df.sort_values(join_field, inplace=True)

    locs = aligned_df.loc[:, [join_field]].squeeze()
    locs_list = locs.tolist()

    spacing_list = []
    for position, loc in enumerate(locs_list[1:]):  # Adjusted to deal with droppped nan value rows
        spacing_list.append(loc - locs_list[position])
    spacing = int(max(set(spacing_list), key=spacing_list.count))
    if spacing < 0:
        spacing == abs(int(spacing))

    for value in value_dict.keys():
        signals = []
        cor_coeffs = []
        colors = ['red', 'blue', 'green', 'pink', 'orange']
        fig_name = out_folder + '\\Key_z_corr_%s.png' % value
        comb = list(combinations(key_zs, 2))

        for z in key_zs:
            z_str = key_z_analysis_functions.float_keyz_format(z)
            z_str_list.append(z_str)

        fig, ax = plt.subplots(len(comb) + 1, 1, sharex=True, sharey=False)
        ax[0].set_title('Cross Correlation of %s signals' % value)
        ax[0].set_ylabel(value)
        ax[len(key_zs)].set_xlabel('Thalweg distance downstream (ft)')

        for count, z in enumerate(key_zs):
            signals.append(aligned_df.loc[:, [value + '_%sft' % z_str_list[count]]].squeeze())

        min_sig = 0
        max_sig = 0
        for count, signal in enumerate(signals):
            ax[0].plot(locs, signal, color=colors[count], label=labels[count] + ' (%sft)' % key_zs[count])
            if np.min(signal) <= min_sig:
                min_sig = np.min(signal)
            if np.max(signal) >= max_sig:
                max_sig = np.max(signal)
        ax[0].grid(True, which='both')
        ax[0].set_xticks(np.arange(0, np.max(locs), 250))
        ax[0].set_xlim(0.0, np.max(locs))
        ax[0].set_yticks(np.arange(round(min_sig, 0), round(max_sig, 0), 1), minor=False)
        ax[0].legend(loc='lower center', ncol=len(signals), fontsize=8)

        min_corr = 0
        max_corr = 0
        for count, combo in enumerate(comb):
            index1 = key_zs.index(combo[0])
            index2 = key_zs.index(combo[1])
            corr = sig.correlate(signals[index1], signals[index2], mode='same')
            ax[count + 1].plot(locs, corr)
            ax[count + 1].set_ylabel('%sft and %sft' % (combo[0], combo[1]))
            ax[count + 1].grid(True, which='both')

            if np.min(corr) <= min_corr:
                min_corr = np.min(corr)
            if np.max(corr) >= max_corr:
                max_corr = np.max(corr)

        for i in range(0, len(comb)):
            ax[i + 1].set_ylim(min_corr, max_corr)
            ax[i + 1].set_yticks(np.arange(round(min_corr, -2), round(max_corr, -2), 50), minor=True)

        fig.set_size_inches(12, 6)
        plt.savefig(fig_name, dpi=300, bbox_inches='tight')
        plt.cla()
    plt.close('all')
    print('Cross-Correlation plots finished!')


def fourier_analysis(in_folder, out_folder, detrend_folder, key_zs, fields=['W_s_Z_s', 'W_s', 'W', 'Z_s'], n=0, in_csv='', by_power=False):
    '''INPUTS:
        N (0 default, accepts int or list. Refers to the # of hamronic components included in the analysis,
        Set the parameter N=list(range(1, N)) for a incrementing range. N=0 does a normal fft and ifft.
        in_csv allows the aligned csv to be explicitly referenced if not 'aligned_locations.csv'
        If by_power == True (bool, False is default)and n != 0, the n highest power frequencies are plotted, else the n lowest frequencies are used.'''

    print('Calculating Pearsons correlation between signals and inverse-FFT plots...')
    value_dict = {}  # Stores Ws and C(Ws,Zs) power spectral density values respectively
    z_str_list = []
    key_zs.sort()
    comb = list(combinations(key_zs, 2))
    for field in fields:
        value_dict[field] = []
    labels = ['Base flow', 'Bank full', 'Valley Fill']
    join_field = 'loc_%sft' % min(key_z_analysis_functions.find_centerline_nums(detrend_folder))

    wb = xl.Workbook()
    lab = n
    if n == 0:
        lab = 'all'
    if by_power == False:
        xl_name = out_folder + '\\%s_harmonic_coefs.xlsx' % lab
    else:
        xl_name = out_folder + '\\%s_harmonic_coefs_by_PSD.xlsx' % lab
    wb.save(xl_name)

    ns = []
    if isinstance(N, int):
        ns.append(N)
    elif isinstance(N, list):
        for i in N:
            ns.append(i)

    if in_csv == '':
        aligned_csv = in_folder + '\\aligned_locations.csv'
    else:
        aligned_csv = in_csv

    aligned_df = pd.read_csv(aligned_csv)
    aligned_df.sort_values(join_field, inplace=True)
    locs = aligned_df.loc[:, [join_field]].squeeze()
    spacing = locs[1] - locs[2]

    for i, value in enumerate(value_dict.keys()):
        if value == 'W_s_Z_s':
            value_for_fig = 'WsZs'
        else:
            value_for_fig = value

        fig, ax = plt.subplots(len(comb), 1, sharex=True, sharey=True)
        ax[0].set_xticks(np.arange(0, np.max(locs), 250))
        if i == 0:
            ws = wb.active
            ws.title = value_for_fig
        else:
            wb.create_sheet(value_for_fig)
            ws = wb[value_for_fig]

        col = 1
        for z in key_zs:
            z_str = key_z_analysis_functions.float_keyz_format(z)
            z_str_list.append(z_str)
            ws.cell(row=1, column=col).value = '%sft cos coefs' % z
            ws.cell(row=1, column=col + 1).value = '%sft sin coefs' % z
            col += 2

        signals = []
        for count, z in enumerate(key_zs):
            signals.append(aligned_df.loc[:, [value + '_%sft' % z_str_list[count]]].squeeze())
        comb = list(combinations(key_zs, 2))

        ax[len(signals) - 1].set_xlabel('Thalweg distance downstream (ft)')

        ymin = 0
        ymax = 0
        col = 1
        for count, signal in enumerate(signals):
            ifft_df = pd.DataFrame()
            ifft_df['raw_series'] = signal
            fft = np.fft.fft(signal)
            if n == 0:
                ifft = np.fft.ifft(fft).real

            elif n != 0 and by_power == False:
                fft = np.fft.fft(signal)
                np.put(fft, range(n, len(fft)), 0.0)
                ifft = np.fft.ifft(fft).real
                cos_coefs = []
                sin_coefs = []
                for index, i in enumerate(fft):
                    if i != 0.0:
                        cos_coefs.append(i.real)
                        sin_coefs.append(i.imag)
                        temp_fft = np.fft.fft(signal)
                        np.put(temp_fft, range(index + 2, len(temp_fft)), 0.0)
                        np.put(temp_fft, range(0, index + 1), 0.0)
                        temp_ifft = np.fft.ifft(temp_fft).real
                        ifft_df['harmonic_%s' % (index + 1)] = temp_ifft
                        if index == (n - 1):
                            ifft_df['all_%s_harmonics' % n] = ifft

            elif n != 0 and by_power == True:
                fft = np.fft.fft(signal)
                psd = np.abs(fft) ** 2
                indices = np.argsort(psd).tolist()
                n_indices = indices[:-n]
                np.put(fft, n_indices, 0.0)
                ifft = np.fft.ifft(fft).real
                cos_coefs = []
                sin_coefs = []
                for index, i in enumerate(fft):
                    if i != 0.0:
                        cos_coefs.append(i.real)
                        sin_coefs.append(i.imag)
                        temp_fft = np.fft.fft(signal)
                        np.put(temp_fft, range(index + 2, len(temp_fft)), 0.0)
                        np.put(temp_fft, range(0, index + 1), 0.0)
                        temp_ifft = np.fft.ifft(temp_fft).real
                        ifft_df['harmonic_%s' % (index + 1)] = temp_ifft
                        if index == (n - 1):
                            ifft_df['all_%s_harmonics' % n] = ifft

            z = key_zs[count]
            z_str = key_z_analysis_functions.float_keyz_format(z)
            ifft_df.to_csv(out_folder + '\\%s_%sft_harmonic_series.csv' % (value_for_fig, z_str))

            for ind, coef in enumerate(cos_coefs):
                row = ind + 2
                ws.cell(row=row, column=col).value = coef
                ws.cell(row=row, column=col + 1).value = sin_coefs[ind]
            col += 2

            print('Cos coefficients for %s Z=%sft: %s' % (value, key_zs[count], cos_coefs))
            print('Sin coefficients for %s, Z=%sft: %s' % (value, key_zs[count], sin_coefs))

            if np.max(ifft) >= ymax or np.max(signal) >= ymax:
                ymax = np.max(np.array([np.max(ifft), np.max(signal)]))
            if np.min(ifft) <= ymin or np.min(signal) <= ymin:
                ymin = np.min(np.array([np.min(ifft), np.min(signal)]))

            ax[count].plot(locs, signal, label='%s signal' % value, color='blue')
            ax[count].plot(locs, ifft, label='Reconstructed %s signal' % value, color='red', linestyle='--')
            if count == 0:
                ax[count].legend(loc='upper center', ncol=2, fontsize=8)
            r_squared = float(np.corrcoef(signal, ifft)[0][1])**2
            ax[count].grid(True, which='both')
            ax[count].text(0.5, 0.2, labels[count], transform=ax[count].transAxes, fontsize=14)
            ax[count].text(0.5, 0.1, ('Pearsons R^2= %s' % round(r_squared, 4)), transform=ax[count].transAxes, fontsize=10)
            ax[count].set_ylabel('%sft stage %s' % (key_zs[count], value))

        wb.save(xl_name)

        if by_power == False:
            fig_name = out_folder + '\\%s_IFFT_N%s_plot.png' % (value_for_fig, lab)

        if by_power == True:
                fig_name = out_folder + '\\%s_IFFT_N%s_by_PSD_plot.png' % (value_for_fig, lab)

        ax[0].set_ylim(ymin, ymax)
        ax[0].set_xlim(0.0, np.max(locs))
        if n == 0:
            fig.suptitle('%s signals and reconstructed inverse-FFT signals' % value, y=0.94)
        else:
            fig.suptitle('%s signals and reconstructed inverse-FFT w/ %s harmonic frequencies signals' % (n, value), y=0.94)
        fig.set_size_inches(12, 6)
        plt.savefig(fig_name, dpi=300, bbox_inches='tight')
        plt.cla()
    plt.close('all')
    print('Correlation plots of inverse Fourier Transform and original signals complete!')


def harmonic_r_square_plot(in_folder, out_folder, detrend_folder, key_zs=[], fields=['W_s_Z_s', 'W_s', 'Z_s'], threshold=0, in_csv='', by_power=False):
    '''This function plots the relationship between the R^2 of the IFFT w/ a given amount of harmonic terms, and the original signal for key zs.
    INPUTS: in_folder containing all_stages_table.csv containing data for each selected key z. out_folder to save fig.
    key_zs can be either float or int.
    fields can be changed from the defaults is csv headers are different.
    threshold (float, 0 is default, max=1) if changed from 0 plots lines marking the # of harmonics where a given R^2 is met
    in_csv allows a specified csv location to be used instead of the default file structure
    by_power (bool, default=False) allows harmonics to be added from highes PSD to lowest'''

    print('Calculating IFFT Pearsons correlation vs # of harmonics')
    value_dict = {}  # Stores Ws and C(Ws,Zs) power spectral density values respectively
    z_str_list = []
    key_zs.sort()

    if len(key_zs) >= 3:
        comb = list(combinations(key_zs, 2))
    else:
        comb = key_zs

    for field in fields:
        value_dict[field] = []
    labels = ['Base flow', 'Bank full', 'Valley Fill']

    join_field = 'loc_%sft' % min(key_z_analysis_functions.find_centerline_nums(detrend_folder))
    threshold_str = str(threshold).split('.')[-1]
    if len(threshold_str) == 1:
        threshold_str = threshold_str + '0'

    if in_csv == '':
        aligned_csv = in_folder + '\\aligned_locations.csv'
    else:
        aligned_csv = in_csv

    aligned_df = pd.read_csv(aligned_csv)
    aligned_df.sort_values(join_field, inplace=True)
    aligned_df.dropna(0, inplace=True)

    for value in value_dict.keys():
        for z in key_zs:
            z_str = key_z_analysis_functions.float_keyz_format(z)
            z_str_list.append(z_str)

            signal = aligned_df.loc[:, [value + '_%sft' % z_str]].squeeze()
            value_dict[value].append(signal)

    for value in value_dict.keys():
        out_df = pd.DataFrame()  # Stores N vs R^2 relationships

        fig, ax = plt.subplots(len(comb), 1, sharex=True, sharey=True)

        if by_power == False:
            if value != 'W_s_Z_s':
                fig_name = out_folder + '\\%s_N_harmonics_r2_t%s_plot.png' % (value, threshold_str)
            else:
                value_for_fig = 'W_s_Z_s'
                fig_name = out_folder + '\\%s_N_harmonics_r2_t%s_plot.png' % (value_for_fig, threshold_str)
        if by_power == True:
            if value != 'W_s_Z_s':
                fig_name = out_folder + '\\%s_N_harmonics_r2_by_PSD_t%s_plot.png' % (value, threshold_str)
            else:
                value_for_fig = 'W_s_Z_s'
                fig_name = out_folder + '\\%s_N_harmonics_r2_by_PSD_t%s_plot.png' % (value_for_fig, threshold_str)

        ax[-1].set_xlabel('# of harmonic terms')
        for count, z in enumerate(key_zs):
            key_z_r_squares = []
            signal = value_dict[value][count]

            if not by_power:
                fft = np.fft.fft(signal)
                for i in range(len(fft)):
                    fft = np.fft.fft(signal)
                    np.put(fft, range(i+1, len(fft)), 0.0)
                    ifft = np.fft.ifft(fft).real
                    r2 = float(np.corrcoef(signal, ifft)[0][1])**2
                    if i == 0:
                        r2 = 0
                    key_z_r_squares.append(r2)

            else:
                fft = np.fft.fft(signal)
                for i in range(len(fft)):
                    fft = np.fft.fft(signal)
                    psd = np.abs(fft) ** 2
                    indices = np.argsort(psd).tolist()
                    n_indices = indices[:-i]
                    np.put(fft, n_indices, 0.0)
                    ifft = np.fft.ifft(fft).real
                    r2 = float(np.corrcoef(signal, ifft)[0][1])**2
                    if i == 0:
                        r2 = 0
                    key_z_r_squares.append(r2)

            if not 'N' in out_df.columns:
                out_df['N'] = np.arange(round(len(fft)/3))

            out_df[labels[count]] = np.asarray(key_z_r_squares)[:round(len(fft)/3)]

            index = 0
            if threshold != 0:
                while key_z_r_squares[index] < threshold and index < len(key_z_r_squares):
                    index += 1

            ax[count].plot(np.arange(round(len(fft)/3)), np.asarray(key_z_r_squares)[:round(len(fft)/3)], color='b')
            ax[count].grid(b=True, which='major', color='grey', linewidth=0.5)
            ax[count].text(0.5, 0.2, ('%sft ' % z) + labels[count], transform=ax[count].transAxes, fontsize=14)
            ax[count].set_ylabel('Pearsons R^2')

            xmax = np.arange(len(fft))[index] / np.arange(round(len(fft)/3))[-1]
            ax[count].axhline(y=key_z_r_squares[index], xmax=xmax, color='r', linestyle='--')
            ax[count].axvline(x=np.arange(len(fft))[index], ymax=key_z_r_squares[index], color='r', linestyle='--')

        print(fig_name[:-9] + '.csv')  # For testing purposes
        out_df.to_csv(fig_name[:-9] + '.csv')
        ax[0].set_xticks(np.arange(0, round(len(fft)/3), 5))
        ax[0].set_yticks(np.arange(0, 1, 0.1))
        ax[0].set_xlim(0.0, round(len(fft)/3))
        ax[0].set_ylim(0.0, 1)
        if threshold == 0:
            fig.suptitle('%s R^2 plot with # of harmonics in a reconstructed signal' % value, y=0.94)
        else:
            fig.suptitle('%s R^2 plot with # of harmonics in a reconstructed signal. Threshold=%s' % (value, threshold), y=0.94)
        fig.set_size_inches(12, 6)
        plt.savefig(fig_name, dpi=300, bbox_inches='tight')
        plt.cla()
    plt.close('all')
    print('Correlation plots of N # of harmonics IFFT and original signals complete!')


def merge_to_csv(in_folder, out_folder, r2_thresholds=[0.90], values=['W_s', 'Z_s', 'W_s_Z_s']):
    """This is a simple function that outputs values for the full sample set csv. Outputs the N number of harmonics associated 0.90 and 0.95 R^2
    thresholds (for by FFT, and PSD). Values can be used for box plotting.
    Threshold must be a list of floats representing R^2 values (i.e 0.9)"""

    prefixs = ['base', 'bf', 'vf']  # Used to match formatting on all sample csv
    values_for_csv = ['ws', 'zs', 'cwz']  # Used to match formatting on all sample csv
    cols = ['Base flow', 'Bank full', 'Valley Fill']  # Input csv column names
    suffix_list = []

    out_x = []
    out_y = []

    for threshold in r2_thresholds:
        threshold_str = str(threshold).split('.')[-1]
        if len(threshold_str) == 1:
            threshold_str = threshold_str + '0'
        suffix_list.append(threshold_str)

    for i, threshold in enumerate(r2_thresholds):
        suffix = suffix_list[i]
        for j, value in enumerate(values):
            val_out = values_for_csv[j]
            csvs = [in_folder + '\\%s_N_harmonics_r2_t%s.csv' % (value, suffix_list[i]),
                    in_folder + '\\%s_N_harmonics_r2_by_PSD_t%s.csv' % (value, suffix_list[i])]  # First represents by FFT method, second is by PSD
            adds = ['', 'PSD_']

            for m, csv in enumerate(csvs):
                add = adds[m]
                df = pd.read_csv(csv)
                for count, col in enumerate(cols):
                    prefix = prefixs[count]
                    col_as_list = df[col].to_list()

                    index = 0
                    if threshold != 0:
                        while col_as_list[index] < threshold and index < len(col_as_list) - 1:
                            index += 1
                    else:
                        print('Error. Must select a non zero R^2 values to find N number of harmonics associated with it')

                    number = df['N'].to_list()[index]  # number of harmonics
                    out_x.append('%s_%s_%sr%s' % (prefix, val_out, add, suffix))
                    out_y.append(number)

    out_dict = {'label': out_x, 'value': out_y}
    out_df = pd.DataFrame.from_dict(out_dict)
    out_df.to_csv(out_folder + '\\harmonics_out.csv')


def class_average_harmonic_curve(class_comid_dict):
    """This function averages, and plots all class averaged N vs R^s curves in square subplots with exagerated Y axis.
    in_folder must contain sub folders with csvs for each comid from the harmonics_r2_analysis. The function averages the csvs as data frames,
    and plots the curves with standard deviation bars"""
    work_on_this = True


comid_list = [17569535, 22514218, 17607553, 17609707, 17609017, 17610661]
SCO_list = ['O1' for i in comid_list]
key_zs_list = [[0.9, 3.0, 5.8], [0.1, 0.9, 5.2], [0.2, 1.1, 2.6], [0.5, 2.0, 5.0], [0.5, 4.2, 7.3], [0.5, 2.1, 8.5]]
signal_process = False

if signal_process:
    for count, comid in enumerate(comid_list):
        key_zs = key_zs_list[count]
        SCO_number = SCO_list[count]
        sc_folder = r"Z:\users\xavierrn\SoCoast_Final_ResearchFiles\SCO%s" % SCO_list[count]
        direct = (r"Z:\users\xavierrn\SoCoast_Final_ResearchFiles\SC%s\COMID%s" % (SCO_number, comid))
        out_folder = direct + r'\LINEAR_DETREND'
        process_footprint = direct + '\\las_footprint.shp'
        table_location = out_folder + "\\gcs_ready_tables"
        channel_clip_poly = out_folder + '\\raster_clip_poly.shp'
        aligned_csv_loc = out_folder + '\\landform_analysis\\aligned_locations.csv'
        landform_folder = out_folder + '\\landform_analysis'
        confine_table = r'Z:\users\xavierrn\Manual classification files\South_200m.shp'
        key_z_dict = {}

        arcpy.env.overwriteOutput = True

        print(comid)
        #powerspec_plotting(in_folder=table_location, out_folder=landform_folder, key_zs=key_zs, fields=['W_s', 'Z_s', 'W_s_Z_s'], smoothing=5)
        #fourier_analysis(in_folder=landform_folder, out_folder=landform_folder, detrend_folder=out_folder, key_zs=key_zs, fields=['W_s_Z_s', 'W_s', 'W', 'Z_s'], n=10, in_csv='', by_power=False)
        #fourier_analysis(in_folder=landform_folder, out_folder=landform_folder, detrend_folder=out_folder, key_zs=key_zs, fields=['W_s_Z_s', 'W_s', 'W', 'Z_s'], n=10, in_csv='', by_power=True)
        harmonic_r_square_plot(in_folder=landform_folder, out_folder=landform_folder, detrend_folder=out_folder, key_zs=key_zs, fields=['W_s_Z_s', 'W_s', 'Z_s'],
                               threshold=0.90, in_csv='', by_power=False)
        #harmonic_r_square_plot(in_folder=landform_folder, out_folder=landform_folder, detrend_folder=out_folder, key_zs=key_zs, fields=['W_s_Z_s', 'W_s', 'Z_s'],
                               threshold=0.90, in_csv='', by_power=True)
        #merge_to_csv(in_folder=landform_folder, out_folder=direct, r2_thresholds=[0.90], values=['W_s', 'Z_s', 'W_s_Z_s'])

