import openpyxl as xl
import arcpy
import os as os
from arcpy import env
from openpyxl import Workbook
from openpyxl import load_workbook
import matplotlib as mpl
from matplotlib import pyplot as plt
import numpy as np
import csv

###### INPUTS ######
# excel file containing xyz data for station points
direct = r'Z:\users\xavierrn\SoCoast_Final_ResearchFiles\SCO2\COMID17586810\window_detrend_test'
xyz_table = direct + '\\XY_elevation_table_300_smooth_3_spaced.xlsx' #change back to 200 to match code!
centerline = direct + '\\las_files\\centerline\\smooth_centerline.shp'
DEM = direct + '\\las_files\\ls_nodt.tif'
#process_footprint = direct + '\\las_footprint.shp'
#detrend_workplace = direct + '\\LINEAR_DETREND_BP1960_3ft_spacing'
#spatial_ref = arcpy.Describe(process_footprint).spatialReference
listofcolumn = ["D", "A", "L", "I", "J"] #For least cost centerlines
######
#Fill lists with necessary data




def prep_xl_file(xyz_table_location, listofcolumn):
    #Import xl file
    id = []
    location = []
    x = []
    y = []
    z = []
    listoflist = [location, id, z, x, y]

    if xyz_table_location[-3:] == "csv":
        print("Input table is a csv, conversion for openpyxl underway...")
        wb = Workbook()
        ws = wb.active
        with open(xyz_table_location, 'r', errors='ignore') as f:
            for row in csv.reader(f):
                ws.append(row)
        wb.save(xyz_table_location[:-3] + "xlsx")
        print("csv converted to xlsx @: %s" % (xyz_table_location[:-3] + "xlsx"))
    else:
        wb = load_workbook(xyz_table_location)
        ws = wb.active
        print("Workbook " + str(xyz_table_location) + " loaded")

    for i in range(0, len(listofcolumn)):
        for cell in ws[listofcolumn[i]]:
            listoflist[i].append(cell.value)
        del listoflist[i][0]

    print(listoflist)

    #define station point spacing, number of station points and index
    point_spacing = int(location[1]) - int(location[0])
    print("Point spacing: " + str(point_spacing))
    number_of_points = int(int(location[-1]) / int(point_spacing))
    print("Number of points: " + str(number_of_points))

    location_np = np.array(location)
    z_np = np.array(z)
    location_np = np.int_(location_np)
    z_np = np.float_(z_np)
    z_np = np.around(z_np, 9)
    print("Z array: %s" % z_np)

    return [location_np,z_np, ws]

def quadratic_fit(location_np, location, z_np, ws):
    print("Applying quadratic fit...")
    constants = np.polyfit(location_np, z_np, 2)

    fit_params = []
    for i in constants:
        fit_params.append(i)
    print(fit_params)

    #Use the found regresssion down channel
    fitted_z = []
    for i in location:
        z_fit_temp = (fit_params[0]*i**(2) + fit_params[1]*i + fit_params[2])
        fitted_z.append(z_fit_temp)

    print(fitted_z)

    residual = []
    if len(fitted_z) == len(z):
        print("Lists are same length, continue as planned...")
        for i in range(len(fitted_z)):
            residual.append(z[i] - fitted_z[i])

    else:
        print("Lists are wrong lengths/not matching, something is wrong...")

    mean_z = (sum(z) / len(z))
    squared_real = []
    squared_res = []

    #Calculate the total sum of squares, sum of residuals, and R^2
    for i in range(len(residual)):
        squared_real.append((z[i] - mean_z)**2)
        squared_res.append(residual[i]**2)

    print("List of squared variance from mean: " + str(squared_real))
    print("List of squared residuals: " + str(squared_res))

    R_squared = 1 - (sum(squared_res)/sum(squared_real))
    print("The coefficient of determination is: " + str(R_squared))

    # Calculate mean and SD for residual to calculate residual z scores as a list
    mean_res = sum(residual)/len(residual)
    sd_res = np.std(residual)
    res_z_score = [(z_res_value - mean_res) * 1.0 / sd_res for z_res_value in residual]
    print(residual)
    print("The mean residual is %.2f" %mean_res)

    # Add fitted z values to the xyz table
    cell_test = ws["D1"]
    print(cell_test.value)
    cell_test.value = "Quadratic fit z values"
    print(cell_test.value)

    if ws["D1"].value == "Quadratic fit z values":
        for i in range(2, len(fitted_z)):
            cell = ws.cell(row=i, column=4)
            cell.value = float(fitted_z[i])
    else:
        print("Something is wrong with writing to the excel sheet")


    wb.save(filename=xyz_table)

    print("Excel file ready for Arc processing!")

def linear_fit(location, z, xyz_table_location, list_of_breakpoints=[]):
    # Applies a linear fit to piecewise sections of the longitudinal profile, each piece is stored in split_list
    # ADD 0 BEFORE ANY ADDED BREAKPOINTS OR THE FUNCTION WILL FAIL!!!!!!!!!!
    print("Applying linear fit")
    list_of_breakpoints.append(int(location[-1]))
    print(list_of_breakpoints)
    fit_params = []
    split_location_list = []
    split_z_list = []
    # Split by breakpoints into a list of lists
    point_spacing = int(location[1]) - int(location[0])

    if xyz_table_location[-3:] == 'csv':
        xyz_table_location = (xyz_table_location[:-3] + "xlsx")

    wb = load_workbook(xyz_table_location)
    ws = wb.active

    #Format input numpy arrays
    location = np.int_(location)
    z = np.float_(z)
    z = np.around(z, 9) # Round z to 9 decimal points

    if len(list_of_breakpoints) > 0:
        slope_break_indices = [int(int(distance)/int(point_spacing)) for distance in list_of_breakpoints]
        for i in range(1, len(slope_break_indices)):
            temp_location_list = []
            temp_z_list = []
            temp_break_index = slope_break_indices[i]
            if i == 1:
                temp_break_index = slope_break_indices[i]
                split_location_list.append(location[:temp_break_index])
                split_z_list.append(z[:temp_break_index])
            #elif i < (len(slope_break_indices)):   FIX THIS SO WE CAN WORK WITH MORE THAN JUST ONE BREAK POINT
                #temp_last_break_index = slope_break_indices[i-1]
                # Splice locations into split_location_list
                #temp_location_list.append(location[temp_last_break_index:temp_break_index])
                #split_location_list.append(temp_location_list)
            else:
                print("In the else: " + str(temp_z_list))
                split_location_list.append(location[slope_break_indices[i-1]:])
                split_z_list.append(z[slope_break_indices[i-1]:])

        print("Breakpoints added...")
        print("Slope break index: " + str(slope_break_indices))
        print("Split location list: " + str(split_location_list))
        print("Split z list: " + str(split_z_list))

        # Get fit parameters for each section of the data
        if len(split_z_list) == len(split_location_list):
            for i in range(len(split_location_list)):
                temp_loc_array_unformatted = np.array(split_location_list[i])
                temp_loc_array = np.int_(temp_loc_array_unformatted)
                temp_z_array_unformatted = np.array(split_z_list[i])
                temp_z_array = np.float_(temp_z_array_unformatted)

                m, b = np.polyfit(temp_loc_array, temp_z_array, 1)
                fit_params.append([m, b])
            print("Fit param list: " + str(fit_params))
        else:
            print("Something went wrong, list lengths do not match...")

    else:
        m, b = np.polyfit(location, z, 1)
        fit_params.append([m, b])
        print("Fit params [m, b]: " + str(fit_params))

    # Make list of lists storing fitted z values
    list_of_lengths = []
    z_fit_list = []
    for i in range(len(split_z_list)):
        list_of_lengths.append(len(split_z_list[i]))

    # Add fitted z's into a list
    print(list_of_lengths)
    i = 0
    while i < len(list_of_lengths):
        print(i)
        print(list_of_lengths[i])
        for j in range(list_of_lengths[i]):
            z_fit_list.append(split_location_list[i][j]*fit_params[i][0]+fit_params[i][1])
        i += 1

    if len(z_fit_list) == len(z):
        print("List lengths are compatible, next we export to excel")
    else:
        print("Something went wrong, length of z =/ z_fit_list")
    print(z_fit_list)

    #Calculate residual and R^2
    residual = []
    for i in range(len(z_fit_list)):
        residual.append(z[i]-z_fit_list[i])
    print(residual)
    mean_z = (sum(z) / len(z))
    squared_real = []
    squared_res = []

    #Calculate the total sum of squares, sum of residuals, and R^2
    for i in range(len(residual)):
        squared_real.append((z[i] - mean_z)**2)
        squared_res.append(residual[i]**2)

    print("List of squared variance from mean: " + str(squared_real))
    print("List of squared residuals: " + str(squared_res))

    R_squared = 1 - (sum(squared_res)/sum(squared_real))
    print("The coefficient of determination is: " + str(R_squared))

    # Calculate mean and SD for residual to calculate residual z scores as a list
    mean_res = sum(residual)/len(residual)
    sd_res = np.std(residual)
    res_z_score = [(z_res_value - mean_res) * 1.0 / sd_res for z_res_value in residual]
    print(residual)
    print("The mean residual is %.2f" % mean_res)

    # Add fitted z values to the xyz table
    '''NOTE: ADJUST SHEET ROW FOR CELL_TEST AND COLUMN FOR THE CELL = WS.CELL COMMAND'''
    cell_test = ws["F1"]
    print(cell_test.value)
    cell_test.value = ("z_fit_%s" % (list_of_breakpoints[1]))
    print(cell_test.value)

    if ws["F1"].value == cell_test.value:
        print("Sheet activated...")
        for i in range(2, len(z_fit_list)):
            cell = ws.cell(row=i, column=6)
            cell.value = float(z_fit_list[i])
    else:
        print("Something is wrong with writing to the excel sheet")
    ws["A1"].value == "OID"


    wb.save(filename=xyz_table_location)

    print("Excel file ready for wetted-polygon processing!")

    return [fit_params, z_fit_list, residual, R_squared]

def moving_window_linear_fit(location, z, xyz_table_location, window_size):
    point_spacing = int(location[1]) - int(location[0]) #Calculate point spacing
    number_of_complete_windows = 0
    while ((number_of_complete_windows*window_size) + window_size) <= location[-1]:
        number_of_complete_windows += 1
    print("Total number of full %sft detrending windows is %s" % (window_size, number_of_complete_windows))

    if xyz_table_location[-3:] == 'csv': #Make sure we are accessing and xlsx file to use openpyxl module
        xyz_table_location = (xyz_table_location[:-3] + "xlsx")

    wb = load_workbook(xyz_table_location)
    ws = wb.active

    # Format input numpy arrays
    location = np.int_(location)
    z = np.float_(z)
    z = np.around(z, 9)  # Round z to 9 decimal points

    #Get window indice value
    window_breakpoints = []
    for num in range(0,number_of_complete_windows):
        window_breakpoints.append(int(num*window_size))
    window_breakpoints.remove(0)
    print("Window breakpoints @ %s" % window_breakpoints)
    window_breakpoint_indices = [int(f/point_spacing) for f in window_breakpoints]
    print("Window breakpoint indices @ %s" % window_breakpoint_indices)

    list_of_fit_params = []
    z_fit_list = []
    window_breakpoint_indices.sort()
    for bp in window_breakpoint_indices:
        if bp == window_breakpoint_indices[0]:
            split_loc_np = np.array(location[:bp])
            split_z_np = np.array(z[:bp])
            m, b = np.polyfit(split_loc_np, split_z_np, 1)
            list_of_fit_params.append([m, b])
            print("Fit params [m, b] for the first window of %sft are: %s" % (window_size, list_of_fit_params[0]))
            for i in location[:bp]:
                z_fit_list.append((i * m) + b) #add fit values to list
        elif bp != window_breakpoint_indices[-1]:
            index = int(window_breakpoint_indices.index(bp))
            last_index = window_breakpoint_indices[int(index - 1)]
            split_loc_np = np.array(location[last_index:bp])
            split_z_np = np.array(z[last_index:bp])
            m, b = np.polyfit(split_loc_np, split_z_np, 1)
            list_of_fit_params.append([m, b])
            print("Fit params [m, b] for window (%s to %s)ft are: %s" % ((point_spacing*window_breakpoint_indices[index-1]),(bp*point_spacing), list_of_fit_params[index]))
            for i in location[last_index:index]:
                z_fit_list.append((i * m) + b)
        elif bp == window_breakpoint_indices[-1]:
            split_loc_np = np.array(location[bp:])
            split_z_np = np.array(z[bp:])
            m, b = np.polyfit(split_loc_np, split_z_np, 1)
            list_of_fit_params.append([m, b])
            print("Fit params [m, b] for the last window from %sft ro the end are: %s" % ((bp*point_spacing), list_of_fit_params[-1]))
            for i in location[bp:]:
                z_fit_list.append((i * m) + b)

    print("Z fit list is ready for excel: %s" % z_fit_list)

    #residual = [location[i]-z_fit_list[i] for i in range(len(location))]

    # Add fitted z values to the xyz table
    cell_test = ws["F1"]
    print(cell_test.value)
    cell_test.value = ("z_fit_window%s" % window_size)
    print(cell_test.value)

    if ws["F1"].value == cell_test.value:
        print("Sheet activated...")
        for i in range(2, len(z_fit_list)): #index begins at two to ignore 0 (doesn't exist in xl) and 1 (xl titles)
            cell = ws.cell(row=i, column=6)
            cell.value = float(z_fit_list[i])
    else:
        print("Something is wrong with writing to the excel sheet")
    ws["A1"].value == "OID"

    wb.save(filename=xyz_table_location)

    print("Excel file ready for wetted-polygon processing!")

    return [z_fit_list]


def detrend_that_raster(detrend_location, fit_z_xl_file, original_dem, stage=0, window_size=0,list_of_breakpoints=[]):
    # Turn fitted xl file to a csv
    wb = xl.load_workbook(fit_z_xl_file)
    ws = wb.active
    arcpy.env.workspace = detrend_location
    arcpy.overwriteoutput = True
    spatial_ref = arcpy.Describe(original_dem).spatialReference
    arcpy.env.extent = arcpy.Describe(original_dem).extent


    csv_name = fit_z_xl_file[:-5] + "_fitted.csv"
    with open(csv_name, 'w', newline="") as f:
        col = csv.writer(f)
        for row in ws.rows:
            col.writerow([cell.value for cell in row])

    if window_size != 0:
        column = ('z_fit_window%s' % window_size)
    elif len(list_of_breakpoints) == 1:
        column = ('z_fit_%s' % list_of_breakpoints[0])
    else:
        column = ('z_fit_breakpoints')

    if stage != 0:
        detrended_raster_file = detrend_location + ("\\rs_dt_s%s.tif" % stage)
    else:
        detrended_raster_file = detrend_location + "\\ras_detren.tif"
        print("0th stage marks non-stage specific centerline")
    points = arcpy.MakeXYEventLayer_management(csv_name, "POINT_X", "POINT_Y", out_layer=("fitted_station_points%s_stage%s" % (i, stage)), spatial_reference=spatial_ref, in_z_field=column)
    points = arcpy.SaveToLayerFile_management(points, ("fitted_station_points%s_stage%sft" % (i, stage)).replace('.csv', '.lyr'))
    points = arcpy.CopyFeatures_management(points)
    print("Creating Thiessen polygons...")
    # Delete non-relevent fields tp reduce errors
    fields = [f.name for f in arcpy.ListFields(points)]
    dont_delete_fields = ['FID', 'Shape', 'POINT_X', 'POINT_Y', column]
    fields2delete = list(set(fields) - set(dont_delete_fields))
    points = arcpy.DeleteField_management(points, fields2delete)

    cell_size1 = arcpy.GetRasterProperties_management(DEM, "CELLSIZEX")
    cell_size = float(cell_size1.getOutput(0))
    thiessen = arcpy.CreateThiessenPolygons_analysis(points, "thiespoly_stage%s.shp" % stage, fields_to_copy='ALL')
    z_fit_raster = arcpy.PolygonToRaster_conversion(thiessen, column, ('theis_raster%s_stage%sft.tif' % (i, stage)), cell_assignment="CELL_CENTER", cellsize=cell_size)
    detrended_DEM = arcpy.Raster(DEM) - arcpy.Raster(z_fit_raster)
    detrended_DEM.save(detrended_raster_file)
    print("DEM DETRENDED!")


###### Define plotting function ######
def diagnostic_quick_plot(location_np, z_np):
    x_plot = location_np
    y_plot = z_np
    plt.plot(x_plot, y_plot, 'r', label="Actual elevation profile")
    plt.xlabel("Thalweg distance downstream (ft)")
    plt.ylabel("Bed elevation (ft)")
    plt.grid(b=True, which='major', color='#666666', linestyle='-')
    plt.minorticks_on()
    plt.grid(b=True, which='minor', color='#999999', linestyle='-', alpha=0.2)
    return plt.show()

def make_quadratic_fit_plot(location_np, z_np, fit_params,stage=0, location=''):
    # Make a plot of the quadratic fit
    # Plot longitudinal elevation profile and detrending fit line
    x_plot = location_np
    y1_plot = z_np
    y2_plot = (fit_params[0]*x_plot**(2) + fit_params[1]*x_plot + fit_params[2])
    plt.plot(x_plot, y1_plot, 'r', label="Actual elevation profile")
    plt.xlabel("Thalweg distance downstream (ft)")
    plt.ylabel("Bed elevation (ft)")
    plt.title("Quadratic detrended elevation profile")
    plt.plot(x_plot, y2_plot, 'b', label= "Elevation profile detrending quadratic: %s * x^2 + %.4f * x + %.4f" % (fit_params[0], fit_params[1], fit_params[2]))
    plt.grid(b=True, which='major', color='#666666', linestyle='-')
    plt.minorticks_on()
    plt.grid(b=True, which='minor', color='#999999', linestyle='-', alpha=0.2)
    plt.legend(loc=1)
    if stage == 0 and location == '':
        plt.show()
    else:
        fig = plt.gcf()
        fig.set_size_inches(12, 6)
        plt.savefig((location + '\\Stage_%s_quadratic_detrend_plot' % stage), dpi=300, bbox_inches='tight')
        plt.cla()

def make_linear_fit_plot(location_np, z_np, fit_params, stage=0, location=''):
    x_plot = location_np
    y1_plot = z_np
    y2_plots = []
    for list in fit_params:
        y2_plots.append(list[0]*x_plot + list[1])
    plt.plot(x_plot, y1_plot, 'r', label="Actual elevation profile")
    plt.xlabel("Thalweg distance downstream (ft)")
    plt.ylabel("Bed elevation (ft)")
    plt.title("Linear piecewise detrended elevation profile")
    for i in range(len(y2_plots)):
        plt.plot(x_plot, y2_plots[i], 'b', label="Elevation profile linear detrending: %.4f * x + %.4f" % (y2_plots[i][0], y2_plots[i][1]))
    plt.grid(b=True, which='major', color='#666666', linestyle='-')
    plt.minorticks_on()
    plt.grid(b=True, which='minor', color='#999999', linestyle='-', alpha=0.2)
    plt.legend(loc=1)

    if stage == 0 and location == '':
        plt.show()
    else:
        fig = plt.gcf()
        fig.set_size_inches(12, 6)
        plt.savefig((location + '\\Stage_%sft_linear_detrend_plot' % stage), dpi=300, bbox_inches='tight')
        plt.cla()

def make_residual_plot(location_np, residual, R_squared, stage=0, location=''):
    '''Plot residuals across longitudinal profile, show R^2. Inputs are a numpy array of location, a list of residuals, and a float for R-squared'''
    x_plot = location_np
    y_plot = np.array(residual)
    y_zero = 0*x_plot
    plt.scatter(x_plot, y_plot, s=1, c='r')
    plt.plot(x_plot, y_zero, c='b')
    plt.xlim(0, max(location_np))
    plt.ylim(min(y_plot), max(y_plot))
    plt.xlabel("Distance down stream centerline (ft)")
    plt.ylabel("Residual")
    plt.title("Residuals: R^2 = %.4f" % R_squared)
    plt.grid(b=True, which='major', color='#666666', linestyle='-')
    plt.minorticks_on()
    plt.grid(b=True, which='minor', color='#999999', linestyle='-', alpha=0.2)

    if stage == 0 and location == '':
        plt.show()
    else:
        fig = plt.gcf()
        fig.set_size_inches(12, 6)
        plt.savefig((location + '\\Stage_%sft_linear_residual_plot' % stage), dpi=300, bbox_inches='tight')
        plt.cla()


################## CALL FUNCTIONS AS NECESSARY ####################

loc = prep_xl_file(xyz_table_location=xyz_table, listofcolumn=listofcolumn)[0]
z = prep_xl_file(xyz_table_location=xyz_table, listofcolumn=listofcolumn)[1]
ws = prep_xl_file(xyz_table_location=xyz_table, listofcolumn=listofcolumn)[2]
#diagnostic_quick_plot(location_np=loc, z_np=z)
#fit_list = linear_fit(location=loc, z=z, xyz_table_location=xyz_table, list_of_breakpoints=[0,6100])
moving_window_linear_fit(location=loc, z=z, xyz_table_location=xyz_table, window_size=500)

#make_linear_fit_plot(location_np=loc, z_np=z, fit_params=fit_list[0], stage=0, location=direct)
#make_residual_plot(location_np=loc, residual=fit_list[2], R_squared=fit_list[3], stage=0, location=direct)
#detrend_that_raster(detrend_location=direct, fit_z_xl_file=xyz_table, original_dem=DEM, stage=0, list_of_breakpoints=[6100])
