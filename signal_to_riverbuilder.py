import pandas as pd
import numpy as np
import openpyxl as xl
import matplotlib
import matplotlib.pyplot as plt
import sys

def ifft_out(signal, fft, ifft_df, n, spacing):
    fft_freqs = np.fft.fftfreq(fft.size, spacing)
    for index, i in enumerate(fft):
        cos_coefs = []
        sin_coefs = []
        freq_list = []
        amp_list = []
        phase_list = []
        if i != 0.0:
            cos_coefs.append(i.real)
            sin_coefs.append(i.imag)
            freq_list.append(fft_freqs[index])
            temp_fft = np.fft.fft(signal)
            np.put(temp_fft, range(index + 2, len(temp_fft)), 0.0)
            np.put(temp_fft, range(0, index + 1), 0.0)
            temp_ifft = np.fft.ifft(temp_fft).real
            amp = abs(np.max(temp_ifft))
            amp_list.append(amp)
            ifft_df['harmonic_%s' % (index + 1)] = temp_ifft

            sub_index = 0
            if temp_ifft[0] < 0:  # Finds when the single FFT component IFFT crossed 0, and therefore it's phase in length units
                while temp_ifft[sub_index] < 0:
                    sub_index += 1
                phase = -1 * sub_index * spacing
            elif temp_ifft[0] > 0:
                while temp_ifft[sub_index] > 0:
                    sub_index += 1
                phase = sub_index * spacing

            phase_list.append(phase)

    return [sin_coefs, cos_coefs, freq_list, amp_list, phase_list, ifft_df]

def by_fft(signal, n, spacing):
    ifft_df = pd.DataFrame()
    ifft_df['raw_series'] = signal
    fft = np.fft.fft(signal)
    if n == 0:
        ifft = np.fft.ifft(fft).real

    fft = np.fft.fft(signal)
    np.put(fft, range(n, len(fft)), 0.0)
    ifft = np.fft.ifft(fft).real
    fft_freqs = np.fft.fftfreq(signal.size, spacing)

    out_list = ifft_out(fft, ifft_df, spacing)
    freqs = out_list[2]
    amps = out_list[3]
    phases = out_list[4]
    ifft_df = out_list[-1]
    ifft_df['all_%s_harmonics' % n] = ifft

    return [ifft, n, out_list[1], out_list[2], ifft_df, freqs, amps, phases]


def by_power(signal, n, spacing):
    ifft_df = pd.DataFrame()
    ifft_df['raw_series'] = signal
    fft = np.fft.fft(signal)
    if n == 0:
        ifft = np.fft.ifft(fft).real

    fft = np.fft.fft(signal)
    psd = np.abs(fft) ** 2
    indices = np.argsort(psd).tolist()
    n_indices = indices[:-n]
    np.put(fft, n_indices, 0.0)
    ifft = np.fft.ifft(fft).real
    fft_freqs = np.fft.fftfreq(signal.size, spacing)

    out_list = ifft_out(fft, ifft_df, spacing)
    freqs = out_list[2]
    amps = out_list[3]
    phases = out_list[4]
    ifft_df = out_list[-1]
    ifft_df['all_%s_harmonics' % n] = ifft

    return [ifft, n, out_list[1], out_list[2], ifft_df, freqs, amps, phases]

def by_power_binned(signal, n, spacing):
    ifft_df = pd.DataFrame()
    ifft_df['raw_series'] = signal
    fft = np.fft.fft(signal)
    if n == 0:
        ifft = np.fft.ifft(fft).real

    psd = (np.abs(fft) ** 2).to_list()
    indices = []

    avg = len(psd) / float(n)
    bins_list = []  # Stores n sub-lists of FFT components from which PSD is calculated
    last = 0.0

    while last < len(psd):
        bins_list.append(psd[int(last):int(last + avg)])
        last += avg

    add = 0
    for sub_list in bins_list:  # Test this to make sure the add thing works out
        add += len(sub_list)
        max_sub_index = psd.index(np.max(psd))
        indices.append(max_sub_index + add)

    np.put(fft, indices, 0.0)
    ifft = np.fft.ifft(fft).real
    fft_freqs = np.fft.fftfreq(signal.size, spacing)

    out_list = ifft_out(fft, ifft_df, spacing)
    freqs = out_list[2]
    amps = out_list[3]
    phases = out_list[4]
    ifft_df = out_list[-1]
    ifft_df['all_%s_harmonics' % n] = ifft

    return [ifft, n, out_list[1], out_list[2], ifft_df, freqs, amps, phases]


def river_builder_harmonics(in_csv, out_folder, index_field, units='', fields=[], field_names=[], r_2=0.95, n=0, methods='ALL', by_power=False, by_bins=False, to_riverbuilder=False):
    """This function plots a N number of Fourier coefficients reconstrution of input signals. Exports coefficients to csv or text file.
    in_csv= A csv file location with evenly spaced values (string).
    sort_by (optional) allows an unsorted csv to be sorted by a input index field header (string)
    out_folder= Folder to which plots and exported coefficients are saved (string).
    index_field is the csv header corresponding to the centerline position, the units parameter can be any length unit and is used strictly for plotting (empty is default).
    fields= A list of csv headers from which signals are plotted and reconstructed (list of strings)
    field_names (optional) if specified must be a list of strings with names for plotting titles
    R_2= R^2 threshold for signal reconstruction (float). 0.95 is default.
    n (0 is default) if not 0 and an int allows a number of Fourier components to specified (int), as opposed to the standard R^2 threshold (default)
    by_power (False is default) if True takes the N highest power components first
    by_bins (False is default) if True splits the FFT components into N bins, and selects the highest power frequency from each
    to_riverbuilder (False"""

    in_df = pd.read_csv(in_csv)

    try:
        in_df.sort_values(index_field, inplace=True)
        index_array = in_df.loc[:, index_field].to_array()
        spacing = float(index_array[1] - index_array[0])
    except:
        print('Could not sort values by  the input index field header: %s. Please either remove sort_by parameter, or correct the input field header.' % sort_by)
        sys.exit()

    if len(fields) == 0:
        print('Error! No field headers input.')
        sys.exit()

    if methods == 'ALL':
        methods_dict = {'by_fft': [], 'by_power': [], 'by_power_binned': []}  # Each list associated with each method stores [ifft, n, sin_coefs, cos_coefs, ifft_df]
    else:
        methods_dict = {methods: []}

    else:
        for count, field in enumerate(fields):
            try:
                field_signal = in_df.loc[:, str(field)].to_array()
            except:
                print('Error! Could not use csv field headers input. Please check csv headers.' % field)

            if len(field_names) == len(fields):
                field_name = field_names[count]
            elif count == 0:
                field_names = []
                field_names.append(field)
                field_name = field_names[count]
            else:
                field_names.append(field)
                field_name = field_names[count]

            if n == 0 and r_2 > 0:
                for method in methods_dict.keys():
                    if method == 'by_fft':
                        for i in range(1, len(field_signal)):
                                out_list = by_fft(field_signal, i, spacing)
                                temp_r2 = np.corrcoef(field_signal, out_list[0])[0][1] ** 2
                                if temp_r2 >= r_2:
                                    for out in out_list:
                                        methods_dict[method].append(out_list)
                                    break

                    if method == 'by_power':
                        for i in range(1, len(field_signal)):
                            out_list = by_power(field_signal, i, spacing)
                            temp_r2 = np.corrcoef(field_signal, out_list[0])[0][1] ** 2
                            if temp_r2 >= r_2:
                                for out in out_list:
                                    methods_dict[method].append(out_list)
                                break

                    if method == 'by_power_binned':
                        for i in range(1, len(field_signal)):
                            out_list = by_power_binned(field_signal, i, spacing)
                            temp_r2 = np.corrcoef(field_signal, out_list[0])[0][1] ** 2
                            if temp_r2 >= r_2:
                                for out in out_list:
                                    methods_dict[method].append(out_list)
                                break

            else:
                for method in methods_dict.keys():
                    if method == 'by_fft':
                        out_list = by_fft(field_signal, n, spacing)
                        temp_r2 = np.corrcoef(field_signal, out_list[0])[0][1] ** 2
                        for out in out_list:
                            methods_dict[method].append(out_list)
                            methods_dict[method].append(temp_r2)  # Stores the R^2 value associated with the given method and set n number of harmonics value

                    if method == 'by_power':
                        out_list = by_power(field_signal, n, spacing)
                        temp_r2 = np.corrcoef(field_signal, out_list[0])[0][1] ** 2
                        for out in out_list:
                            methods_dict[method].append(out_list)
                            methods_dict[method].append(temp_r2)

                    if method == 'by_power':
                        out_list = by_power(field_signal, n, spacing)
                        temp_r2 = np.corrcoef(field_signal, out_list[0])[0][1] ** 2
                        for out in out_list:
                            methods_dict[method].append(out_list)
                            methods_dict[method].append(temp_r2)


        for method in methods_dict.keys():
            wb = xl.Workbook()
            wb.save(out_folder + '\\harmonics_coefs_%s' % method)

            for count, field in enumerate(fields):
                text_file = open(out_folder + '%s_%s_to_riverbuilder.txt' % (field, method), 'w+')
                field_name = field_names[count]
                if count == 0:
                    ws = wb.active
                    ws.title = field_names[count]
                else:
                        wb.create_sheet(field_names[count])
                        ws = wb[field_names[count]]

                list = methods_dict[method][count]
                list[4].to_csv(out_folder + '\\%s_harmonics_%s.csv' % (field, method))

                ws.cell(row=1, column=1).value = 'Cosine coefs'
                ws.cell(row=1, column=1).value = 'Sine coefs'

                for ind, coef in enumerate(list[3]):
                    row = ind + 2
                    ws.cell(row=row, column=1).value = coef
                    ws.cell(row=row, column=2).value = list[2][ind]

                plt.plot(index_array, in_df.loc[:, str(field)].to_array(), color='blue', label='Signal')
                plt.plot(index_array, list[0], color='red', linestyle='--', label='Reconstructed signal')

                if units != '':
                    add_units = 'in %s' % units
                else:
                    add_units = ''
                plt.xlabel('Distance along centerline %s' % add_units)
                plt.ylabel('Value')
                plt.title('%s, %s method, N=%s component harmonic reconstruction' % (field_name, method, list[1]))
                plt.grid(b=True, which='major', color='#666666', linestyle='-')
                plt.minorticks_on()
                plt.legend(loc='lower center')

                for num, amp in enumerate(list[-2]):
                    text_file.write('COS%s=(%s, %s, %s, MASK0)' % (num, amp, list[-3], list[-1]))  # Writes in the form of COS#=(a, f, ps, MASK0) for river builder inputs
                text_file.save()
                text_file.close()
                
            wb.save(out_folder + '\\harmonics_coefs_%s' % method)
                




                





