import arcpy
import csv
import os
import scipy
import math
from scipy.stats import variation
from os import listdir
from os.path import isfile, join
from matplotlib import pyplot as plt
import matplotlib.colors as colors_module
from mpl_toolkits.axes_grid1.axes_divider import make_axes_locatable
import matplotlib.cm
import plotly
import plotly.graph_objects as go
import plotly.express as pex
from itertools import combinations
import seaborn as sns
import numpy as np
import file_functions
import descriptive_statistics_functions
import create_station_lines
from create_station_lines import *
import GCS_analysis
import pandas
import openpyxl as xl
import gcs_generating_functions
import classify_landforms_GUI
import DEM_detrending_functions


def flip_tables(table_folder, aligned_table):
    """This function can be used to flip the index on the gcs csv out tables when dist_down corresponds to dist upstream by error."""
    table_list = [i for i in listdir(table_folder) if i[-4:] == '.csv']

    for table in table_list:
        table_loc = table_folder + '\\%s' % table
        table_df = pd.read_csv(table_loc)
        max_dist = np.nanmax(table_df.loc[:, 'dist_down'].to_numpy())
        dist_list = table_df.loc[:, ['dist_down']].squeeze().to_list()
        loc_np = np.array([int(max_dist - i) for i in dist_list])
        table_df['dist_down'] = loc_np
        table_df.sort_values('dist_down', inplace=True)
        table_df.to_csv(table_loc)

    print('non-aligned GCS tables distance down field flipped!')

    aligned_df = pd.read_csv(aligned_table)
    loc_fields = [j for j in list(aligned_df.columns.values) if j[:3] == 'loc']
    loc_nums = []

    for loc_field in loc_fields:
        if loc_field[5] == 'f':
            loc_nums.append(loc_field[4])
        else:
            loc_nums.append(loc_field[4:6])
        temp_max = np.nanmax(aligned_df.loc[:, loc_field].to_numpy())
        dist_list = aligned_df.loc[:, [loc_field]].squeeze().to_list()
        loc_np = np.array([int(temp_max - i) for i in dist_list])
        aligned_df[loc_field] = loc_np

    min_loc = loc_fields[loc_nums.index(min(loc_nums, key=int))]
    aligned_df.sort_values(str(min_loc), inplace=True)
    aligned_df.to_csv(aligned_table)

    print('Aligned locations table flipped successfully!')


def gcs_plotter(table_folder, out_folder, key_zs, key_z_meanings=['Baseflow', 'Bankfull', 'Valley Fill'], fields=['W_s', 'Z_s', 'W_s_Z_s']):
    """This function makes longitudinal profile plots for given fields across each key z saving them to a folder.
     If aligned_table is defined as the aligned csv, plots showing each key z profile as sub-plots for a given field are saved as well."""
    key_zs.sort()
    colors = ['black', 'blue', 'grey', 'orange', 'red']
    landforms = ['Oversized', 'Const. Pool', 'Normal', 'Wide Bar', 'Nozzle']
    print('Plotting fields: %s for each key z' % fields)
    for field in fields:
        xs = []
        ys = [[] for i in key_zs]
        full_ys = []
        for count, z in enumerate(key_zs):
            ys.append([])
            z_str = float_keyz_format(z)
            table_loc = table_folder + '\\%sft_WD_analysis_table.csv' % z_str
            table_df = pd.read_csv(table_loc)
            table_df.sort_values('dist_down', inplace=True)
            xs.append(table_df.loc[:, 'dist_down'].to_numpy())
            codes = table_df.loc[:, 'code'].to_numpy()

            for num in range(-2, 3):  # Make arrays representing each landform type
                y_temp = table_df.loc[:, field].to_numpy()
                y_temp[codes != num] = np.nan
                ys[count].append(y_temp)
                table_df = pd.read_csv(table_loc)
                table_df.sort_values('dist_down', inplace=True)

            full_ys.append(table_df.loc[:, field].to_numpy())

        fig, ax = plt.subplots(len(key_zs), sharey=True)
        fig.subplots_adjust(hspace=0.4)
        fig_name = out_folder + '\\%s_GCS_plots.png' % field
        ax[0].set_title('%s signals' % field)

        for count, z in enumerate(key_zs):
            x = xs[count]
            ymax = 0
            ax[count].plot(x, full_ys[count], color=colors[2])
            for i, y in enumerate(ys[count]):
                ax[count].plot(x, y, color=colors[i], label=landforms[i])
                temp_max = np.amax(np.array([np.abs(np.nanmin(y)), np.abs(np.nanmax(y))]))
                if temp_max >= ymax and ymax <= 5:
                    ymax = math.ceil(temp_max)
                elif ymax > 5:
                    ymax = 5
            ax[count].set_ylim(-1*ymax, ymax)
            ax[count].set_ylabel(field)
            ax[count].set_title(key_z_meanings[count])
            ax[count].set_yticks(np.arange(-1 * ymax, ymax, 1), minor=False)
            ax[count].grid(True, which='both', color='gainsboro', linestyle='--')
            ax[count].set_xlim(0.0, np.max(x))
            ax[count].set_xticks(np.arange(0, np.max(x), 250))

        ax[count].set_xlabel('Thalweg distance downstream (ft)')
        ax[count].legend(loc='lower center', ncol=len(landforms), fontsize=8)  # Adds legend to the bottom plot
        fig.set_size_inches(12, 6)
        plt.savefig(fig_name, dpi=300, bbox_inches='tight')
        plt.cla()
    plt.close('all')
    print('GCS plots saved @ %s' % out_folder)


def box_plots(in_csv, out_folder, fields=[], field_units=[], field_title='fields', sort_by_field='', sort_by_title='', ylim=0, single_plots=False):
    """This function takes a csv and creates box and whisker plots for each field.
    in_csv must contained data readable by pandas. Out folder is where the plots are saved.
    Fields is a list of fields to make plots from. If single_plots==True(False is default), each field is on the same plots for a single sort_by_field.
    sort_by_field (int, or str) is the field that contains values to separate plots or for comparison (ex: log(catchment area) or geo_class.
    field_units is a list that MUST be filled with units corresponding to each input field
    field_title comes into play when single_plots == True and designates the plot title and filename. Default is 'fields'
    ! If single_plots=False, a plot is made comparing each class for a given column, if single_plots=True all fields are shown on one plot for each class (assuming sorted by class)"""
    in_df = pd.read_csv(in_csv)
    box_dict = {}  # Either formatted as fields:[[sort_unique1],[],...] if single_plots==False OR sort_uniques:[[field1],[],...] if single_plots == True

    if sort_by_title == '':  # Allows plot titles to be manually defined
        sort_by_title = sort_by_field

    if not os.path.exists(out_folder):
        os.makedirs(out_folder)

    sort_uniques = in_df[sort_by_field].unique()

    print('Making box and whisker plots...')
    if single_plots == False:
        for field in fields:
            box_dict[field] = []

            for unique in sort_uniques:
                temp_df = in_df.query('%s == "%s"' % (sort_by_field, unique))
                temp_np = temp_df.loc[:, field].to_numpy()
                filtered_np = temp_np[~np.isnan(temp_np)]
                box_dict[field].append(filtered_np)

        for count, field in enumerate(fields):
            plot_dir = '\\%s_by_%s_boxplots.png' % (field, sort_by_field)
            fig, ax = plt.subplots()
            ax.set_title('%s sorted by %s' % (field, sort_by_field))
            ax.set_xlabel(sort_by_title)
            ax.set_ylabel(field_units[count])

            if ylim != 0:
                ax.set_ylim(0, ylim)

            plot = plt.boxplot(box_dict[field], medianprops=dict(color='black'), patch_artist=True, showfliers=False)
            colors = ['royalblue', 'yellow', 'indianred', 'violet', 'limegreen']
            for patch, color in zip(plot['boxes'], colors):
                patch.set_facecolor(color)
            plt.savefig(out_folder + plot_dir, dpi=400, bbox_inches='tight')
            plt.close(fig)

    if single_plots == True:
        print('MAKE SURE ALL FIELDS HAVE THE SAME UNITS and it is input as a single value in the field_units list!')

        for unique in sort_uniques:
            box_dict[unique] = []
            temp_df = in_df.query('%s == "%s"' % (sort_by_field, unique))
            for field in fields:
                temp_np = temp_df.loc[:, field].to_numpy()
                filtered_np = temp_np[~np.isnan(temp_np)]
                box_dict[unique].append(filtered_np)

        for count, unique in enumerate(sort_uniques):
            plot_dir = '\\%s_%s_by_stage_boxplots.png' % (unique, field_title)
            fig, ax = plt.subplots()
            ax.set_title('%s %s values' % (unique, field_title))
            ax.set_xlabel('Stage')
            ax.set_ylabel(field_units[0])
            ax.set_xticks([1, 2, 3], ['Base', 'BF', 'VF'])
            ax.set_xticklabels(['Base', 'BF', 'VF'])

            if ylim != 0:
                ax.set_ylim(0, ylim)

            plot = plt.boxplot(box_dict[unique], widths=0.6, medianprops=dict(color='black'), patch_artist=True, showfliers=False)
            colors = ['b', 'y', 'r', 'purple', 'g']
            for patch, color in zip(plot['boxes'], colors):
                patch.set_facecolor(color)
            plt.savefig(out_folder + plot_dir, dpi=400, bbox_inches='tight')
            plt.close(fig)

    print('Plots saved @ %s' % out_folder)


def landform_pie_charts(base_folder, comids, out_folder, class_title='', geo_classes=[], all_key_zs=[]):
    """This function plots relative landform abundance across key z stages as pie sub-plots. If len(comids) > 1, values are averaged across multiple COMIDS.
    base_folder must contain COMID#### folders. comids must be a list of ints. out_folder is the path to which figures are saved. class_title, if comids in list
    represent a geomorphic class can be set to a string for plot labeling. geo_classes must be a list with the same length of comids, containing \\SC0# strings.
    key_zs must be a list of lists containing key z stage heights for each input comid."""
    titles = {0: 'Baseflow', 1: 'Bankfull', 2: 'Valley Fill'}
    labels = ['Oversized', 'Const.Pool', 'Normal', 'Wide Bar', 'Nozzle']
    colors = ['black', 'blue', 'grey', 'orange', 'red']

    if not os.path.exists(out_folder):
        os.makedirs(out_folder)

    if isinstance(comids, int) or len(comids) == 1:
        if isinstance(comids, int):
            comid = comids
        else:
            comid = comids[0]
        percents = []  # [oversized, cons. pool, normal, wide bar, nozzle] as a sub list for each key z stage
        if isinstance(all_key_zs[0], float):
            key_zs = all_key_zs
        elif isinstance(all_key_zs[0], list):
            key_zs = all_key_zs[0]

        geo_class = geo_classes[0]

        for index, z in enumerate(key_zs):
            z_str = float_keyz_format(z)
            temp_percents = []

            data = base_folder + '\\DATASET_by_reach\\%s\\COMID%s\\Tables\\%sft_stats_table.xlsx' % (geo_class, comid, z_str)

            wb = xl.load_workbook(data)
            ws = wb.active

            for row_num in range(8, 48, 8):
                temp_percents.append(float(ws.cell(row=row_num, column=8).value))
            percents.append(np.array(temp_percents))

        fig, axs = plt.subplots(ncols=int(len(key_zs)), figsize=(10, 3))
        fig.subplots_adjust(wspace=0.2, left=0.07, right=0.93)

        for count, ax in enumerate(axs):
            ax.pie(percents[count], labels=labels, labeldistance=None, autopct='%1.1f%%', textprops={'color': "w"}, colors=colors)
            ax.set_title(titles[count])
            ax.title.set_position([0.5, 0.92])

        axs[1].legend(bbox_to_anchor=(0.32, 0.07), bbox_transform=ax.transAxes, ncol=len(labels), fontsize=8)  # Adds legend to the bottom plot

        save_title = out_folder + '\\comid%s_piecharts.png' % comid
        fig = plt.gcf()
        fig.set_size_inches(10, 10)
        plt.savefig(save_title, dpi=300, bbox_inches='tight')
        plt.clf()
        plt.close('all')
        print('Pie plots for COMID%s landform abundances @ %s' % (comid, save_title))

    if isinstance(comids, list) and len(comids) > 1:
        totals = [[], [], []]
        for i in totals:
            for land in labels:
                i.append(0)

        for count, comid in enumerate(comids):
            geo_class = geo_classes[count]
            key_zs = all_key_zs[count]

            for index, z in enumerate(key_zs):
                z_str = float_keyz_format(z)
                temp_percents = []

                data = base_folder + '\\DATASET_by_reach\\%s\\COMID%s\\Tables\\%sft_stats_table.xlsx' % (geo_class, comid, z_str)

                wb = xl.load_workbook(data)
                ws = wb.active

                for ind, row_num in enumerate(range(8, 48, 8)):
                    totals[index][ind] += float(ws.cell(row=row_num, column=8).value)

        percents = []
        for sub_list in totals:
            percents.append([float(i/5) for i in sub_list])

        fig, axs = plt.subplots(ncols=int(len(key_zs)), figsize=(10, 3))
        fig.subplots_adjust(wspace=0.2, left=0.07, right=0.93)

        for count, ax in enumerate(axs):
            ax.pie(percents[count], labels=labels, labeldistance=None, autopct='%1.1f%%', textprops={'color': "w"}, colors=colors)
            ax.set_title(titles[count])
            ax.title.set_position([0.5, 0.92])

        axs[1].legend(bbox_to_anchor=(0.32, 0.07), bbox_transform=ax.transAxes, ncol=len(labels), fontsize=8)  # Adds legend to the bottom plot

        save_title = out_folder + '\\%s_piecharts.png' % class_title
        fig = plt.gcf()
        fig.set_size_inches(10, 10)
        plt.savefig(save_title, dpi=300, bbox_inches='tight')
        plt.clf()
        plt.close('all')
        print('Pie plots for %s landform abundances @ %s' % (class_title, save_title))


def nested_landform_analysis(aligned_csv, key_zs):
    '''IN: An aligned csv with landform codes for each XS.
    A list (key_zs) containing three stages. All key_zs must already be aligned into table.
    RETURNS: A xl table containing the abundance of each unique nested landform set'''
    landform_folder = str(os.path.dirname(aligned_csv))
    code_dict = {-2: 'O', -1: 'CP', 0: 'NC', 1: 'WB', 2: 'NZ'}  # code number and corresponding MU
    print('Starting nested landform analysis...')

    if len(key_zs) == 0:
        return print('No key Zs selected, please add parameters')

    key_zs.sort()
    aligned_df = pd.read_csv(aligned_csv)
    data = aligned_df.dropna()

    code_df_list = []  # code_df_list[0] is baseflow, [1] is bankful, and [2] is flood stage
    for key_z in key_zs:
        if isinstance(key_z, float):
            if key_z >= 10.0:
                key_z = (str(key_z)[0:2] + 'p' + str(key_z)[3])
            else:
                key_z = (str(key_z)[0] + 'p' + str(key_z)[2])
        code_df_temp = data.loc[:, [('code_%sft' % key_z)]].squeeze()
        code_df_list.append(code_df_temp.values.tolist())

    nested_landforms = list(zip(code_df_list[0], code_df_list[1], code_df_list[2]))
    unique_nests = list(set(nested_landforms))

    unique_nest_counts = list(np.zeros(len(unique_nests), dtype=int))  # initialize list of lists to count abundance

    for nest in nested_landforms:
        i = unique_nests.index(nest)
        unique_nest_counts[i] += 1

    nest_abundances = list(zip(unique_nests, unique_nest_counts))
    nest_abundances.sort(key=lambda x: x[1], reverse=True)

    nested_analysis_xl = (landform_folder + '\\nested_landforms.xlsx')
    wb = xl.Workbook()
    wb.save(nested_analysis_xl)
    ws = wb.active
    ws.title = 'Nested landforms abundances'
    ws.cell(row=1, column=1).value = 'Nested landform set [baseflow, BF, flood]'
    ws.cell(row=1, column=2).value = 'Abundances'
    ws.cell(row=1, column=3).value = '% of unique sets'
    ws.cell(row=1, column=4).value = 'Total unique sets(125 possible):'
    ws.cell(row=2, column=4).value = len(unique_nests)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['D'].width = 20

    for count, unique_set in enumerate(nest_abundances):
        string = '%s, %s, %s' % (code_dict[unique_set[0][0]], code_dict[unique_set[0][1]], code_dict[unique_set[0][2]])
        ws.cell(row=2 + count, column=1).value = str(string)
        ws.cell(row=2 + count, column=2).value = unique_set[1]
        ws.cell(row=2 + count, column=3).value = round((unique_set[1] / len(nested_landforms)) * 100, 2)

    wb.save(nested_analysis_xl)
    wb.close()
    print('Nested landform analysis complete. Results @ %s' % nested_analysis_xl)


def nested_landform_sankey(base_folder, comids, out_folder, class_title='', geo_classes=[], all_key_zs=[], ignore_normal=False):
    """Creates Sankey diagrams showing nested landform relationships. Can be done across a class, transition occurences are normalized as a % for each reach."""
    code_dict = {-2: 'O', -1: 'CP', 0: 'NC', 1: 'WB', 2: 'NZ'}  # code number and corresponding MU
    print('Sankey landform diagram plotting comencing...')

    source = []
    target = []
    value = []

    if len(comids) == 1 and class_title == '':
        class_title = 'comid%s' % comids[0]
    elif class_title.isdigit():
        class_title = 'class%s' % class_title

    if not ignore_normal:
        title = '%s\\%s_sankey_plot' % (out_folder, class_title)
    else:
        title = '%s\\%s_sankey_plot_no_normal' % (out_folder, class_title)

    for k, comid in enumerate(comids):
        geo_class = geo_classes[k]
        key_zs = all_key_zs[k]
        key_zs.sort()

        aligned_df = pd.read_csv(base_folder + '\\DATASET_by_reach\\%s\\COMID%s\\Tables\\aligned_locations.csv' % (geo_class, comid))
        data = aligned_df.dropna()

        code_df_list = []  # code_df_list[0] is baseflow, [1] is bankful, and [2] is flood stage
        for key_z in key_zs:
            if isinstance(key_z, float):
                if key_z >= 10.0:
                    key_z = (str(key_z)[0:2] + 'p' + str(key_z)[3])
                else:
                    key_z = (str(key_z)[0] + 'p' + str(key_z)[2])

            code_df_temp = data.loc[:, [('code_%sft' % key_z)]].squeeze()
            code_df_list.append(code_df_temp.values.tolist())

        base_to_bf = list(zip(code_df_list[0], code_df_list[1]))
        bf_to_vf = list(zip(code_df_list[0], code_df_list[2]))
        both = [base_to_bf, bf_to_vf]

        unique_nests = [list(set(base_to_bf)), list(set(bf_to_vf))]
        unique_nest_counts = [list(np.zeros(len(i), dtype=int)) for i in unique_nests]  # initialize list of lists to count abundance

        for j, zipped in enumerate(both):  # Calculates totals of occurences for each possible base->bf, and bf->vf combination
            for pair in zipped:
                    i = unique_nests[j].index(pair)
                    unique_nest_counts[j][i] += 1

        nest_abundances = [list(zip(unique_nests[0], unique_nest_counts[0])), list(zip(unique_nests[1], unique_nest_counts[1]))]

        if ignore_normal == False:
            nodes = {
                    "label": ['Oversized', 'Const.Pool', 'Normal', 'Wide Bar', 'Nozzle', 'Oversized', 'Const.Pool', 'Normal', 'Wide Bar', 'Nozzle', 'Oversized', 'Const.Pool', 'Normal', 'Wide Bar', 'Nozzle'],
                    "x": [0.1, 0.1, 0.1, 0.1, 0.1, 0.5, 0.5, 0.5, 0.5, 0.5, 0.9, 0.9, 0.9, 0.9, 0.9],
                    "y": [0.1, 0.3, 0.5, 0.7, 0.9, 0.1, 0.3, 0.5, 0.7, 0.9, 0.1, 0.3, 0.5, 0.7, 0.9],
                    "color": ['black', 'blue', 'grey', 'orange', 'red', 'black', 'blue', 'grey', 'orange', 'red', 'black', 'blue', 'grey', 'orange', 'red'],
                    'pad': 15}
        else:
            nodes = {
                "label": ['Oversized', 'Const.Pool',  'Wide Bar', 'Nozzle', 'Oversized', 'Const.Pool', 'Wide Bar', 'Nozzle', 'Oversized', 'Const.Pool',  'Wide Bar', 'Nozzle'],
                "x": [0.1, 0.1, 0.1, 0.1, 0.5, 0.5, 0.5, 0.5, 0.9, 0.9, 0.9, 0.9],
                "y": [0.2, 0.4, 0.6, 0.8, 0.2, 0.4,  0.6, 0.8, 0.2, 0.4, 0.6, 0.8],
                "color": ['black', 'blue', 'orange', 'red', 'black', 'blue', 'orange', 'red', 'black',
                          'blue', 'orange', 'red'],
                'pad': 15}

        for j, nests in enumerate(nest_abundances):

            for i in nests:
                if ignore_normal == False:
                    index_adjust = 2 + j * 5
                    source.append(int(i[0][0] + index_adjust))
                    target.append(int(i[0][1] + index_adjust + 5))
                    value.append(float(i[1] + index_adjust))
                elif 0 not in i[0]:
                    index_adjust = 2 + j * 4
                    if int(i[0][0]) < 0:
                        source.append(int(i[0][0] + index_adjust))
                    else:
                        source.append(int(i[0][0] + index_adjust - 1))
                    if int(i[0][1]) < 0:
                        target.append(int(i[0][1] + index_adjust + 4))
                    else:
                        target.append(int(i[0][1] + index_adjust + 3))
                    value.append(float(i[1] + index_adjust))

    fig = go.Figure(go.Sankey(
        arrangement="snap",
        node=nodes,  # 10 Pixels
        link={
            "source": source,
            "target": target,
            "value": value}))

    fig.write_image(title + '.pdf')
    fig.write_image(title + '.png', scale=5)
    print('Sankey plots saved @ %s' % title)


def heat_plotter(base_folder, comids, out_folder, class_title='', geo_classes=[], key_zs=[]):
    '''IN: a list with either one or multiple comids. Key_zs list if filled makes subplots, if not a plot for each stage is made.
    If multiple comids are present, key_zs list structure is important. EXAMPLE: comids= [123,456,789], key_zs = [[baseflow, BF, flood],[baseflow, BF, flood],[baseflow, BF, flood]
    Heatmaps plotted and saved in landform folder, or new folder for class averaged figures'''
    titles = {0: 'Baseflow', 1: 'Bankfull', 2: 'Valley Fill'}

    if len(comids) > 1:
        if class_title == '':
            class_title = 'multiple_reaches'
        if len(geo_classes) != len(comids):
            print('Enter a sub folder string associated with a COMID[comid]')
        if len(key_zs) != len(comids):
            print('Enter a sub list of key zs (ex: key_zs=[[0.5, 1.2, 4.9]...]) for each input reach comid')

        x_list_of_arrays = [[], [], []]  # Initialize list containing [baseflow, bankful, valley fill] values
        y_list_of_arrays = [[], [], []]

        for count, comid in enumerate(comids):
            geo_class = geo_classes[count]

            for index, z in enumerate(key_zs[count]):
                z_str = float_keyz_format(z)

                data = pd.read_csv(base_folder + '\\DATASET_by_reach\\%s\\COMID%s\\Tables\\%sft_WD_analysis_table.csv' % (geo_class, comid, z_str))
                x_temp = data.loc[:, ['W_s']].squeeze().to_list()
                y_temp = data.loc[:, ['Z_s']].squeeze().to_list()
                for value in range(len(x_temp)):
                    x_list_of_arrays[index].append(x_temp[value])
                    y_list_of_arrays[index].append(y_temp[value])

        fig, axs = plt.subplots(ncols=int(len(key_zs[0])), figsize=(10, 3))
        fig.subplots_adjust(hspace=0.5, wspace=0.3, left=0.07, right=0.93)

        for count, ax in enumerate(axs):
            x = np.asarray(x_list_of_arrays[count])
            y = np.asarray(y_list_of_arrays[count])

            ax.set_aspect('equal', adjustable='box')
            ax.hexbin(x, y, gridsize=30, cmap='YlOrRd', extent=(-3, 3, -3, 3))
            ax.set(xlim=(-3, 3), ylim=(-3, 3))
            ax.axhline(y=0.5, xmin=0, xmax=0.4167, color='#9e9e9e', linestyle='--')
            ax.axhline(y=0.5, xmin=0.583, xmax=1, color='#9e9e9e', linestyle='--')
            ax.axhline(y=-0.5, xmin=0, xmax=0.4167, color='#9e9e9e', linestyle='--')
            ax.axhline(y=-0.5, xmin=0.583, xmax=1, color='#9e9e9e', linestyle='--')

            ax.axvline(x=-0.5, ymin=0, ymax=0.4167, color='#9e9e9e', linestyle='--')
            ax.axvline(x=-0.5, ymin=0.583, ymax=1, color='#9e9e9e', linestyle='--')
            ax.axvline(x=0.5, ymin=0, ymax=0.4167, color='#9e9e9e', linestyle='--')
            ax.axvline(x=0.5, ymin=0.583, ymax=1, color='#9e9e9e', linestyle='--')

            ax.text(0.20, 0.05, 'Const. Pool', horizontalalignment='center', verticalalignment='center',
                    transform=ax.transAxes)
            ax.text(0.18, 0.95, 'Nozzle', horizontalalignment='center', verticalalignment='center',
                    transform=ax.transAxes)
            ax.text(0.82, 0.95, 'Wide Bar', horizontalalignment='center', verticalalignment='center',
                    transform=ax.transAxes)
            ax.text(0.82, 0.05, 'Oversized', horizontalalignment='center', verticalalignment='center',
                    transform=ax.transAxes)
            ax.set_xlabel('Standardized width (Ws)')
            ax.set_ylabel('Standardized detrended elevation (Zs)')

        save_title = out_folder + '\\class%s_heatplots.png' % class_title
        fig = plt.gcf()
        fig.set_size_inches(10, 10)
        plt.savefig(save_title, dpi=300, bbox_inches='tight')
        plt.clf()
        plt.close('all')
        print('Plot comparing key Zs for all class %s reaches completed. Located @ %s' % (class_title, save_title))

    elif len(key_zs) != 0:
        fig, axs = plt.subplots(ncols=int(len(key_zs)), figsize=(10, 3))
        fig.subplots_adjust(hspace=0.5, wspace=0.3, left=0.07, right=0.93)

        for count, ax in enumerate(axs):
            z = key_zs[count]
            z_str = float_keyz_format(z)

            data = pd.read_csv(base_folder + '%s\\COMID%s\\LINEAR_DETREND\\gcs_ready_tables\\%sft_WD_analysis_table.csv' % (geo_classes[0], comids[0], z_str))
            data = data.loc[:, ~data.columns.str.contains('^Unnamed')]

            x = data.loc[:, ['W_s']].to_numpy()
            y = data.loc[:, ['Z_s']].to_numpy()

            ax.set_aspect('equal', adjustable='box')
            ax.hexbin(x, y, gridsize=30, cmap='YlOrRd', extent=(-3, 3, -3, 3))
            ax.set(xlim=(-3, 3), ylim=(-3, 3))
            ax.axhline(y=0.5, xmin=0, xmax=0.4167, color='#9e9e9e', linestyle='--')
            ax.axhline(y=0.5, xmin=0.583, xmax=1, color='#9e9e9e', linestyle='--')
            ax.axhline(y=-0.5, xmin=0, xmax=0.4167, color='#9e9e9e', linestyle='--')
            ax.axhline(y=-0.5, xmin=0.583, xmax=1, color='#9e9e9e', linestyle='--')

            ax.axvline(x=-0.5, ymin=0, ymax=0.4167, color='#9e9e9e', linestyle='--')
            ax.axvline(x=-0.5, ymin=0.583, ymax=1, color='#9e9e9e', linestyle='--')
            ax.axvline(x=0.5, ymin=0, ymax=0.4167, color='#9e9e9e', linestyle='--')
            ax.axvline(x=0.5, ymin=0.583, ymax=1, color='#9e9e9e', linestyle='--')

            ax.text(0.20, 0.05, 'Const. Pool', horizontalalignment='center', verticalalignment='center', transform=ax.transAxes)
            ax.text(0.18, 0.95, 'Nozzle', horizontalalignment='center', verticalalignment='center', transform=ax.transAxes)
            ax.text(0.82, 0.95, 'Wide Bar', horizontalalignment='center', verticalalignment='center', transform=ax.transAxes)
            ax.text(0.82, 0.05, 'Oversized', horizontalalignment='center', verticalalignment='center', transform=ax.transAxes)

            ax.set_title(titles[count])
            ax.set_xlabel('Standardized width (Ws)')
            ax.set_ylabel('Standardized detrended elevation (Zs)')

        if class_title == '':
            save_title = out_folder + '\\comid%s_heatplots.png' % comids[0]
        else:
            save_title = out_folder + '\\%s_heatplots.png'

        fig = plt.gcf()
        plt.savefig(save_title, dpi=300, bbox_inches='tight')
        plt.clf()
        plt.close('all')
        print('Plot comparing key Zs %s for comid %s. Located @ %s' % (key_zs, comids[0], save_title))

    else:
        print('Error, please input key Z floats associated with flow stage')


def vector_plotter(base_folder, comids, out_folder, class_title='', geo_classes=[], key_zs=[]):
    titles = {0: 'Base->BF', 1: 'BF->VF', 2: 'Base->VF'}

    if len(comids) >= 1:
        if class_title == '':
            class_title = 'multiple_reaches'
        if len(geo_classes) != len(comids):
            print('Enter a sub folder string associated with a COMID[comid]')
        if len(key_zs) != len(comids):
            print('Enter a sub list of key zs (ex: key_zs=[[0.5, 1.2, 4.9]...]) for each input reach comid')

        x_list_of_arrays = [[], [], []]  # Initialize list containing [baseflow, bankful, valley fill] values
        y_list_of_arrays = [[], [], []]
        c_list_of_arrays = [[], [], []]

        for count, comid in enumerate(comids):
            geo_class = geo_classes[count]
            df = pd.read_csv(base_folder + '\\DATASET_by_reach\\%s\\COMID%s\\Tables\\aligned_locations.csv' % (geo_class, comid))
            data = df.dropna()

            for index, z in enumerate(key_zs[count]):
                z_str = float_keyz_format(z)
                x_temp = data.loc[:, ['W_s_%sft' % z_str]].squeeze().to_list()
                y_temp = data.loc[:, ['Z_s_%sft' % z_str]].squeeze().to_list()
                c_temp = data.loc[:, ['W_s_Z_s_%sft' % z_str]].squeeze().to_list()
                for value in range(len(x_temp)):
                    x_list_of_arrays[index].append(x_temp[value])
                    y_list_of_arrays[index].append(y_temp[value])
                    c_list_of_arrays[index].append(c_temp[value])

        permutations = list(combinations([0, 1, 2], 2))
        permutations[1], permutations[2] = permutations[2], permutations[1]

        x = np.arange(-3, 3, 0.05)  # plots go from -3 to 3
        y = np.arange(-3, 3, 0.05)
        x_axis, y_axis = np.meshgrid(x, y)

        fig, axs = plt.subplots(ncols=int(len(key_zs[0])), figsize=(10, 3))
        fig.subplots_adjust(hspace=0.5, wspace=0.3, left=0.07, right=0.93)

        for count, ax in enumerate(axs):
            perm = np.array(list(permutations[count]))
            max, min = np.amax(perm), np.amin(perm)
            x_velocities = []
            y_velocities = []
            delta_cov = []  # Change in covariance

            for ind in range(0, len(x_list_of_arrays[0])):
                x_velocities.append(float(x_list_of_arrays[max][ind] - x_list_of_arrays[min][ind]))
                y_velocities.append(float(y_list_of_arrays[max][ind] - y_list_of_arrays[min][ind]))
                delta_cov.append(float(c_list_of_arrays[max][ind] - c_list_of_arrays[min][ind]))

            #  Delta_cov values are normalized from -5 to 5
            ax.quiver(x_list_of_arrays[min], y_list_of_arrays[min], x_velocities, y_velocities, delta_cov, norm=colors_module.Normalize(vmin=-5, vmax=5), cmap='seismic', width=0.022, units='xy', scale=10)
            ax.set_title(titles[count])

            ax.set_aspect('equal', adjustable='box')
            ax.set(xlim=(-3, 3), ylim=(-3, 3))
            ax.axhline(y=0.5, xmin=0, xmax=0.4167, color='#9e9e9e', linestyle='--')
            ax.axhline(y=0.5, xmin=0.583, xmax=1, color='#9e9e9e', linestyle='--')
            ax.axhline(y=-0.5, xmin=0, xmax=0.4167, color='#9e9e9e', linestyle='--')
            ax.axhline(y=-0.5, xmin=0.583, xmax=1, color='#9e9e9e', linestyle='--')

            ax.axvline(x=-0.5, ymin=0, ymax=0.4167, color='#9e9e9e', linestyle='--')
            ax.axvline(x=-0.5, ymin=0.583, ymax=1, color='#9e9e9e', linestyle='--')
            ax.axvline(x=0.5, ymin=0, ymax=0.4167, color='#9e9e9e', linestyle='--')
            ax.axvline(x=0.5, ymin=0.583, ymax=1, color='#9e9e9e', linestyle='--')

            ax.text(0.20, 0.05, 'Const. Pool', horizontalalignment='center', verticalalignment='center',
                    transform=ax.transAxes)
            ax.text(0.18, 0.95, 'Nozzle', horizontalalignment='center', verticalalignment='center',
                    transform=ax.transAxes)
            ax.text(0.82, 0.95, 'Wide Bar', horizontalalignment='center', verticalalignment='center',
                    transform=ax.transAxes)
            ax.text(0.82, 0.05, 'Oversized', horizontalalignment='center', verticalalignment='center',
                    transform=ax.transAxes)
            ax.set_xlabel('Standardized width (Ws)')
            ax.set_ylabel('Standardized detrended elevation (Zs)')

        #fig.colorbar(matplotlib.cm.ScalarMappable(norm=colors_module.Normalize(vmin=-5, vmax=5), cmap='seismic'), ticks=range(-5, 6, 1), orientation='horizantal', ax=axs)
        if sc_class == '':
            save_title = out_folder + '\\comid%s_vectorplot.png' % comid
        else:
            save_title = out_folder + '\\class%s_vectorplot.png' % class_title

        fig.set_size_inches(10, 10)
        plt.savefig(save_title, dpi=300, bbox_inches='tight')
        plt.clf()
        plt.close('all')
        print('Vector plot comparing key Zs for all class %s reaches completed. Located @ %s' % (class_title, save_title))



def ww_runs_test(in_folder, out_folder, key_zs=[], fields=['W_s', 'Z_s', 'W_s_Z_s']):
    '''INPUTS: in_folder = Main directory (LINEAR_DETREND folder).
                out_folder is the directory where the xlsx with runs test results is saved
            A list of float or int stage heights for the runs test to be run on.
            A list of fields to do the WW runs test on.
    RETURNS: A xlxs file containing the following for each field (separated in sheets):
            number of runs
            number of expected runs (if random)
            expected standard deviation of number of runs (if random)
            Z: number of standard deviations difference between actual and expected number of run (standard deviation of num. of runs if random)'''

    gcs_folder = in_folder
    xl_loc = out_folder + '\\WW_runs_tests.xlsx'

    base_row = 1
    gap = len(fields) + 2
    wb = xl.Workbook()
    wb.save(xl_loc)
    ws = wb.active
    for z in key_zs:
        print('Runs test with values %s being ran for a %sft stage' % (fields, z))
        ws.cell(row=base_row, column=1).value = '%sft' % z
        ws.cell(row=base_row, column=2).value = 'Field'

        z_str = float_keyz_format(z)

        data_csv = gcs_folder + '\\%sft_WD_analysis_table.csv' % z_str
        data_df = pd.read_csv(data_csv)
        data_df.sort_values(by=['dist_down'], inplace=True)
        spacing = int(data_df.iloc[1]['dist_down'] - data_df.iloc[0]['dist_down'])

        for count, field in enumerate(fields):
            row = base_row + count + 1
            ws.cell(row=row, column=2).value = field
            series = data_df.loc[:, [field]].squeeze()
            out_dict = GCS_analysis.runs_test(series, spacing=int(spacing))
            if count == 0:  # Set heading for a given flow using the output names from the WW runs dict
                for i, key in enumerate(out_dict.keys()):
                    ws.cell(row=base_row, column=(3 + i)).value = str(key)
            for j, key in enumerate(out_dict.values()):  # Inserts corresponding values
                ws.cell(row=row, column=(3 + j)).value = str(key)

        base_row += gap
        wb.save(xl_loc)
        print('%sft stage runs test complete!' % z)

    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['H'].width = 18
    wb.save(xl_loc)
    wb.close()
    print('Wald-Wolfowitz runs test for values below/above median finished for all inputs. Located @ %s' % xl_loc)


###### INPUTS ######
#  SCO3 folder list
#  SCO4 folder list ['\\SCO1', '\\SCO1', '\\SCO2', '\\SCO3', '\\SCO3', '\\SCO4', '\\SCO5', '\\SCO5']
#  SCO5 folder list ['\\SCO1', '\\SCO1', '\\SCO1', '\\SCO3', '\\SCO3', '\\SCO4', '\\SCO4', '\\SCO5', '\\SCO5']

#  SCO1 fake grouping [17569535, 22514218, 17607553, 17609707, 17609017, 17610661]
#  SCO2 fake grouping [17586504, 17610257, 17573013, 17573045, 17586810, 17609015]
#  SCO3 fake grouping [17586610, 17610235, 17595173, 17607455, 17586760]
#  SCO4 fake grouping [17563722, 17594703, 17609699, 17570395, 17585756, 17611423]
#  SCO5 fake grouping [17609755, 17569841, 17563602, 17610541, 17610721, 17610671]
#  SC00_new_adds fake grouping [17567211, 17633478, 17562556, 17609947, 17637906, 17570347]

#  SCO1 class [17573013, 17573045, 17567211, 17633478, 17562556, 17609947]
#  SCO2 class [17609707, 17586810, 17609015, 17586760, 17610671, 17637906]
#  SCO3 class [17569535, 17586504, 17594703, 17609699, 17570395, 17609755, 17570347]
#  SCO4 class [17610541, 22514218, 17610257, 17610235, 17595173, 17563722, 17569841, 17563602]
#  SCO5 class [17607553, 17609017, 17610661, 17586610, 17607455, 17585756, 17611423, 17610721]

comid_list = [17607553, 17609017, 17610661, 17586610, 17607455, 17585756, 17611423, 17610721]
sc_class = '05'
SCO_list = [sc_class for i in comid_list]

key_zs = [[0.2, 1.1, 2.6], [0.5, 4.2, 7.3], [0.5, 2.1, 8.5], [0.5, 1.7, 5.4], [0.3, 1.4, 4.2], [0.8, 2.0, 4.3], [0.8, 1.8, 6.0], [0.4, 1.3, 4.1]]

#  SCO1 fake grouping [[0.9, 3.0, 5.8], [0.1, 0.9, 5.2], [0.2, 1.1, 2.6], [0.5, 2.0, 5.0], [0.5, 4.2, 7.3], [0.5, 2.1, 8.5]]
#  SCO2 fake grouping [[0.7, 2.9, 4.9], [0.4, 2.5, 4.9], [0.2, 2.2, 5.1], [0.6, 3.1, 12.0], [0.6, 3.6, 8.1], [0.3, 3.4, 10.3]]
#  SCO3 fake grouping [[0.5, 1.7, 5.4], [0.4, 1.9, 3.8], [0.0, 1.0, 4.6], [0.3, 1.4, 4.2], [0.7, 2.7, 5.0]]
#  SCO4 fake grouping [[0.7, 1.6, 4.8], [0.5, 2.9, 5.6], [0.5, 2.2, 5.6], [0.2, 1.1, 5.0], [0.8, 2.0, 4.3], [0.8, 1.8, 6.0]]
#  SCO5 fake grouping [[0.2, 1.0, 3.5], [0.3, 1.5, 5.0], [0.6, 1.2, 6.0], [0.5, 2.3, 5.9], [0.4, 1.3, 4.1], [0.4, 2.7, 8.0]]
#  SC00_new_adds fake grouping [[0.1, 0.9, 2.6], [0.1, 1.0, 3.1], [0.3, 3.0], [0.2, 0.7, 2.6], [0.3, 1.2, 5.3], [0.6, 3.2, 6.0]]

# SCO1 class [[0.2, 2.2, 5.1], [0.6, 3.1, 12.0], [0.1, 0.9, 2.6], [0.1, 1.0, 3.1], [0.3, 3.0], [0.2, 0.7, 2.6]]
# SCO2 class [[0.5, 2.0, 5.0], [0.6, 3.6, 8.1], [0.3, 3.4, 10.3], [0.7, 2.7, 5.0], [0.4, 2.7, 8.0], [0.3, 1.2, 5.3]]
# SCO3 class [[0.9, 3.0, 5.8], [0.7, 2.9, 4.9], [0.5, 2.9, 5.6], [0.5, 2.2, 5.6], [0.2, 1.1, 5.0], [0.2, 1.0, 3.5], [0.6, 3.2, 6.0]]
# SCO4 class [[0.5, 2.3, 5.9], [0.1, 0.9, 5.2], [0.4, 2.5, 4.9], [0.4, 1.9, 3.8], [0.0, 1.0, 4.6], [0.7, 1.6, 4.8], [0.3, 1.5, 5.0], [0.6, 1.2, 6.0]]
# SCO5 class [[0.2, 1.1, 2.6], [0.5, 4.2, 7.3], [0.5, 2.1, 8.5], [0.5, 1.7, 5.4], [0.3, 1.4, 4.2], [0.8, 2.0, 4.3], [0.8, 1.8, 6.0], [0.4, 1.3, 4.1]]


bf_zs = []
roots = ['cwz_above_0', 'ws_above_0p5', 'zs_above_0p5', 'cwz_above_0p5', 'ws_above_1', 'zs_above_1', 'cwz_above_1',
                'ws_below_0p5', 'zs_below_0p5', 'cwz_below_0p5', 'ws_below_1', 'zs_below_1', 'cwz_below_1', 'abs_ws_above_0p5',
                'abs_zs_above_0p5', 'abs_cwz_above_0p5', 'abs_ws_above_1', 'abs_zs_above_1', 'abs_cwz_above_1']
y_labels = ['base', 'bf', 'vf']
full_x_labels = []

for label in y_labels:
    for root in roots:
        full_x_labels.append('%s_%s' % (label, root))

single_plot_lists = []
for root in roots:
    temp_list = []
    for label in y_labels:
        temp_list.append('%s_%s' % (label, root))
    single_plot_lists.append(temp_list)

analysis_plotting = False

if analysis_plotting == True:
    arcpy.env.overwriteOutput = True

    for count, comid in enumerate(comid_list):
        SCO_number = SCO_list[count]
        base = r'Z:\users\xavierrn\SoCoast_Final_ResearchFiles'
        sc_folder = r"Z:\users\xavierrn\SoCoast_Final_ResearchFiles\SC%s" % SCO_list[count]
        direct = (r"Z:\users\xavierrn\SoCoast_Final_ResearchFiles\SC%s\COMID%s" % (SCO_number, comid))
        out_folder = direct + r'\LINEAR_DETREND'
        process_footprint = direct + '\\las_footprint.shp'
        table_location = out_folder + "\\gcs_ready_tables"
        channel_clip_poly = out_folder + '\\raster_clip_poly.shp'
        aligned_csv_loc = out_folder + '\\landform_analysis\\aligned_locations.csv'
        landform_folder = out_folder + '\\landform_analysis'
        confine_table = r'Z:\users\xavierrn\Manual classification files\South_200m.shp'
        wetted_top_folder = out_folder + '\\wetted_polygons'
        key_z_dict = {}

    base = r'Z:\rivers\eFlows\6_South_Coast_Ephemeral_Data'
    sample_table = r'Z:\rivers\eFlows\6_South_Coast_Ephemeral_Data\classified_sampled_reaches.csv'
    out = base + '\\Boxplots\\stage_comparison_box_plots_class_sorted\\scaled_0_to_100'

    more_roots = ['mean_cwz', 'mean_cwz', 'mean_cwz']
    items = ['base_mean_cwz', 'bf_mean_cwz', 'vf_mean_cwz']

    box_plots(in_csv=sample_table, out_folder=out, fields=items, field_units=['Mean C(Ws,Zs)' for i in items], ylim=1, field_title='mean_cwz_no_cap', sort_by_field='manual_class', sort_by_title='Channel type class', single_plots=True)
    #heat_plotter(base_folder=base, comids=comid_list, out_folder=out, class_title='SC5', geo_classes=folder_list, key_zs=key_zs)
    #landform_pie_charts(base_folder=base, comids=comid_list, out_folder=out, class_title='SC1', geo_classes=folder_list, all_key_zs=key_zs)
    #vector_plotter(base_folder=base, comids=comid_list, out_folder=out, class_title=class_plot_title, geo_classes=folder_list, key_zs=key_zs)
    #nested_landform_sankey(base_folder=base, comids=comid_list, out_folder=out, class_title=class_plot_title, geo_classes=folder_list, all_key_zs=key_zs, ignore_normal=False)
    #nested_landform_sankey(base_folder=base, comids=comid_list, out_folder=out, class_title=class_plot_title, geo_classes=folder_list, all_key_zs=key_zs, ignore_normal=True)

